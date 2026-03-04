"""
Gestão Processual Routes - Planejamento Acaiaca
Handles CRUD operations for Process Management
"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from io import BytesIO
import uuid
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER

router = APIRouter(prefix="/api", tags=["Gestão Processual"])

# Process modalities
MODALIDADES_CONTRATACAO = [
    "Pregão Eletrônico",
    "Pregão Presencial",
    "Concorrência",
    "Tomada de Preços",
    "Convite",
    "Dispensa de Licitação",
    "Inexigibilidade",
    "Leilão",
    "Concurso",
    "Diálogo Competitivo",
    "Adesão a Ata de Registro de Preços",
    "Chamamento Público"
]

# Process statuses
STATUS_PROCESSO = [
    "Em Elaboração",
    "Aguardando Aprovação",
    "Aprovado",
    "Em Andamento",
    "Publicado",
    "Em Fase de Lances",
    "Em Análise de Propostas",
    "Homologado",
    "Adjudicado",
    "Contratado",
    "Em Execução",
    "Concluído",
    "Cancelado",
    "Suspenso",
    "Deserto",
    "Fracassado"
]


def setup_processos_routes(db, get_current_user, Processo, ProcessoCreate, ProcessoUpdate):
    """Configure Processos routes with injected dependencies"""

    # ===== YEARS =====
    @router.get("/processos/anos")
    async def get_anos_processos(request: Request):
        """Returns available years in processes"""
        user = await get_current_user(request)
        pipeline = [
            {"$group": {"_id": "$ano"}},
            {"$sort": {"_id": -1}}
        ]
        anos = await db.processos.aggregate(pipeline).to_list(100)
        anos_list = [a['_id'] for a in anos if a['_id']]
        if not anos_list:
            anos_list = [datetime.now().year]
        return anos_list

    # ===== MODALITIES AND STATUS =====
    @router.get("/processos/modalidades/lista")
    async def get_modalidades(request: Request):
        """List all contracting modalities"""
        user = await get_current_user(request)
        return MODALIDADES_CONTRATACAO

    @router.get("/processos/status/lista")
    async def get_status_lista(request: Request):
        """List all possible statuses"""
        user = await get_current_user(request)
        return STATUS_PROCESSO

    # ===== CRUD =====
    @router.get("/processos", response_model=List[Processo])
    async def get_processos(
        request: Request,
        ano: int = None,
        status: str = None,
        modalidade: str = None,
        secretaria: str = None
    ):
        """List all processes"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        if status:
            query['status'] = status
        if modalidade:
            query['modalidade_contratacao'] = modalidade
        if secretaria:
            query['secretaria'] = {'$regex': secretaria, '$options': 'i'}
        
        processos = await db.processos.find(query, {'_id': 0}).sort('created_at', -1).to_list(10000)
        return [Processo(**p) for p in processos]

    @router.get("/processos/paginado")
    async def get_processos_paginado(
        request: Request,
        ano: int = None,
        status: str = None,
        modalidade: str = None,
        secretaria: str = None,
        busca: str = None,
        page: int = 1,
        limit: int = 20
    ):
        """List processes with pagination"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        if status:
            query['status'] = status
        if modalidade:
            query['modalidade_contratacao'] = modalidade
        if secretaria:
            query['secretaria'] = {'$regex': secretaria, '$options': 'i'}
        if busca:
            query['$or'] = [
                {'numero_processo': {'$regex': busca, '$options': 'i'}},
                {'objeto': {'$regex': busca, '$options': 'i'}},
                {'responsavel': {'$regex': busca, '$options': 'i'}}
            ]
        
        skip = (page - 1) * limit
        total = await db.processos.count_documents(query)
        processos = await db.processos.find(query, {'_id': 0}).sort('created_at', -1).skip(skip).limit(limit).to_list(limit)
        
        return {
            'data': processos,
            'total': total,
            'page': page,
            'limit': limit,
            'pages': (total + limit - 1) // limit
        }

    @router.get("/processos/stats")
    async def get_processos_stats(request: Request, ano: int = None):
        """Returns process statistics"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        
        processos = await db.processos.find(query, {'_id': 0}).to_list(10000)
        
        stats = {
            'total': len(processos),
            'por_status': {},
            'por_modalidade': {},
            'por_secretaria': {},
            'por_mes': {}
        }
        
        for proc in processos:
            status = proc.get('status', 'Não definido')
            modalidade = proc.get('modalidade_contratacao', 'Não definida')
            secretaria = proc.get('secretaria', 'Não definida')
            
            stats['por_status'][status] = stats['por_status'].get(status, 0) + 1
            stats['por_modalidade'][modalidade] = stats['por_modalidade'].get(modalidade, 0) + 1
            stats['por_secretaria'][secretaria] = stats['por_secretaria'].get(secretaria, 0) + 1
            
            # By month
            created = proc.get('created_at')
            if created:
                if isinstance(created, datetime):
                    mes = created.strftime('%Y-%m')
                else:
                    mes = str(created)[:7]
                stats['por_mes'][mes] = stats['por_mes'].get(mes, 0) + 1
        
        return stats

    @router.post("/processos", response_model=Processo)
    async def create_processo(processo: ProcessoCreate, request: Request):
        """Creates a new process"""
        user = await get_current_user(request)
        proc_dict = processo.model_dump()
        proc_dict['processo_id'] = f"proc_{uuid.uuid4().hex[:12]}"
        proc_dict['user_id'] = user.user_id
        
        # Extract year from process number if not provided
        if not proc_dict.get('ano') and proc_dict.get('numero_processo'):
            match = re.search(r'/(\d{4})$', proc_dict['numero_processo'])
            if match:
                proc_dict['ano'] = int(match.group(1))
            else:
                proc_dict['ano'] = datetime.now().year
        
        proc_dict['created_at'] = datetime.now(timezone.utc)
        proc_dict['updated_at'] = datetime.now(timezone.utc)
        
        await db.processos.insert_one(proc_dict)
        proc_dict.pop('_id', None)
        return Processo(**proc_dict)

    @router.get("/processos/{processo_id}", response_model=Processo)
    async def get_processo(processo_id: str, request: Request):
        """Gets a specific process"""
        user = await get_current_user(request)
        processo = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0})
        if not processo:
            raise HTTPException(status_code=404, detail="Processo não encontrado")
        return Processo(**processo)

    @router.put("/processos/{processo_id}", response_model=Processo)
    async def update_processo(processo_id: str, processo: ProcessoUpdate, request: Request):
        """Updates a process"""
        user = await get_current_user(request)
        existing = await db.processos.find_one({'processo_id': processo_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Processo não encontrado")
        
        update_data = {k: v for k, v in processo.model_dump().items() if v is not None}
        update_data['updated_at'] = datetime.now(timezone.utc)
        
        await db.processos.update_one({'processo_id': processo_id}, {'$set': update_data})
        updated = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0})
        return Processo(**updated)

    @router.delete("/processos/{processo_id}")
    async def delete_processo(processo_id: str, request: Request):
        """Deletes a process"""
        user = await get_current_user(request)
        result = await db.processos.delete_one({'processo_id': processo_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Processo não encontrado")
        return {'message': 'Processo excluído com sucesso'}

    # ===== MIGRATION =====
    @router.post("/processos/migrate-fields")
    async def migrate_processos_fields(request: Request):
        """Migrates legacy fields to new format"""
        user = await get_current_user(request)
        if not user.is_admin:
            raise HTTPException(status_code=403, detail="Admin access required")
        
        processos = await db.processos.find({}, {'_id': 0}).to_list(10000)
        migrated = 0
        
        for proc in processos:
            updates = {}
            
            # Migrate modalidade -> modalidade_contratacao
            if proc.get('modalidade') and not proc.get('modalidade_contratacao'):
                updates['modalidade_contratacao'] = proc['modalidade']
            
            # Migrate situacao -> status
            if proc.get('situacao') and not proc.get('status'):
                updates['status'] = proc['situacao']
            
            if updates:
                await db.processos.update_one(
                    {'processo_id': proc['processo_id']},
                    {'$set': updates}
                )
                migrated += 1
        
        return {'message': f'Migrados {migrated} processos'}

    # ===== IMPORT =====
    @router.post("/processos/import")
    async def import_processos(request: Request, file: UploadFile = File(...)):
        """Imports processes from XLSX file"""
        user = await get_current_user(request)
        
        try:
            contents = await file.read()
            wb = load_workbook(filename=BytesIO(contents))
            ws = wb.active
            
            imported = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                try:
                    # Extract year from process number
                    numero = str(row[0])
                    ano = datetime.now().year
                    match = re.search(r'/(\d{4})$', numero)
                    if match:
                        ano = int(match.group(1))
                    
                    proc_dict = {
                        'processo_id': f"proc_{uuid.uuid4().hex[:12]}",
                        'user_id': user.user_id,
                        'numero_processo': numero,
                        'modalidade_contratacao': str(row[1]) if row[1] else 'Pregão Eletrônico',
                        'status': str(row[2]) if row[2] else 'Em Elaboração',
                        'objeto': str(row[3]) if row[3] else '',
                        'responsavel': str(row[4]) if row[4] else '',
                        'secretaria': str(row[5]) if row[5] else '',
                        'secretario': str(row[6]) if row[6] else '',
                        'ano': ano,
                        'observacoes': str(row[7]) if len(row) > 7 and row[7] else '',
                        'created_at': datetime.now(timezone.utc),
                        'updated_at': datetime.now(timezone.utc)
                    }
                    await db.processos.insert_one(proc_dict)
                    imported += 1
                except Exception as e:
                    errors.append(f"Linha: {str(e)}")
            
            return {'success': True, 'imported': imported, 'errors': errors}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

    # ===== EXPORT PDF =====
    @router.get("/processos/export/pdf")
    async def export_processos_pdf(request: Request, ano: int = None, status: str = None):
        """Exports processes to PDF"""
        user = await get_current_user(request)
        
        query = {}
        if ano:
            query['ano'] = ano
        if status:
            query['status'] = status
        
        processos = await db.processos.find(query, {'_id': 0}).sort('numero_processo', 1).to_list(10000)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=15*mm, bottomMargin=10*mm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=12, textColor=colors.HexColor('#1F4E78'), alignment=TA_CENTER, spaceAfter=8)
        
        titulo = f"RELATÓRIO DE PROCESSOS - {'ANO ' + str(ano) if ano else 'TODOS OS ANOS'}"
        elements.append(Paragraph(titulo, title_style))
        elements.append(Spacer(1, 5))
        
        # Table
        table_data = [['#', 'Número', 'Modalidade', 'Status', 'Objeto', 'Secretaria', 'Responsável']]
        
        for i, proc in enumerate(processos, 1):
            table_data.append([
                str(i),
                proc.get('numero_processo', '')[:15],
                proc.get('modalidade_contratacao', '')[:20],
                proc.get('status', '')[:15],
                proc.get('objeto', '')[:40],
                proc.get('secretaria', '')[:20],
                proc.get('responsavel', '')[:20]
            ])
        
        table = Table(table_data, colWidths=[20, 55, 70, 55, 150, 75, 75])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Total de processos: {len(processos)}", ParagraphStyle('Footer', fontSize=9, alignment=TA_CENTER)))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", ParagraphStyle('Footer2', fontSize=8, alignment=TA_CENTER, textColor=colors.gray)))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"Processos_{ano or 'todos'}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # ===== EXPORT XLSX =====
    @router.get("/processos/export/xlsx")
    async def export_processos_xlsx(request: Request, ano: int = None, status: str = None):
        """Exports processes to XLSX"""
        user = await get_current_user(request)
        
        query = {}
        if ano:
            query['ano'] = ano
        if status:
            query['status'] = status
        
        processos = await db.processos.find(query, {'_id': 0}).sort('numero_processo', 1).to_list(10000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Processos"
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RELATÓRIO DE PROCESSOS - {ano or 'TODOS OS ANOS'}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['#', 'Número', 'Modalidade', 'Status', 'Objeto', 'Secretaria', 'Responsável', 'Observações']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # Data
        for row, proc in enumerate(processos, 4):
            ws.cell(row=row, column=1, value=row-3).border = thin_border
            ws.cell(row=row, column=2, value=proc.get('numero_processo', '')).border = thin_border
            ws.cell(row=row, column=3, value=proc.get('modalidade_contratacao', '')).border = thin_border
            ws.cell(row=row, column=4, value=proc.get('status', '')).border = thin_border
            ws.cell(row=row, column=5, value=proc.get('objeto', '')).border = thin_border
            ws.cell(row=row, column=6, value=proc.get('secretaria', '')).border = thin_border
            ws.cell(row=row, column=7, value=proc.get('responsavel', '')).border = thin_border
            ws.cell(row=row, column=8, value=proc.get('observacoes', '')).border = thin_border
        
        # Adjust widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 30
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Processos_{ano or 'todos'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    return router
