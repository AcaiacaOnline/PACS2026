"""
PAC Individual Routes - Planejamento Acaiaca
Handles CRUD operations for Individual PACs (Planos Anuais de Contratação)
"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from io import BytesIO
import uuid
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT

router = APIRouter(prefix="/api", tags=["PAC Individual"])


def setup_pac_routes(db, get_current_user, PAC, PACItem, PACCreate, PACUpdate, PACItemCreate, PACItemUpdate):
    """Configure PAC Individual routes with injected dependencies"""

    # ===== PAC YEARS =====
    @router.get("/pacs/anos")
    async def get_anos_pacs(request: Request):
        """Returns available years in PACs"""
        user = await get_current_user(request)
        pipeline = [
            {"$group": {"_id": "$ano"}},
            {"$sort": {"_id": -1}}
        ]
        anos = await db.pacs.aggregate(pipeline).to_list(100)
        anos_list = [a['_id'] for a in anos if a['_id']]
        if not anos_list:
            anos_list = [str(datetime.now().year)]
        return anos_list

    # ===== PAC CRUD =====
    @router.get("/pacs", response_model=List[PAC])
    async def get_pacs(request: Request, ano: str = None, secretaria: str = None):
        """List all user PACs"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        if secretaria:
            query['secretaria'] = {'$regex': secretaria, '$options': 'i'}
        pacs = await db.pacs.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
        return pacs

    @router.get("/pacs/paginado")
    async def get_pacs_paginado(
        request: Request,
        ano: str = None,
        secretaria: str = None,
        page: int = 1,
        limit: int = 20
    ):
        """List PACs with pagination"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        if secretaria:
            query['secretaria'] = {'$regex': secretaria, '$options': 'i'}
        
        skip = (page - 1) * limit
        total = await db.pacs.count_documents(query)
        pacs = await db.pacs.find(query, {'_id': 0}).sort('created_at', -1).skip(skip).limit(limit).to_list(limit)
        
        return {
            'data': pacs,
            'total': total,
            'page': page,
            'limit': limit,
            'pages': (total + limit - 1) // limit
        }

    @router.get("/pacs/stats")
    async def get_pacs_stats(request: Request, ano: str = None):
        """Returns PAC statistics"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        
        pacs = await db.pacs.find(query, {'_id': 0}).to_list(1000)
        
        stats = {
            'total_pacs': len(pacs),
            'por_secretaria': {},
            'valor_total': 0,
            'total_items': 0
        }
        
        for pac in pacs:
            secretaria = pac.get('secretaria', 'Não definida')
            if secretaria not in stats['por_secretaria']:
                stats['por_secretaria'][secretaria] = {
                    'quantidade': 0,
                    'valor_total': 0,
                    'quantidade_items': 0
                }
            
            stats['por_secretaria'][secretaria]['quantidade'] += 1
            
            items = await db.pac_items.find({'pac_id': pac['pac_id']}, {'_id': 0}).to_list(10000)
            valor_pac = sum(item.get('valorTotal', 0) for item in items)
            
            stats['por_secretaria'][secretaria]['valor_total'] += valor_pac
            stats['por_secretaria'][secretaria]['quantidade_items'] += len(items)
            stats['valor_total'] += valor_pac
            stats['total_items'] += len(items)
        
        return stats

    @router.post("/pacs", response_model=PAC)
    async def create_pac(pac: PACCreate, request: Request):
        """Creates a new PAC"""
        user = await get_current_user(request)
        pac_dict = pac.model_dump()
        pac_dict['pac_id'] = f"pac_{uuid.uuid4().hex[:12]}"
        pac_dict['user_id'] = user.user_id
        pac_dict['total_value'] = 0.0
        pac_dict['stats'] = None
        pac_dict['created_at'] = datetime.now(timezone.utc)
        pac_dict['updated_at'] = datetime.now(timezone.utc)
        await db.pacs.insert_one(pac_dict)
        pac_dict.pop('_id', None)
        return PAC(**pac_dict)

    @router.get("/pacs/{pac_id}", response_model=PAC)
    async def get_pac(pac_id: str, request: Request):
        """Gets a specific PAC"""
        user = await get_current_user(request)
        pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        return PAC(**pac)

    @router.put("/pacs/{pac_id}", response_model=PAC)
    async def update_pac(pac_id: str, pac: PACUpdate, request: Request):
        """Updates a PAC"""
        user = await get_current_user(request)
        existing = await db.pacs.find_one({'pac_id': pac_id})
        if not existing:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        
        update_data = {k: v for k, v in pac.model_dump().items() if v is not None}
        update_data['updated_at'] = datetime.now(timezone.utc)
        
        await db.pacs.update_one({'pac_id': pac_id}, {'$set': update_data})
        updated = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
        return PAC(**updated)

    @router.delete("/pacs/{pac_id}")
    async def delete_pac(pac_id: str, request: Request):
        """Deletes a PAC and its items"""
        user = await get_current_user(request)
        result = await db.pacs.delete_one({'pac_id': pac_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        await db.pac_items.delete_many({'pac_id': pac_id})
        return {'message': 'PAC excluído com sucesso'}

    # ===== PAC ITEMS =====
    @router.get("/pacs/{pac_id}/items", response_model=List[PACItem])
    async def get_pac_items(pac_id: str, request: Request):
        """Lists items of a PAC"""
        user = await get_current_user(request)
        items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(10000)
        return items

    @router.post("/pacs/{pac_id}/items", response_model=PACItem)
    async def create_pac_item(pac_id: str, item: PACItemCreate, request: Request):
        """Adds an item to a PAC"""
        user = await get_current_user(request)
        pac = await db.pacs.find_one({'pac_id': pac_id})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        
        item_dict = item.model_dump()
        item_dict['item_id'] = f"item_{uuid.uuid4().hex[:12]}"
        item_dict['pac_id'] = pac_id
        item_dict['valorTotal'] = round(item_dict['quantidade'] * item_dict['valorUnitario'], 2)
        item_dict['created_at'] = datetime.now(timezone.utc)
        
        await db.pac_items.insert_one(item_dict)
        item_dict.pop('_id', None)
        return PACItem(**item_dict)

    @router.put("/pacs/{pac_id}/items/{item_id}", response_model=PACItem)
    async def update_pac_item(pac_id: str, item_id: str, item: PACItemUpdate, request: Request):
        """Updates a PAC item"""
        user = await get_current_user(request)
        existing = await db.pac_items.find_one({'item_id': item_id, 'pac_id': pac_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        
        update_data = {k: v for k, v in item.model_dump().items() if v is not None}
        
        # Recalculate total if quantity or unit value changed
        quantidade = update_data.get('quantidade', existing.get('quantidade', 0))
        valor_unitario = update_data.get('valorUnitario', existing.get('valorUnitario', 0))
        update_data['valorTotal'] = round(quantidade * valor_unitario, 2)
        
        await db.pac_items.update_one({'item_id': item_id}, {'$set': update_data})
        updated = await db.pac_items.find_one({'item_id': item_id}, {'_id': 0})
        return PACItem(**updated)

    @router.delete("/pacs/{pac_id}/items/{item_id}")
    async def delete_pac_item(pac_id: str, item_id: str, request: Request):
        """Deletes a PAC item"""
        user = await get_current_user(request)
        result = await db.pac_items.delete_one({'item_id': item_id, 'pac_id': pac_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        return {'message': 'Item excluído com sucesso'}

    # ===== DASHBOARD STATS =====
    @router.get("/dashboard/stats")
    async def get_dashboard_stats(request: Request):
        """Returns dashboard statistics"""
        user = await get_current_user(request)
        
        pacs_count = await db.pacs.count_documents({})
        pacs_geral_count = await db.pacs_geral.count_documents({})
        pacs_obras_count = await db.pacs_geral_obras.count_documents({})
        processos_count = await db.processos.count_documents({})
        
        # Calculate total values
        pac_items = await db.pac_items.find({}, {'valorTotal': 1, '_id': 0}).to_list(100000)
        total_pac = sum(i.get('valorTotal', 0) for i in pac_items)
        
        pac_geral_items = await db.pac_geral_items.find({}, {'valorTotal': 1, '_id': 0}).to_list(100000)
        total_pac_geral = sum(i.get('valorTotal', 0) for i in pac_geral_items)
        
        pac_obras_items = await db.pac_obras_items.find({}, {'valorTotal': 1, '_id': 0}).to_list(100000)
        total_pac_obras = sum(i.get('valorTotal', 0) for i in pac_obras_items)
        
        return {
            'pacs': {
                'total': pacs_count,
                'valor_total': total_pac
            },
            'pacs_geral': {
                'total': pacs_geral_count,
                'valor_total': total_pac_geral
            },
            'pacs_obras': {
                'total': pacs_obras_count,
                'valor_total': total_pac_obras
            },
            'processos': {
                'total': processos_count
            },
            'valor_total_geral': total_pac + total_pac_geral + total_pac_obras
        }

    # ===== EXPORT XLSX =====
    @router.get("/pacs/{pac_id}/export/xlsx")
    async def export_pac_xlsx(pac_id: str, request: Request):
        """Exports PAC to XLSX"""
        user = await get_current_user(request)
        
        pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        
        items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(10000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "PAC Individual"
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = f"PAC INDIVIDUAL - {pac.get('secretaria', 'N/A')} - {pac.get('ano', '2026')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # PAC info
        ws['A3'] = 'Secretário:'
        ws['B3'] = pac.get('secretario', '')
        ws['A4'] = 'Fiscal:'
        ws['B4'] = pac.get('fiscal', '')
        ws['A5'] = 'Email:'
        ws['B5'] = pac.get('email', '')
        ws['A6'] = 'Telefone:'
        ws['B6'] = pac.get('telefone', '')
        
        # Items header
        headers = ['#', 'Tipo', 'CATMAT', 'Descrição', 'Unidade', 'Qtd', 'Valor Unit.', 'Valor Total', 'Prioridade', 'Justificativa']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Items data
        total = 0
        for row, item in enumerate(items, 9):
            ws.cell(row=row, column=1, value=row-8).border = thin_border
            ws.cell(row=row, column=2, value=item.get('tipo', '')).border = thin_border
            ws.cell(row=row, column=3, value=item.get('catmat', '')).border = thin_border
            ws.cell(row=row, column=4, value=item.get('descricao', '')).border = thin_border
            ws.cell(row=row, column=5, value=item.get('unidade', '')).border = thin_border
            ws.cell(row=row, column=6, value=item.get('quantidade', 0)).border = thin_border
            ws.cell(row=row, column=7, value=item.get('valorUnitario', 0)).border = thin_border
            ws.cell(row=row, column=8, value=item.get('valorTotal', 0)).border = thin_border
            ws.cell(row=row, column=9, value=item.get('prioridade', '')).border = thin_border
            ws.cell(row=row, column=10, value=item.get('justificativa', '')).border = thin_border
            total += item.get('valorTotal', 0)
        
        # Total row
        total_row = 9 + len(items)
        ws.cell(row=total_row, column=7, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total).font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 30
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"PAC_{pac.get('secretaria', 'individual')}_{pac.get('ano', '2026')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # ===== EXPORT PDF =====
    @router.get("/pacs/{pac_id}/export/pdf")
    async def export_pac_pdf(pac_id: str, request: Request):
        """Exports PAC to PDF"""
        user = await get_current_user(request)
        
        pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        
        items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(10000)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=20*mm, bottomMargin=15*mm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#1F4E78'), alignment=TA_CENTER, spaceAfter=10)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=5)
        
        # Header
        elements.append(Paragraph(f"PAC INDIVIDUAL - {pac.get('secretaria', 'N/A')}", title_style))
        elements.append(Paragraph(f"Ano: {pac.get('ano', '2026')} | Secretário: {pac.get('secretario', 'N/A')}", subtitle_style))
        elements.append(Spacer(1, 10))
        
        # Items table
        table_data = [['#', 'Tipo', 'CATMAT', 'Descrição', 'Unid.', 'Qtd', 'V.Unit', 'V.Total', 'Prior.']]
        
        total = 0
        for i, item in enumerate(items, 1):
            valor_total = item.get('valorTotal', 0)
            total += valor_total
            table_data.append([
                str(i),
                item.get('tipo', '')[:10],
                item.get('catmat', '')[:10],
                item.get('descricao', '')[:50],
                item.get('unidade', '')[:5],
                str(item.get('quantidade', 0)),
                f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                item.get('prioridade', '')[:5]
            ])
        
        # Total row
        table_data.append(['', '', '', '', '', '', 'TOTAL:', f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), ''])
        
        table = Table(table_data, colWidths=[20, 40, 45, 150, 30, 30, 55, 65, 35])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"PAC_{pac.get('secretaria', 'individual')}_{pac.get('ano', '2026')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # ===== TEMPLATE DOWNLOAD =====
    @router.get("/template/download")
    async def download_template(request: Request):
        """Downloads PAC import template"""
        user = await get_current_user(request)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Template PAC"
        
        headers = ['Tipo', 'CATMAT', 'Descrição', 'Unidade', 'Quantidade', 'Valor Unitário', 'Prioridade', 'Justificativa']
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Example row
        ws.cell(row=2, column=1, value='Material')
        ws.cell(row=2, column=2, value='123456')
        ws.cell(row=2, column=3, value='Descrição do item')
        ws.cell(row=2, column=4, value='UN')
        ws.cell(row=2, column=5, value=10)
        ws.cell(row=2, column=6, value=100.00)
        ws.cell(row=2, column=7, value='Alta')
        ws.cell(row=2, column=8, value='Justificativa da aquisição')
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template_pac.xlsx"}
        )

    # ===== IMPORT FROM XLSX =====
    @router.post("/pacs/{pac_id}/import")
    async def import_pac_items(pac_id: str, request: Request, file: UploadFile = File(...)):
        """Imports items from XLSX file"""
        user = await get_current_user(request)
        
        pac = await db.pacs.find_one({'pac_id': pac_id})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC não encontrado")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")
        
        try:
            contents = await file.read()
            wb = load_workbook(filename=BytesIO(contents))
            ws = wb.active
            
            imported = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                try:
                    item_dict = {
                        'item_id': f"item_{uuid.uuid4().hex[:12]}",
                        'pac_id': pac_id,
                        'tipo': str(row[0]) if row[0] else 'Material',
                        'catmat': str(row[1]) if row[1] else '',
                        'descricao': str(row[2]) if row[2] else '',
                        'unidade': str(row[3]) if row[3] else 'UN',
                        'quantidade': float(row[4]) if row[4] else 0,
                        'valorUnitario': float(row[5]) if row[5] else 0,
                        'valorTotal': round((float(row[4]) if row[4] else 0) * (float(row[5]) if row[5] else 0), 2),
                        'prioridade': str(row[6]) if row[6] else 'Média',
                        'justificativa': str(row[7]) if row[7] else '',
                        'created_at': datetime.now(timezone.utc)
                    }
                    await db.pac_items.insert_one(item_dict)
                    imported += 1
                except Exception as e:
                    errors.append(f"Linha {row}: {str(e)}")
            
            return {
                'success': True,
                'imported': imported,
                'errors': errors
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

    return router
