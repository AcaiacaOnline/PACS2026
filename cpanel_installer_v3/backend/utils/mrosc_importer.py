"""
MROSC Excel Importer - Importador de Planilhas Excel para Prestação de Contas
Planejamento Acaiaca - Sistema de Gestão Municipal
"""
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Dict, Any, Tuple
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger("planejamento_acaiaca.mrosc_importer")


class MROSCExcelImporter:
    """
    Importador de planilhas Excel para o módulo MROSC.
    Suporta importação de:
    - Recursos Humanos
    - Despesas/Materiais
    - Cronograma de Desembolso
    """
    
    # Mapeamento de colunas esperadas para RH
    RH_COLUMNS = {
        'nome_funcao': ['função', 'funcao', 'cargo', 'nome_funcao', 'nome da função'],
        'regime_contratacao': ['regime', 'contratação', 'contratacao', 'regime_contratacao', 'tipo contrato'],
        'carga_horaria': ['carga horária', 'carga_horaria', 'ch', 'horas semanais'],
        'salario_bruto': ['salário', 'salario', 'salário bruto', 'salario_bruto', 'remuneração'],
        'encargos': ['encargos', 'provisões', 'provisoes', 'encargos mensais'],
        'meses': ['meses', 'quantidade meses', 'qtd meses', 'numero_meses'],
        'observacoes': ['observações', 'observacoes', 'obs', 'notas'],
    }
    
    # Mapeamento de colunas esperadas para Despesas
    DESPESA_COLUMNS = {
        'natureza_despesa': ['natureza', 'natureza_despesa', 'código natureza', 'codigo natureza'],
        'item': ['item', 'item_despesa', 'descrição item'],
        'descricao': ['descrição', 'descricao', 'especificação', 'especificacao'],
        'unidade': ['unidade', 'un', 'unid', 'unidade medida'],
        'quantidade': ['quantidade', 'qtd', 'qtde', 'quant'],
        'valor_unitario': ['valor unitário', 'valor_unitario', 'preço unitário', 'preco_unitario'],
        'orcamento_1': ['orçamento 1', 'orcamento_1', 'cotação 1', 'cotacao_1'],
        'orcamento_2': ['orçamento 2', 'orcamento_2', 'cotação 2', 'cotacao_2'],
        'orcamento_3': ['orçamento 3', 'orcamento_3', 'cotação 3', 'cotacao_3'],
        'justificativa': ['justificativa', 'just', 'motivo'],
    }
    
    def __init__(self):
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, Any]] = []
        self.imported_rh: List[Dict] = []
        self.imported_despesas: List[Dict] = []
    
    def _normalize_header(self, header: str) -> str:
        """Normaliza o nome da coluna para comparação"""
        if header is None:
            return ""
        return str(header).lower().strip().replace("_", " ")
    
    def _find_column_index(self, headers: List[str], column_mappings: List[str]) -> int:
        """Encontra o índice da coluna baseado nos mapeamentos possíveis"""
        normalized_headers = [self._normalize_header(h) for h in headers]
        for mapping in column_mappings:
            normalized_mapping = mapping.lower().strip()
            for idx, header in enumerate(normalized_headers):
                if normalized_mapping in header or header in normalized_mapping:
                    return idx
        return -1
    
    def _parse_number(self, value: Any, default: float = 0.0) -> float:
        """Converte valor para número"""
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        try:
            # Remove formatação de moeda
            cleaned = str(value).replace("R$", "").replace(".", "").replace(",", ".").strip()
            return float(cleaned) if cleaned else default
        except (ValueError, AttributeError):
            return default
    
    def _parse_int(self, value: Any, default: int = 0) -> int:
        """Converte valor para inteiro"""
        if value is None:
            return default
        if isinstance(value, int):
            return value
        try:
            return int(float(str(value).strip()))
        except (ValueError, AttributeError):
            return default
    
    def _validate_cnpj(self, cnpj: str) -> bool:
        """Valida formato básico de CNPJ"""
        if not cnpj:
            return False
        cnpj_clean = ''.join(filter(str.isdigit, str(cnpj)))
        return len(cnpj_clean) == 14
    
    def _validate_cpf(self, cpf: str) -> bool:
        """Valida formato básico de CPF"""
        if not cpf:
            return False
        cpf_clean = ''.join(filter(str.isdigit, str(cpf)))
        return len(cpf_clean) == 11
    
    def parse_rh_sheet(self, workbook, sheet_name: str = None) -> List[Dict]:
        """
        Parseia a aba de Recursos Humanos
        
        Args:
            workbook: Workbook do openpyxl
            sheet_name: Nome da aba (opcional, tenta detectar automaticamente)
        
        Returns:
            Lista de dicionários com os dados de RH
        """
        # Tentar encontrar a aba de RH
        rh_sheet = None
        rh_sheet_names = ['rh', 'recursos humanos', 'pessoal', 'equipe', 'funcionários', 'funcionarios']
        
        if sheet_name:
            rh_sheet = workbook[sheet_name]
        else:
            for name in workbook.sheetnames:
                if self._normalize_header(name) in rh_sheet_names:
                    rh_sheet = workbook[name]
                    break
        
        if not rh_sheet:
            # Usar primeira aba
            rh_sheet = workbook.active
            self.warnings.append({
                'type': 'sheet_detection',
                'message': f'Aba de RH não encontrada, usando aba ativa: {rh_sheet.title}'
            })
        
        # Ler cabeçalhos
        headers = [cell.value for cell in rh_sheet[1]]
        
        # Mapear colunas
        column_map = {}
        for field, possible_names in self.RH_COLUMNS.items():
            idx = self._find_column_index(headers, possible_names)
            if idx >= 0:
                column_map[field] = idx
        
        if 'nome_funcao' not in column_map:
            self.errors.append({
                'type': 'missing_column',
                'sheet': rh_sheet.title,
                'message': 'Coluna de função/cargo não encontrada'
            })
            return []
        
        # Processar linhas
        results = []
        for row_idx, row in enumerate(rh_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Linha vazia
                continue
            
            nome_funcao = row[column_map.get('nome_funcao', 0)]
            if not nome_funcao:
                continue
            
            rh_item = {
                'recurso_id': f"rh_{uuid.uuid4().hex[:12]}",
                'nome_funcao': str(nome_funcao).strip(),
                'regime_contratacao': str(row[column_map.get('regime_contratacao', 1)] or 'CLT').strip() if 'regime_contratacao' in column_map else 'CLT',
                'carga_horaria_semanal': self._parse_int(row[column_map.get('carga_horaria', -1)] if 'carga_horaria' in column_map else None, 40),
                'salario_bruto': self._parse_number(row[column_map.get('salario_bruto', -1)] if 'salario_bruto' in column_map else None),
                'encargos_mensais': self._parse_number(row[column_map.get('encargos', -1)] if 'encargos' in column_map else None),
                'numero_meses': self._parse_int(row[column_map.get('meses', -1)] if 'meses' in column_map else None, 12),
                'observacoes': str(row[column_map.get('observacoes', -1)] or '').strip() if 'observacoes' in column_map else '',
                'row_number': row_idx
            }
            
            # Calcular custo total
            custo_mensal = rh_item['salario_bruto'] + rh_item['encargos_mensais']
            rh_item['custo_mensal_total'] = custo_mensal
            rh_item['custo_total_projeto'] = custo_mensal * rh_item['numero_meses']
            
            results.append(rh_item)
        
        self.imported_rh = results
        return results
    
    def parse_despesas_sheet(self, workbook, sheet_name: str = None) -> List[Dict]:
        """
        Parseia a aba de Despesas
        
        Args:
            workbook: Workbook do openpyxl
            sheet_name: Nome da aba (opcional, tenta detectar automaticamente)
        
        Returns:
            Lista de dicionários com os dados de despesas
        """
        # Tentar encontrar a aba de despesas
        despesas_sheet = None
        despesas_sheet_names = ['despesas', 'materiais', 'custos', 'orçamento', 'orcamento', 'planilha']
        
        if sheet_name:
            despesas_sheet = workbook[sheet_name]
        else:
            for name in workbook.sheetnames:
                if self._normalize_header(name) in despesas_sheet_names:
                    despesas_sheet = workbook[name]
                    break
        
        if not despesas_sheet:
            # Tentar segunda aba
            if len(workbook.sheetnames) > 1:
                despesas_sheet = workbook[workbook.sheetnames[1]]
            else:
                despesas_sheet = workbook.active
            self.warnings.append({
                'type': 'sheet_detection',
                'message': f'Aba de despesas não encontrada, usando: {despesas_sheet.title}'
            })
        
        # Ler cabeçalhos
        headers = [cell.value for cell in despesas_sheet[1]]
        
        # Mapear colunas
        column_map = {}
        for field, possible_names in self.DESPESA_COLUMNS.items():
            idx = self._find_column_index(headers, possible_names)
            if idx >= 0:
                column_map[field] = idx
        
        if 'descricao' not in column_map and 'item' not in column_map:
            self.errors.append({
                'type': 'missing_column',
                'sheet': despesas_sheet.title,
                'message': 'Coluna de descrição/item não encontrada'
            })
            return []
        
        # Processar linhas
        results = []
        for row_idx, row in enumerate(despesas_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Linha vazia
                continue
            
            descricao = row[column_map.get('descricao', column_map.get('item', 0))]
            if not descricao:
                continue
            
            valor_unitario = self._parse_number(row[column_map.get('valor_unitario', -1)] if 'valor_unitario' in column_map else None)
            quantidade = self._parse_number(row[column_map.get('quantidade', -1)] if 'quantidade' in column_map else None, 1)
            
            despesa_item = {
                'despesa_id': f"desp_{uuid.uuid4().hex[:12]}",
                'natureza_despesa': str(row[column_map.get('natureza_despesa', -1)] or '339030').strip() if 'natureza_despesa' in column_map else '339030',
                'item_despesa': str(row[column_map.get('item', -1)] or '').strip() if 'item' in column_map else '',
                'descricao': str(descricao).strip(),
                'unidade': str(row[column_map.get('unidade', -1)] or 'UN').strip() if 'unidade' in column_map else 'UN',
                'quantidade': quantidade,
                'valor_unitario': valor_unitario,
                'valor_total': valor_unitario * quantidade,
                'orcamento_1': self._parse_number(row[column_map.get('orcamento_1', -1)] if 'orcamento_1' in column_map else None),
                'orcamento_2': self._parse_number(row[column_map.get('orcamento_2', -1)] if 'orcamento_2' in column_map else None),
                'orcamento_3': self._parse_number(row[column_map.get('orcamento_3', -1)] if 'orcamento_3' in column_map else None),
                'justificativa': str(row[column_map.get('justificativa', -1)] or '').strip() if 'justificativa' in column_map else '',
                'row_number': row_idx
            }
            
            # Calcular média dos orçamentos
            orcamentos = [despesa_item['orcamento_1'], despesa_item['orcamento_2'], despesa_item['orcamento_3']]
            orcamentos_validos = [o for o in orcamentos if o > 0]
            if orcamentos_validos:
                despesa_item['media_orcamentos'] = sum(orcamentos_validos) / len(orcamentos_validos)
            else:
                despesa_item['media_orcamentos'] = valor_unitario
            
            results.append(despesa_item)
        
        self.imported_despesas = results
        return results
    
    def import_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """
        Importa dados de um arquivo Excel em bytes
        
        Args:
            file_bytes: Bytes do arquivo Excel
        
        Returns:
            Dicionário com resultados da importação
        """
        self.errors = []
        self.warnings = []
        self.imported_rh = []
        self.imported_despesas = []
        
        try:
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
            
            # Log das abas encontradas
            logger.info(f"Planilha carregada com {len(workbook.sheetnames)} abas: {workbook.sheetnames}")
            
            # Parsear RH
            rh_data = self.parse_rh_sheet(workbook)
            
            # Parsear Despesas
            despesas_data = self.parse_despesas_sheet(workbook)
            
            # Calcular totais
            total_rh = sum(item.get('custo_total_projeto', 0) for item in rh_data)
            total_despesas = sum(item.get('valor_total', 0) for item in despesas_data)
            
            return {
                'success': len(self.errors) == 0,
                'rh': {
                    'count': len(rh_data),
                    'total': total_rh,
                    'items': rh_data
                },
                'despesas': {
                    'count': len(despesas_data),
                    'total': total_despesas,
                    'items': despesas_data
                },
                'totals': {
                    'rh': total_rh,
                    'despesas': total_despesas,
                    'geral': total_rh + total_despesas
                },
                'errors': self.errors,
                'warnings': self.warnings,
                'sheets_found': workbook.sheetnames
            }
            
        except Exception as e:
            logger.error(f"Erro ao importar planilha: {e}")
            self.errors.append({
                'type': 'parse_error',
                'message': f'Erro ao processar planilha: {str(e)}'
            })
            return {
                'success': False,
                'rh': {'count': 0, 'total': 0, 'items': []},
                'despesas': {'count': 0, 'total': 0, 'items': []},
                'totals': {'rh': 0, 'despesas': 0, 'geral': 0},
                'errors': self.errors,
                'warnings': self.warnings,
                'sheets_found': []
            }
    
    def generate_template(self) -> bytes:
        """
        Gera um template Excel para importação
        
        Returns:
            Bytes do arquivo Excel template
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aba de Recursos Humanos
        ws_rh = wb.active
        ws_rh.title = "Recursos Humanos"
        
        rh_headers = [
            "Função/Cargo", "Regime de Contratação", "Carga Horária Semanal",
            "Salário Bruto (R$)", "Encargos Mensais (R$)", "Quantidade de Meses", "Observações"
        ]
        
        for col, header in enumerate(rh_headers, 1):
            cell = ws_rh.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_rh.column_dimensions[get_column_letter(col)].width = 20
        
        # Exemplo de RH
        ws_rh.append(["Coordenador do Projeto", "CLT", 40, 5000.00, 2000.00, 12, ""])
        ws_rh.append(["Assistente Administrativo", "CLT", 40, 2500.00, 1000.00, 12, ""])
        
        # Aba de Despesas
        ws_despesas = wb.create_sheet(title="Despesas")
        
        despesas_headers = [
            "Natureza da Despesa", "Item", "Descrição/Especificação", "Unidade",
            "Quantidade", "Valor Unitário (R$)", "Orçamento 1 (R$)", 
            "Orçamento 2 (R$)", "Orçamento 3 (R$)", "Justificativa"
        ]
        
        for col, header in enumerate(despesas_headers, 1):
            cell = ws_despesas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_despesas.column_dimensions[get_column_letter(col)].width = 18
        
        # Exemplos de despesas
        ws_despesas.append(["339030", "01", "Material de escritório", "Kit", 10, 150.00, 145.00, 155.00, 150.00, "Uso administrativo"])
        ws_despesas.append(["339039", "02", "Serviço de internet", "Mês", 12, 200.00, 190.00, 210.00, 200.00, "Comunicação da equipe"])
        
        # Aba de Instruções
        ws_inst = wb.create_sheet(title="Instruções")
        ws_inst.merge_cells('A1:F1')
        ws_inst['A1'] = "INSTRUÇÕES PARA PREENCHIMENTO"
        ws_inst['A1'].font = Font(bold=True, size=14, color="1F4E78")
        
        instructions = [
            "",
            "1. Preencha a aba 'Recursos Humanos' com todos os profissionais do projeto",
            "2. Preencha a aba 'Despesas' com todos os materiais e serviços necessários",
            "3. Use valores numéricos sem formatação de moeda (ex: 1500.00)",
            "4. Não altere os nomes das colunas na primeira linha",
            "",
            "NATUREZAS DE DESPESA COMUNS:",
            "- 339030: Material de Consumo",
            "- 339039: Outros Serviços de Terceiros - PJ",
            "- 339036: Outros Serviços de Terceiros - PF",
            "- 449052: Equipamentos e Material Permanente",
            "",
            "REGIMES DE CONTRATAÇÃO:",
            "- CLT: Consolidação das Leis do Trabalho",
            "- MEI: Microempreendedor Individual",
            "- Autônomo: Prestador de serviços autônomo",
            "- Estágio: Estagiário"
        ]
        
        for row, text in enumerate(instructions, 2):
            ws_inst.cell(row=row, column=1, value=text)
        
        # Salvar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# Instância global do importador
mrosc_importer = MROSCExcelImporter()
