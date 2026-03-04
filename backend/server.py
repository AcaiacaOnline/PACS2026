from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, BackgroundTasks, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import re
import logging
import hashlib
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from io import BytesIO
import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from striprtf.striprtf import rtf_to_text

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ============ IMPORTAÇÃO DE MODELOS REFATORADOS ============
# Os modelos agora são importados do pacote 'models/' para reduzir duplicação
from models import (
    # User models
    User as UserModel, UserCreate as UserCreateModel, UserUpdate as UserUpdateModel,
    UserLogin as UserLoginModel, UserListItem as UserListItemModel,
    UserPermissions as UserPermissionsModel, UserSignatureData as UserSignatureDataModel,
    # PAC models
    PAC as PACModel, PACCreate as PACCreateModel, PACUpdate as PACUpdateModel,
    PACItem as PACItemModel, PACItemCreate as PACItemCreateModel, PACItemUpdate as PACItemUpdateModel,
    PACGeral as PACGeralModel, PACGeralCreate as PACGeralCreateModel, PACGeralUpdate as PACGeralUpdateModel,
    PACGeralItem as PACGeralItemModel, PACGeralItemCreate as PACGeralItemCreateModel, PACGeralItemUpdate as PACGeralItemUpdateModel,
    # PAC Obras models
    PACGeralObras as PACGeralObrasModel, PACGeralObrasCreate as PACGeralObrasCreateModel, PACGeralObrasUpdate as PACGeralObrasUpdateModel,
    PACGeralObrasItem as PACGeralObrasItemModel, PACGeralObrasItemCreate as PACGeralObrasItemCreateModel, PACGeralObrasItemUpdate as PACGeralObrasItemUpdateModel,
    CLASSIFICACAO_OBRAS_SERVICOS,
    # Processo models
    Processo as ProcessoModel, ProcessoCreate as ProcessoCreateModel, ProcessoUpdate as ProcessoUpdateModel,
    PaginatedResponse as PaginatedResponseModel,
    # MROSC models
    ProjetoMROSC as ProjetoMROSCModel, ProjetoMROSCCreate as ProjetoMROSCCreateModel, ProjetoMROSCUpdate as ProjetoMROSCUpdateModel,
    RecursoHumanoMROSC as RecursoHumanoMROSCModel, RecursoHumanoMROSCCreate as RecursoHumanoMROSCCreateModel,
    DespesaMROSC as DespesaMROSCModel, DespesaMROSCCreate as DespesaMROSCCreateModel,
    DocumentoMROSC as DocumentoMROSCModel, DocumentoMROSCCreate as DocumentoMROSCCreateModel,
    NATUREZAS_DESPESA_MROSC,
)

# ============ CONFIGURAÇÕES DE EMAIL ============
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'mail.acaiaca.mg.gov.br')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '465'))
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', 'naoresponda@acaiaca.mg.gov.br')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_USE_SSL = os.environ.get('SMTP_USE_SSL', 'true').lower() == 'true'

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'pac-acaiaca-secret-key-2026')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_DAYS = 7

# ===== CONFIGURAÇÃO DO FASTAPI COM SWAGGER =====
app = FastAPI(
    title="Planejamento Acaiaca - Sistema de Gestão Municipal",
    description="""
## Sistema de Gestão Municipal - Prefeitura de Acaiaca/MG

API REST completa para gestão municipal incluindo:

### 📋 Módulos Disponíveis
- **PAC Individual** - Planos Anuais de Contratação por Secretaria
- **PAC Geral** - Visão Consolidada de todas as Secretarias
- **PAC Obras** - Obras e Serviços de Engenharia
- **Gestão Processual** - Controle de Processos Licitatórios
- **MROSC** - Prestação de Contas de OSCs (Lei 13.019/2014)
- **Portal de Transparência** - Acesso Público aos Dados

### 🔐 Autenticação
Utilize o endpoint `/api/auth/login` para obter um token JWT.
Inclua o token no header `Authorization: Bearer {token}` em todas as requisições autenticadas.

### 📊 WebSocket
Notificações em tempo real disponíveis em `/api/ws/notifications/{user_id}`

### 📧 Contato
- Email: planejamento@acaiaca.mg.gov.br
- CNPJ: 18.295.287/0001-90
    """,
    version="2.0.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    openapi_url="/api/openapi.json",
    openapi_tags=[
        {"name": "Autenticação", "description": "Login, registro e gerenciamento de sessão"},
        {"name": "Usuários", "description": "Gerenciamento de usuários do sistema"},
        {"name": "PAC Individual", "description": "Planos Anuais de Contratação por Secretaria"},
        {"name": "PAC Geral", "description": "PAC consolidado de todas as secretarias"},
        {"name": "PAC Obras", "description": "Obras e serviços de engenharia"},
        {"name": "Gestão Processual", "description": "Processos licitatórios"},
        {"name": "MROSC", "description": "Prestação de Contas - Lei 13.019/2014"},
        {"name": "Portal da Transparência", "description": "Endpoints públicos para acesso cidadão"},
        {"name": "Dashboard Analítico", "description": "Estatísticas e métricas do sistema"},
        {"name": "Sistema de Alertas", "description": "Alertas e notificações automatizadas"},
        {"name": "Relatórios Gerenciais", "description": "Geração de relatórios PDF/XLSX"},
        {"name": "Backup", "description": "Backup e restauração de dados"},
        {"name": "WebSocket Notifications", "description": "Notificações em tempo real via WebSocket"},
        {"name": "Assinaturas Digitais", "description": "Assinatura e validação de documentos conforme Lei 14.063"},
        {"name": "Validação de Documentos", "description": "Validação de autenticidade de documentos assinados"},
    ],
    contact={
        "name": "Prefeitura Municipal de Acaiaca",
        "url": "https://acaiaca.mg.gov.br",
        "email": "planejamento@acaiaca.mg.gov.br"
    },
    license_info={
        "name": "Uso Restrito - Prefeitura de Acaiaca/MG",
        "url": "https://acaiaca.mg.gov.br"
    }
)

# Importar logging primeiro
from utils.logging_config import get_logger, log_info, log_error, request_logger
logger = get_logger("server")

# Importar e registrar middleware de performance
try:
    from middleware.performance import PerformanceMiddleware, rate_limiter, response_cache
    app.add_middleware(PerformanceMiddleware)
    logger.info("✅ Middleware de performance carregado")
except ImportError as e:
    logger.warning(f"⚠️ Middleware de performance não disponível: {e}")

# Importar e registrar WebSocket router
from utils.websocket import router as ws_router, manager as ws_manager
from utils.websocket import (
    notify_processo_created, notify_mrosc_submitted, notify_mrosc_approved,
    notify_mrosc_correction, notify_system_alert, notify_backup_completed,
    Notification, NotificationType
)

# Importar utilitários de PDF (funções refatoradas)
from utils.pdf_utils import (
    mask_cpf as mask_cpf_util,
    generate_validation_code as generate_validation_code_util,
    get_professional_styles as get_professional_styles_util,
    draw_signature_seal as draw_signature_seal_util,
    create_signature_page_mrosc as create_signature_page_mrosc_util
)

api_router = APIRouter(prefix="/api")

# Models
# NOTA: Modelos refatorados agora são importados do pacote 'models/'
# Os aliases abaixo mantêm compatibilidade com o código existente

# ============ ALIASES PARA MODELOS REFATORADOS ============
# Estes aliases apontam para os modelos no pacote 'models/' para manter
# compatibilidade com o código existente enquanto a refatoração é concluída.

# User models - usando aliases do pacote models/
UserPermissions = UserPermissionsModel
UserSignatureData = UserSignatureDataModel
User = UserModel
UserCreate = UserCreateModel
UserUpdate = UserUpdateModel
UserLogin = UserLoginModel
UserListItem = UserListItemModel

# PAC models - usando aliases do pacote models/
PAC = PACModel
PACCreate = PACCreateModel
PACUpdate = PACUpdateModel
PACItem = PACItemModel
PACItemCreate = PACItemCreateModel
PACItemUpdate = PACItemUpdateModel

# PAC Geral models - usando aliases do pacote models/
PACGeral = PACGeralModel
PACGeralCreate = PACGeralCreateModel
PACGeralUpdate = PACGeralUpdateModel
PACGeralItem = PACGeralItemModel
PACGeralItemCreate = PACGeralItemCreateModel
PACGeralItemUpdate = PACGeralItemUpdateModel

# PAC Obras models - usando aliases do pacote models/
PACGeralObras = PACGeralObrasModel
PACGeralObrasCreate = PACGeralObrasCreateModel
PACGeralObrasUpdate = PACGeralObrasUpdateModel
PACGeralObrasItem = PACGeralObrasItemModel
PACGeralObrasItemCreate = PACGeralObrasItemCreateModel
PACGeralObrasItemUpdate = PACGeralObrasItemUpdateModel

# Processo models - usando aliases do pacote models/
Processo = ProcessoModel
ProcessoCreate = ProcessoCreateModel
ProcessoUpdate = ProcessoUpdateModel
PaginatedResponse = PaginatedResponseModel

# MROSC models - usando aliases do pacote models/
ProjetoMROSC = ProjetoMROSCModel
ProjetoMROSCCreate = ProjetoMROSCCreateModel
ProjetoMROSCUpdate = ProjetoMROSCUpdateModel
RecursoHumanoMROSC = RecursoHumanoMROSCModel
RecursoHumanoMROSCCreate = RecursoHumanoMROSCCreateModel
DespesaMROSC = DespesaMROSCModel
DespesaMROSCCreate = DespesaMROSCCreateModel
DocumentoMROSC = DocumentoMROSCModel
DocumentoMROSCCreate = DocumentoMROSCCreateModel

# ============ MODELOS ESPECÍFICOS DO SERVER (NÃO REFATORADOS) ============
# Estes modelos são usados apenas internamente e não foram movidos para o pacote models/

class SignatureRequest(BaseModel):
    """Solicitação de assinatura digital com data opcional"""
    confirmar_assinatura: bool = False  # Deve ser True para confirmar
    data_assinatura: Optional[str] = None  # Formato: DD/MM/YYYY HH:MM:SS (retroativa ou futura)
    observacoes: Optional[str] = None  # Observações opcionais

class SignaturePayload(BaseModel):
    """Payload para assinatura de documento"""
    data_assinatura: Optional[str] = None  # Data customizada de assinatura

# ============ CONSTANTES ============
# Constantes de classificação são importadas do pacote models/

# ============ FUNÇÕES DE AUTENTICAÇÃO ============
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(user_id: str) -> str:
    payload = {
        'user_id': user_id,
        'exp': datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRATION_DAYS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

# Alias para função refatorada
mask_cpf = mask_cpf_util

async def get_current_user(request: Request) -> User:
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = auth_header.split(' ')[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        user_data = await db.users.find_one({'user_id': user_id}, {'_id': 0, 'password_hash': 0})
        if not user_data:
            raise HTTPException(status_code=401, detail="User not found")
        if not user_data.get('is_active', True):
            raise HTTPException(status_code=403, detail="User account is disabled")
        return User(**user_data)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    except Exception as e:
        raise HTTPException(status_code=401, detail=str(e))

@app.on_event("startup")
async def create_admin_user():
    admin_email = "cristiano.abdo@acaiaca.mg.gov.br"
    admin_exists = await db.users.find_one({'email': admin_email})
    if not admin_exists:
        admin_user = {
            'user_id': f"user_{uuid.uuid4().hex[:12]}",
            'email': admin_email,
            'name': "Cristiano Abdo de Souza",
            'password_hash': hash_password("Cris@820906"),
            'is_admin': True,
            'is_active': True,
            'picture': None,
            'created_at': datetime.now(timezone.utc)
        }
        await db.users.insert_one(admin_user)
        logging.info(f"Admin user created: {admin_email}")

@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({'email': user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    
    # Definir permissões baseadas no tipo de usuário
    tipo_usuario = getattr(user_data, 'tipo_usuario', 'SERVIDOR')
    permissions = {
        'can_view': True,
        'can_edit': tipo_usuario == 'SERVIDOR',
        'can_delete': False,
        'can_export': tipo_usuario == 'SERVIDOR',
        'can_manage_users': False,
        'is_full_admin': False,
        'mrosc_only': tipo_usuario == 'PESSOA_EXTERNA'  # Pessoa externa só acessa MROSC
    }
    
    user_doc = {
        'user_id': user_id,
        'email': user_data.email,
        'name': user_data.name,
        'password_hash': hash_password(user_data.password),
        'is_admin': False,
        'is_active': True,
        'tipo_usuario': tipo_usuario,
        'picture': None,
        'permissions': permissions,
        'created_at': datetime.now(timezone.utc)
    }
    await db.users.insert_one(user_doc)
    user_doc.pop('password_hash')
    return User(**user_doc)

@api_router.post("/auth/login")
async def login(credentials: UserLogin, response: Response):
    user_doc = await db.users.find_one({'email': credentials.email}, {'_id': 0})
    if not user_doc or not verify_password(credentials.password, user_doc['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_jwt_token(user_doc['user_id'])
    response.set_cookie(key='session_token', value=token, httponly=True, secure=True, samesite='none', max_age=JWT_EXPIRATION_DAYS*24*60*60, path='/')
    user_doc.pop('password_hash')
    return {'token': token, 'user': User(**user_doc)}

@api_router.get("/auth/me", response_model=User)
async def get_me(request: Request):
    return await get_current_user(request)

@api_router.get("/auth/oauth/session")
async def oauth_session(request: Request, response: Response):
    session_id = request.headers.get('X-Session-ID')
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing session ID")
    async with httpx.AsyncClient() as client:
        try:
            auth_response = await client.get('https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data', headers={'X-Session-ID': session_id})
            auth_response.raise_for_status()
            data = auth_response.json()
        except Exception as e:
            logging.error(f"OAuth error: {str(e)}")
            raise HTTPException(status_code=401, detail=f"OAuth failed: {str(e)}")
    
    # Buscar ou criar usuário
    user_doc = await db.users.find_one({'email': data['email']}, {'_id': 0})
    if not user_doc:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            'user_id': user_id, 
            'email': data['email'], 
            'name': data['name'], 
            'password_hash': None, 
            'is_admin': False, 
            'is_active': True, 
            'picture': data.get('picture'), 
            'created_at': datetime.now(timezone.utc)
        }
        await db.users.insert_one(user_doc)
        # Remover _id após insert
        user_doc.pop('_id', None)
    
    # Criar JWT token para o usuário OAuth
    token_data = {
        "user_id": user_doc['user_id'],
        "exp": datetime.now(timezone.utc) + timedelta(days=7)
    }
    jwt_token = jwt.encode(token_data, JWT_SECRET, algorithm=JWT_ALGORITHM)
    
    # Criar sessão
    session_doc = {
        'user_id': user_doc['user_id'], 
        'session_token': data['session_token'], 
        'expires_at': datetime.now(timezone.utc) + timedelta(days=7), 
        'created_at': datetime.now(timezone.utc)
    }
    await db.user_sessions.insert_one(session_doc)
    response.set_cookie(key='session_token', value=data['session_token'], httponly=True, secure=True, samesite='none', max_age=7*24*60*60, path='/')
    
    # Preparar resposta (remover campos sensíveis)
    user_doc.pop('password_hash', None)
    user_doc.pop('_id', None)
    
    # Converter datetime para string se necessário
    if 'created_at' in user_doc and isinstance(user_doc['created_at'], datetime):
        user_doc['created_at'] = user_doc['created_at'].isoformat()
    
    # Retornar usuário com token
    return {
        **user_doc,
        "token": jwt_token
    }

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    token = request.cookies.get('session_token')
    if token:
        await db.user_sessions.delete_many({'session_token': token})
    response.delete_cookie('session_token', path='/')
    return {'message': 'Logged out'}

# ============ CLASSIFICAÇÃO ORÇAMENTÁRIA (LEI 14.133/2021) ============

@api_router.get("/classificacao/codigos")
async def get_classificacao_codigos():
    """
    Retorna todos os códigos de classificação orçamentária conforme Lei 14.133/2021
    """
    codigos = {
        "339030": {
            "nome": "Material de Consumo",
            "subitens": [
                "Material de Consumo - Combustíveis e Lubrificantes Automotivos",
                "Combustíveis e Lubrificantes de Aviação",
                "Gás Engarrafado",
                "Explosivos e Munições",
                "Alimentos para Animais",
                "Gêneros de Alimentação",
                "Animais para Pesquisa e Abate",
                "Material Farmacológico",
                "Material Odontológico",
                "Material Químico",
                "Material de Coudelaria / Zootécnico",
                "Material de Caça e Pesca",
                "Material Educativo e Esportivo",
                "Material para Festividades e Homenagens",
                "Material de Expediente",
                "Material de Processamento de Dados",
                "Material de Acondicionamento e Embalagem",
                "Material de Cama, Mesa e Banho",
                "Material de Copa e Cozinha",
                "Material de Limpeza e Higienização",
                "Uniformes, Tecidos e Aviamentos",
                "Material para Manutenção de Bens Imóveis",
                "Material para Manutenção de Bens Móveis",
                "Material Elétrico e Eletrônico",
                "Material de Proteção e Segurança",
                "Material para Áudio, Vídeo e Foto",
                "Sementes, Mudas e Insumos",
                "Material Hospitalar",
                "Material para Manutenção de Veículos",
                "Ferramentas",
                "Material de Sinalização Visual"
            ]
        },
        "339036": {
            "nome": "Outros Serviços de Terceiros (Pessoa Física)",
            "subitens": [
                "Diárias a Colaboradores Eventuais",
                "Serviços Técnicos Profissionais",
                "Estagiários",
                "Locação de Imóveis",
                "Locação de Bens Móveis",
                "Manutenção e Conservação de Equipamentos",
                "Manutenção e Conservação de Veículos",
                "Manutenção e Conservação de Imóveis",
                "Serviços de Limpeza e Conservação",
                "Serviços Médicos e Odontológicos",
                "Serviços de Apoio Administrativo/Técnico",
                "Confecção de Uniformes/Bandeiras",
                "Fretes e Transportes de Encomendas",
                "Jetons"
            ]
        },
        "339039": {
            "nome": "Outros Serviços de Terceiros (Pessoa Jurídica)",
            "subitens": [
                "Assinaturas de Periódicos",
                "Condomínios",
                "Serviços Técnicos Profissionais",
                "Manutenção de Software",
                "Locação de Imóveis",
                "Locação de Máquinas e Equipamentos",
                "Manutenção e Conservação de Imóveis",
                "Manutenção e Conservação de Máquinas",
                "Manutenção e Conservação de Veículos",
                "Exposições, Congressos e Festividades",
                "Serviços de Energia Elétrica",
                "Serviços de Água e Esgoto",
                "Serviços de Telecomunicações",
                "Serviços Gráficos",
                "Seguros em Geral",
                "Confecção de Uniformes e Bandeiras",
                "Vale-Transporte",
                "Vigilância Ostensiva",
                "Limpeza e Conservação",
                "Serviços Bancários",
                "Serviços de Cópias e Reprodução",
                "Publicidade e Propaganda"
            ]
        },
        "449052": {
            "nome": "Equipamentos e Material Permanente",
            "subitens": [
                "Aparelhos de Medição e Orientação",
                "Aparelhos e Equipamentos de Comunicação",
                "Equipamentos Médico-Hospitalares",
                "Aparelhos e Utensílios Domésticos",
                "Coleções e Materiais Bibliográficos",
                "Mobiliário em Geral",
                "Equipamentos de Processamento de Dados",
                "Máquinas e Utensílios de Escritório",
                "Máquinas, Ferramentas e Utensílios de Oficina",
                "Equipamentos Hidráulicos e Elétricos",
                "Máquinas Agrícolas e Rodoviárias",
                "Veículos de Tração Mecânica",
                "Veículos Diversos",
                "Acessórios para Automóveis"
            ]
        }
    }
    return codigos

# ============ USER MANAGEMENT ROUTES (ADMIN ONLY) ============

async def require_admin(request: Request) -> User:
    user = await get_current_user(request)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

@api_router.get("/users", response_model=List[UserListItem])
async def get_users(request: Request):
    await require_admin(request)
    users = await db.users.find({}, {'_id': 0, 'password_hash': 0}).to_list(1000)
    return [UserListItem(**u) for u in users]

@api_router.post("/users", response_model=UserListItem)
async def create_user_admin(user_data: UserCreate, request: Request):
    await require_admin(request)
    existing = await db.users.find_one({'email': user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    
    # Preparar permissões
    permissions_data = None
    if user_data.permissions:
        permissions_data = user_data.permissions.model_dump()
    elif user_data.is_admin:
        # Admin tem todas as permissões por padrão
        permissions_data = {
            'can_view': True,
            'can_edit': True,
            'can_delete': True,
            'can_export': True,
            'can_manage_users': True,
            'is_full_admin': True
        }
    else:
        # Usuário padrão só pode visualizar
        permissions_data = {
            'can_view': True,
            'can_edit': False,
            'can_delete': False,
            'can_export': False,
            'can_manage_users': False,
            'is_full_admin': False
        }
    
    user_doc = {
        'user_id': user_id,
        'email': user_data.email,
        'name': user_data.name,
        'password_hash': hash_password(user_data.password),
        'is_admin': user_data.is_admin,
        'is_active': True,
        'picture': None,
        'permissions': permissions_data,
        'signature_data': user_data.signature_data.model_dump() if user_data.signature_data else None,
        'created_at': datetime.now(timezone.utc)
    }
    await db.users.insert_one(user_doc)
    user_doc.pop('password_hash')
    return UserListItem(**user_doc)

@api_router.put("/users/{user_id}", response_model=UserListItem)
async def update_user_admin(user_id: str, user_update: UserUpdate, request: Request):
    admin = await require_admin(request)
    if user_id == admin.user_id and user_update.is_admin is False:
        raise HTTPException(status_code=400, detail="Cannot remove your own admin rights")
    user = await db.users.find_one({'user_id': user_id}, {'_id': 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {}
    for k, v in user_update.model_dump().items():
        if v is not None:
            if k == 'permissions' and isinstance(v, dict):
                update_data['permissions'] = v
            elif k == 'signature_data' and isinstance(v, dict):
                update_data['signature_data'] = v
            else:
                update_data[k] = v
    
    if 'password' in update_data:
        update_data['password_hash'] = hash_password(update_data.pop('password'))
    
    # Se is_full_admin for marcado, atualiza is_admin também
    if update_data.get('permissions', {}).get('is_full_admin'):
        update_data['is_admin'] = True
    
    await db.users.update_one({'user_id': user_id}, {'$set': update_data})
    updated_user = await db.users.find_one({'user_id': user_id}, {'_id': 0, 'password_hash': 0})
    return UserListItem(**updated_user)

@api_router.delete("/users/{user_id}")
async def delete_user_admin(user_id: str, request: Request):
    admin = await require_admin(request)
    if user_id == admin.user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    result = await db.users.delete_one({'user_id': user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    await db.user_sessions.delete_many({'user_id': user_id})
    return {'message': 'User deleted'}

# ============ PAC ROUTES (WITH PERMISSIONS) ============

# ===== CONSTANTES DE MARGENS PADRONIZADAS (Lei 14.133/2021) =====
# Margens: 5cm (esquerda/direita), 3cm (superior/inferior)
REPORT_MARGIN_LEFT = 50*mm    # 5cm
REPORT_MARGIN_RIGHT = 50*mm   # 5cm
REPORT_MARGIN_TOP = 30*mm     # 3cm
REPORT_MARGIN_BOTTOM = 30*mm  # 3cm

# ===== CONSTANTES PARA PAGINAÇÃO DE ITENS =====
ITEMS_PER_PAGE = 15  # Máximo de 15 itens por página conforme solicitado

# ===== ESTILOS DE PDF PROFISSIONAIS =====
def get_professional_styles():
    """Retorna estilos profissionais para relatórios PDF"""
    styles = getSampleStyleSheet()
    
    # Cores do tema
    cor_primaria = colors.HexColor('#1F4E78')  # Azul escuro institucional
    cor_secundaria = colors.HexColor('#2E75B6')  # Azul médio
    cor_destaque = colors.HexColor('#FFC000')  # Amarelo dourado
    cor_texto = colors.HexColor('#333333')  # Cinza escuro
    cor_subtexto = colors.HexColor('#666666')  # Cinza médio
    
    custom_styles = {
        'title': ParagraphStyle(
            'ProfTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=cor_primaria,
            alignment=TA_CENTER,
            spaceAfter=4,
            fontName='Helvetica-Bold',
            leading=22
        ),
        'subtitle': ParagraphStyle(
            'ProfSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=cor_primaria,
            alignment=TA_CENTER,
            spaceAfter=3,
            fontName='Helvetica-Bold'
        ),
        'section_header': ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading3'],
            fontSize=11,
            textColor=cor_primaria,
            alignment=TA_LEFT,
            spaceAfter=6,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            borderPadding=4
        ),
        'body': ParagraphStyle(
            'ProfBody',
            parent=styles['Normal'],
            fontSize=10,
            textColor=cor_texto,
            alignment=TA_JUSTIFY,
            spaceAfter=6,
            leading=14
        ),
        'small': ParagraphStyle(
            'ProfSmall',
            parent=styles['Normal'],
            fontSize=8,
            textColor=cor_subtexto,
            alignment=TA_LEFT,
            spaceAfter=2
        ),
        'legal': ParagraphStyle(
            'Legal',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique',
            spaceAfter=8
        ),
        'footer': ParagraphStyle(
            'ProfFooter',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        ),
        'table_cell': ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontSize=7,
            textColor=cor_texto,
            alignment=TA_LEFT,
            leading=9
        ),
        'table_header': ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
    }
    
    return custom_styles, cor_primaria, cor_secundaria, cor_destaque

def create_professional_header(pac_data: dict, styles: dict, is_pac_geral: bool = False):
    """Cria cabeçalho no estilo oficial (Diário Oficial Eletrônico Municipal)"""
    from utils.pdf_utils import create_oficial_header_elements
    
    # Título do documento
    if is_pac_geral:
        titulo = 'PAC GERAL - PLANO ANUAL DE CONTRATAÇÕES CONSOLIDADO'
    else:
        titulo = 'PAC - PLANO ANUAL DE CONTRATAÇÕES'
    
    ano = pac_data.get("ano", datetime.now().year)
    subtitulo = f'Exercício {ano} - Lei Federal nº 14.133/2021'
    
    header_elements = create_oficial_header_elements(
        styles,
        titulo_documento=titulo,
        subtitulo=subtitulo
    )
    
    return header_elements

def create_info_box(pac_data: dict, styles: dict, is_pac_geral: bool = False):
    """Cria caixa de informações do PAC"""
    info_style = styles['small']
    cor_primaria = colors.HexColor('#1F4E78')
    
    if is_pac_geral:
        info_data = [
            [
                Paragraph(f'<b>PAC Geral:</b> {pac_data.get("nome_secretaria", "")}', info_style),
                Paragraph(f'<b>Exercício:</b> {pac_data.get("ano", "")}', info_style),
            ],
            [
                Paragraph(f'<b>Secretário(a):</b> {pac_data.get("secretario", "")}', info_style),
                Paragraph(f'<b>Fiscal do Contrato:</b> {pac_data.get("fiscal_contrato", "")}', info_style),
            ],
            [
                Paragraph(f'<b>Telefone:</b> {pac_data.get("telefone", "")}', info_style),
                Paragraph(f'<b>E-mail:</b> {pac_data.get("email", "")}', info_style),
            ]
        ]
        col_widths = [14*cm, 14*cm]
    else:
        info_data = [
            [
                Paragraph(f'<b>Secretaria/Órgão:</b> {pac_data.get("secretaria", "")}', info_style),
                Paragraph(f'<b>Exercício:</b> {pac_data.get("ano", "")}', info_style),
            ],
            [
                Paragraph(f'<b>Secretário(a) Responsável:</b> {pac_data.get("secretario", "")}', info_style),
                Paragraph(f'<b>Fiscal do Contrato:</b> {pac_data.get("fiscal", "")}', info_style),
            ],
            [
                Paragraph(f'<b>Telefone:</b> {pac_data.get("telefone", "")}', info_style),
                Paragraph(f'<b>E-mail:</b> {pac_data.get("email", "")}', info_style),
            ],
            [
                Paragraph(f'<b>Endereço:</b> {pac_data.get("endereco", "")}', info_style),
                '',
            ]
        ]
        col_widths = [14*cm, 14*cm]
    
    info_table = Table(info_data, colWidths=col_widths)
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F5F5')),
        ('BOX', (0, 0), (-1, -1), 1.5, cor_primaria),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    return info_table

def create_items_table_paginated(items: list, styles: dict, orientation: str = 'landscape', start_index: int = 1):
    """
    Cria tabela de itens com formatação profissional.
    Retorna lista de elementos (tabelas) para uma página.
    """
    cor_primaria = colors.HexColor('#1F4E78')
    cor_destaque = colors.HexColor('#FFC000')
    
    # Cabeçalhos
    headers = [
        '#', 
        'Código\nCATMAT', 
        'Descrição do Objeto', 
        'Unidade', 
        'Qtd', 
        'Valor\nUnitário', 
        'Valor\nTotal', 
        'Prioridade',
        'Justificativa',
        'Classificação\nOrçamentária'
    ]
    
    table_data = [headers]
    
    for idx, item in enumerate(items, start=start_index):
        # Classificação Orçamentária
        classificacao_text = ''
        if item.get('codigo_classificacao'):
            classificacao_text = f"{item['codigo_classificacao']}"
            if item.get('subitem_classificacao'):
                classificacao_text += f"\n{item['subitem_classificacao']}"
        
        descricao = item.get('descricao', '')
        justificativa = item.get('justificativa', '') or 'N/I'
        
        table_data.append([
            str(idx),
            item.get('catmat', ''),
            Paragraph(f"<font size=7>{descricao}</font>", styles['table_cell']),
            item.get('unidade', ''),
            str(int(item.get('quantidade', 0))),
            f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {item.get('valorTotal', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            item.get('prioridade', ''),
            Paragraph(f"<font size=6>{justificativa}</font>", styles['table_cell']),
            Paragraph(f"<font size=6>{classificacao_text}</font>", styles['table_cell'])
        ])
    
    # Larguras de coluna otimizadas
    if orientation.lower() == 'portrait':
        col_widths = [0.5*cm, 1.3*cm, 3.5*cm, 0.9*cm, 0.7*cm, 1.5*cm, 1.5*cm, 1*cm, 2.5*cm, 2.5*cm]
    else:
        col_widths = [0.5*cm, 1.4*cm, 5.5*cm, 1*cm, 0.8*cm, 1.6*cm, 1.8*cm, 1.2*cm, 5*cm, 4.5*cm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), cor_primaria),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Corpo
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # #
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),  # Unidade, Qtd
        ('ALIGN', (5, 1), (6, -1), 'RIGHT'),   # Valores
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),  # Prioridade
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        
        # Linhas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#D6EAF8')]),
        
        # Bordas
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX', (0, 0), (-1, -1), 1, cor_primaria),
        
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    return table

def create_total_row(total_value: float, styles: dict, orientation: str = 'landscape'):
    """Cria linha de total com destaque"""
    cor_destaque = colors.HexColor('#FFC000')
    
    total_formatted = f"R$ {total_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    
    if orientation.lower() == 'portrait':
        col_widths = [0.5*cm, 1.3*cm, 3.5*cm, 0.9*cm, 0.7*cm, 1.5*cm, 1.5*cm, 1*cm, 2.5*cm, 2.5*cm]
    else:
        col_widths = [0.5*cm, 1.4*cm, 5.5*cm, 1*cm, 0.8*cm, 1.6*cm, 1.8*cm, 1.2*cm, 5*cm, 4.5*cm]
    
    total_data = [['', '', Paragraph('<b>VALOR TOTAL ESTIMADO:</b>', styles['body']), '', '', '', total_formatted, '', '', '']]
    
    total_table = Table(total_data, colWidths=col_widths)
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), cor_destaque),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (6, 0), (6, 0), 'RIGHT'),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1F4E78')),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    return total_table

def create_signature_section(pac_data: dict, styles: dict, is_pac_geral: bool = False):
    """Cria seção de assinaturas profissional"""
    elements = []
    
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph('<b>ASSINATURAS E VALIDAÇÃO</b>', styles['section_header']))
    elements.append(Spacer(1, 4*mm))
    
    if is_pac_geral:
        responsavel = pac_data.get('secretario', '')
        fiscal = pac_data.get('fiscal_contrato', '')
    else:
        responsavel = pac_data.get('secretario', '')
        fiscal = pac_data.get('fiscal', '')
    
    sig_data = [
        ['_' * 45, '_' * 45],
        [Paragraph(f'<b>{responsavel}</b>', styles['body']), Paragraph(f'<b>{fiscal}</b>', styles['body'])],
        ['Secretário(a) Responsável', 'Fiscal do Contrato'],
        ['', ''],
        [f'Acaiaca/MG, ___/___/______', f'Acaiaca/MG, ___/___/______']
    ]
    
    sig_table = Table(sig_data, colWidths=[10*cm, 10*cm])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTSIZE', (0, 2), (-1, 2), 8),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.grey),
        ('TOPPADDING', (0, 0), (-1, 0), 0),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
        ('TOPPADDING', (0, 4), (-1, 4), 10),
    ]))
    
    elements.append(sig_table)
    
    return elements

def create_footer_text():
    """Cria texto de rodapé padrão"""
    return f'<font size=7><i>Documento gerado eletronicamente pelo Sistema Planejamento Acaiaca em {datetime.now().strftime("%d/%m/%Y às %H:%M")} | Desenvolvido por Cristiano Abdo de Souza - Assessor de Planejamento</i></font>'

@api_router.get("/pacs/anos")
async def get_pacs_anos(request: Request):
    """Retorna lista de anos disponíveis nos PACs individuais"""
    user = await get_current_user(request)
    
    # Buscar anos distintos dos PACs
    pipeline = [
        {'$match': {'ano': {'$exists': True, '$ne': None}}},
        {'$group': {'_id': '$ano'}},
        {'$sort': {'_id': -1}}
    ]
    
    result = await db.pacs.aggregate(pipeline).to_list(100)
    anos = [r['_id'] for r in result if r['_id']]
    
    # Converter para inteiros se necessário
    anos_int = []
    for ano in anos:
        try:
            anos_int.append(int(ano))
        except (ValueError, TypeError):
            pass
    
    # Garantir que 2025 e o ano atual estejam na lista
    ano_atual = datetime.now().year
    anos_base = list(range(2025, ano_atual + 1))
    
    for ano in anos_base:
        if ano not in anos_int:
            anos_int.append(ano)
    
    anos_int.sort(reverse=True)
    return {'anos': anos_int, 'ano_atual': ano_atual}

@api_router.get("/pacs", response_model=List[PAC])
async def get_pacs(request: Request, ano: str = None):
    user = await get_current_user(request)
    # Usuários padrão veem todos os PACs (mas só podem editar os seus)
    # Admin vê e pode editar todos
    
    query = {}
    if ano:
        query['ano'] = str(ano)
    
    pacs = await db.pacs.find(query, {'_id': 0}).to_list(1000)
    return [PAC(**p) for p in pacs]

@api_router.get("/pacs/paginado")
async def get_pacs_paginado(
    request: Request, 
    ano: str = None,
    page: int = 1,
    page_size: int = 20,
    search: str = None
):
    """Retorna PACs com paginação"""
    user = await get_current_user(request)
    
    query = {}
    if ano:
        query['ano'] = str(ano)
    
    if search:
        query['$or'] = [
            {'secretaria': {'$regex': search, '$options': 'i'}},
            {'secretario': {'$regex': search, '$options': 'i'}},
            {'fiscal': {'$regex': search, '$options': 'i'}}
        ]
    
    # Contar total
    total = await db.pacs.count_documents(query)
    
    # Buscar com paginação
    skip = (page - 1) * page_size
    pacs = await db.pacs.find(query, {'_id': 0}).sort('created_at', -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        'items': [PAC(**p) for p in pacs],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size
    }

@api_router.get("/pacs/stats")
async def get_pacs_stats(request: Request):
    """
    Retorna estatísticas agregadas de todos os PACs individuais,
    agrupadas por Subitem de Classificação Orçamentária.
    """
    user = await get_current_user(request)
    
    # Buscar todos os items de PAC individual
    all_items = await db.pac_items.find({}, {'_id': 0}).to_list(10000)
    
    # Agrupar por subitem_classificacao
    stats_by_subitem = {}
    total_geral = 0
    
    for item in all_items:
        subitem = item.get('subitem_classificacao', 'Não Classificado') or 'Não Classificado'
        codigo = item.get('codigo_classificacao', '') or ''
        valor = item.get('valorTotal', 0) or 0
        
        key = f"{codigo} - {subitem}" if codigo else subitem
        
        if key not in stats_by_subitem:
            stats_by_subitem[key] = {
                'subitem': subitem,
                'codigo': codigo,
                'valor_total': 0,
                'quantidade_items': 0
            }
        
        stats_by_subitem[key]['valor_total'] += valor
        stats_by_subitem[key]['quantidade_items'] += 1
        total_geral += valor
    
    stats_list = list(stats_by_subitem.values())
    stats_list.sort(key=lambda x: x['valor_total'], reverse=True)
    
    # Contar total de PACs
    total_pacs = await db.pacs.count_documents({})
    
    return {
        'stats_by_subitem': stats_list,
        'total_geral': total_geral,
        'total_items': len(all_items),
        'total_pacs': total_pacs
    }

@api_router.post("/pacs", response_model=PAC)
async def create_pac(pac_data: PACCreate, request: Request):
    user = await get_current_user(request)
    pac_id = f"pac_{uuid.uuid4().hex[:12]}"
    pac_doc = {'pac_id': pac_id, 'user_id': user.user_id, **pac_data.model_dump(), 'total_value': 0.0, 'stats': {'consumo': 0, 'consumoQtd': 0, 'permanente': 0, 'permanenteQtd': 0, 'servicos': 0, 'servicosQtd': 0, 'obras': 0, 'obrasQtd': 0}, 'created_at': datetime.now(timezone.utc), 'updated_at': datetime.now(timezone.utc)}
    await db.pacs.insert_one(pac_doc)
    return PAC(**pac_doc)

@api_router.get("/pacs/{pac_id}", response_model=PAC)
async def get_pac(pac_id: str, request: Request):
    user = await get_current_user(request)
    # CORREÇÃO 1.2: Remover filtro por user_id para permitir visualização de todos os PACs
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    return PAC(**pac)

@api_router.put("/pacs/{pac_id}", response_model=PAC)
async def update_pac(pac_id: str, pac_update: PACUpdate, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    # Apenas admin ou dono pode editar
    if not user.is_admin and pac['user_id'] != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied")
    update_data = {k: v for k, v in pac_update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc)
    await db.pacs.update_one({'pac_id': pac_id}, {'$set': update_data})
    updated_pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    return PAC(**updated_pac)

@api_router.delete("/pacs/{pac_id}")
async def delete_pac(pac_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    # Apenas admin ou dono pode excluir
    if not user.is_admin and pac['user_id'] != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied")
    await db.pacs.delete_one({'pac_id': pac_id})
    await db.pac_items.delete_many({'pac_id': pac_id})
    return {'message': 'PAC deleted'}

@api_router.get("/pacs/{pac_id}/items", response_model=List[PACItem])
async def get_pac_items(pac_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(1000)
    return [PACItem(**i) for i in items]

@api_router.post("/pacs/{pac_id}/items", response_model=PACItem)
async def create_pac_item(pac_id: str, item_data: PACItemCreate, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    # Apenas admin ou dono pode adicionar itens
    if not user.is_admin and pac['user_id'] != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied")
    item_id = f"item_{uuid.uuid4().hex[:12]}"
    valorTotal = item_data.quantidade * item_data.valorUnitario
    item_doc = {'item_id': item_id, 'pac_id': pac_id, **item_data.model_dump(), 'valorTotal': valorTotal, 'created_at': datetime.now(timezone.utc)}
    await db.pac_items.insert_one(item_doc)
    await recalculate_pac_totals(pac_id)
    return PACItem(**item_doc)

@api_router.put("/pacs/{pac_id}/items/{item_id}", response_model=PACItem)
async def update_pac_item(pac_id: str, item_id: str, item_update: PACItemUpdate, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    # Apenas admin ou dono pode editar
    if not user.is_admin and pac['user_id'] != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied")
    item = await db.pac_items.find_one({'item_id': item_id, 'pac_id': pac_id}, {'_id': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    update_data = {k: v for k, v in item_update.model_dump().items() if v is not None}
    if 'quantidade' in update_data or 'valorUnitario' in update_data:
        qtd = update_data.get('quantidade', item['quantidade'])
        val = update_data.get('valorUnitario', item['valorUnitario'])
        update_data['valorTotal'] = qtd * val
    await db.pac_items.update_one({'item_id': item_id}, {'$set': update_data})
    await recalculate_pac_totals(pac_id)
    updated_item = await db.pac_items.find_one({'item_id': item_id}, {'_id': 0})
    return PACItem(**updated_item)

@api_router.delete("/pacs/{pac_id}/items/{item_id}")
async def delete_pac_item(pac_id: str, item_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    # Apenas admin ou dono pode excluir
    if not user.is_admin and pac['user_id'] != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied")
    result = await db.pac_items.delete_one({'item_id': item_id, 'pac_id': pac_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    await recalculate_pac_totals(pac_id)
    return {'message': 'Item deleted'}

async def recalculate_pac_totals(pac_id: str):
    items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(1000)
    stats = {'totalGeral': 0, 'consumo': 0, 'consumoQtd': 0, 'permanente': 0, 'permanenteQtd': 0, 'servicos': 0, 'servicosQtd': 0, 'obras': 0, 'obrasQtd': 0}
    for item in items:
        total = item['valorTotal']
        stats['totalGeral'] += total
        if item['tipo'] == 'Material de Consumo':
            stats['consumo'] += total
            stats['consumoQtd'] += 1
        elif item['tipo'] == 'Material Permanente':
            stats['permanente'] += total
            stats['permanenteQtd'] += 1
        elif item['tipo'] == 'Serviço':
            stats['servicos'] += total
            stats['servicosQtd'] += 1
        elif item['tipo'] == 'Obras':
            stats['obras'] += total
            stats['obrasQtd'] += 1
    await db.pacs.update_one({'pac_id': pac_id}, {'$set': {'total_value': stats['totalGeral'], 'stats': stats, 'updated_at': datetime.now(timezone.utc)}})

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    user = await get_current_user(request)
    pacs = await db.pacs.find({'user_id': user.user_id}, {'_id': 0}).to_list(1000)
    stats = {'totalGeral': 0, 'consumo': {'valor': 0, 'qtd': 0}, 'permanente': {'valor': 0, 'qtd': 0}, 'servicos': {'valor': 0, 'qtd': 0}, 'obras': {'valor': 0, 'qtd': 0}, 'totalPacs': len(pacs)}
    for pac in pacs:
        stats['totalGeral'] += pac.get('total_value', 0)
        if pac.get('stats'):
            s = pac['stats']
            stats['consumo']['valor'] += s.get('consumo', 0)
            stats['consumo']['qtd'] += s.get('consumoQtd', 0)
            stats['permanente']['valor'] += s.get('permanente', 0)
            stats['permanente']['qtd'] += s.get('permanenteQtd', 0)
            stats['servicos']['valor'] += s.get('servicos', 0)
            stats['servicos']['qtd'] += s.get('servicosQtd', 0)
            stats['obras']['valor'] += s.get('obras', 0)
            stats['obras']['qtd'] += s.get('obrasQtd', 0)
    return stats

@api_router.get("/pacs/{pac_id}/export/xlsx")
async def export_xlsx(pac_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PAC 2026"
    
    # Estilos
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho
    ws.merge_cells('A1:K1')
    ws['A1'] = 'PREFEITURA MUNICIPAL DE ACAIACA - MG'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:K2')
    ws['A2'] = 'PAC ACAIACA 2026 - PLANO ANUAL DE CONTRATAÇÕES'
    ws['A2'].font = Font(name='Arial', size=14, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    
    ws.merge_cells('A3:K3')
    ws['A3'] = 'Lei Federal nº 14.133/2021'
    ws['A3'].font = Font(name='Arial', size=10, italic=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Informações do PAC
    current_row = 5
    ws[f'A{current_row}'] = 'DADOS DA SECRETARIA'
    ws[f'A{current_row}'].font = subheader_font
    ws.merge_cells(f'A{current_row}:K{current_row}')
    
    current_row += 1
    info_data = [
        ('Secretaria:', pac['secretaria']),
        ('Secretário(a):', pac['secretario']),
        ('Fiscal do Contrato:', pac['fiscal']),
        ('Telefone:', pac['telefone']),
        ('E-mail:', pac['email']),
        ('Endereço:', pac['endereco']),
        ('Ano:', pac['ano'])
    ]
    
    for label, value in info_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = bold_font
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].font = normal_font
        ws.merge_cells(f'B{current_row}:K{current_row}')
        current_row += 1
    
    # Tabela de itens
    current_row += 1
    headers = ['#', 'Tipo', 'Código', 'Descrição', 'Unidade', 'Qtd', 'Valor Unit (R$)', 'Valor Total (R$)', 'Prioridade', 'Justificativa', 'Classificação']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[current_row].height = 30
    
    # Dados dos itens
    for idx, item in enumerate(items, start=1):
        current_row += 1
        # Formatar classificação orçamentária
        classificacao_text = ''
        if item.get('codigo_classificacao'):
            classificacao_text = f"{item['codigo_classificacao']}"
            if item.get('subitem_classificacao'):
                classificacao_text += f" - {item['subitem_classificacao']}"
        
        row_data = [
            idx,
            item['tipo'],
            item['catmat'],
            item['descricao'],
            item['unidade'],
            item['quantidade'],
            item['valorUnitario'],
            item['valorTotal'],
            item['prioridade'],
            item['justificativa'],
            classificacao_text
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [1,2,5,6,9] else 'left', vertical='center', wrap_text=True)
            if col in [7, 8]:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Totais
    current_row += 1
    total = sum(item['valorTotal'] for item in items)
    
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = 'TOTAL GERAL ESTIMADO:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'A{current_row}'].fill = total_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{current_row}'].border = border
    
    ws[f'H{current_row}'] = total
    ws[f'H{current_row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'H{current_row}'].fill = total_fill
    ws[f'H{current_row}'].number_format = 'R$ #,##0.00'
    ws[f'H{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'H{current_row}'].border = border
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 30
    
    # Configurar impressão
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToWidth = 1
    ws.print_options.horizontalCentered = True
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename=PAC_{pac["secretaria"].replace(" ", "_")}_2026.xlsx'})

@api_router.get("/pacs/{pac_id}/export/pdf")
async def export_pdf(pac_id: str, request: Request, orientation: str = "landscape", assinar: bool = True, data: str = None):
    """
    Exporta PAC Individual para PDF com design profissional e paginação de 15 itens por página.
    Args:
        orientation: 'landscape' (paisagem) ou 'portrait' (retrato)
        assinar: Se True, adiciona assinatura digital
        data: Data da assinatura (formato DD/MM/YYYY HH:MM:SS) - retroativa se necessário
    """
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='PAC - PLANO ANUAL DE CONTRATAÇÕES',
        subtitulo=f'Exercício {pac.get("ano", "2026")} - Lei Federal nº 14.133/2021',
        total_pages=1
    )
    
    # Margens ajustadas para acomodar selo de assinatura na direita
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=35*mm,  # Espaço para cabeçalho oficial
        bottomMargin=25*mm,  # Espaço para rodapé oficial
        title=f'PAC {pac["secretaria"]} {pac.get("ano", "2026")}'
    )
    
    elements = []
    custom_styles, cor_primaria, cor_secundaria, cor_destaque = get_professional_styles()
    
    # Caixa de informações
    elements.append(create_info_box(pac, custom_styles, is_pac_geral=False))
    elements.append(Spacer(1, 6*mm))
    
    # Título da seção de itens
    elements.append(Paragraph('<b>DETALHAMENTO DOS ITENS DO PLANO ANUAL DE CONTRATAÇÕES</b>', custom_styles['section_header']))
    elements.append(Spacer(1, 3*mm))
    
    # Tabela de TODOS os itens (sem paginação forçada para economizar papel)
    if items:
        items_table = create_items_table_paginated(
            items, 
            custom_styles, 
            orientation, 
            start_index=1
        )
        elements.append(items_table)
    
    # Linha de total após todos os itens
    elements.append(Spacer(1, 4*mm))
    total = sum(item['valorTotal'] for item in items)
    elements.append(create_total_row(total, custom_styles, orientation))
    
    # Seção de assinaturas
    elements.extend(create_signature_section(pac, custom_styles, is_pac_geral=False))
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    # Adicionar assinatura digital ao PDF
    try:
        signed_buffer, validation_code = await add_signature_to_pdf(
            buffer, user, f"PAC - {pac['secretaria']}", pac_id, None, data
        )
        return StreamingResponse(
            signed_buffer, 
            media_type='application/pdf', 
            headers={
                'Content-Disposition': f'attachment; filename=PAC_{pac["secretaria"].replace(" ", "_")}_{pac.get("ano", "2026")}.pdf',
                'X-Validation-Code': validation_code
            }
        )
    except HTTPException:
        # Re-raise HTTPException para validação de CPF/Cargo
        raise
    except Exception as e:
        logging.error(f"Erro ao adicionar assinatura ao PDF: {e}")
        buffer.seek(0)
        return StreamingResponse(buffer, media_type='application/pdf', headers={'Content-Disposition': f'attachment; filename=PAC_{pac["secretaria"].replace(" ", "_")}_{pac.get("ano", "2026")}.pdf'})

@api_router.get("/template/download")
async def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template PAC"
    headers = ['Tipo', 'Código CATMAT/CATSER', 'Descrição', 'Unidade', 'Quantidade', 'Valor Unitário', 'Prioridade', 'Justificativa']
    ws.append(headers)
    ws.append(['Material de Consumo', '123456', 'Exemplo de item', 'Unidade', 10, 50.00, 'Alta', 'Justificativa exemplo'])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=Template_PAC.xlsx'})

@api_router.post("/pacs/{pac_id}/import")
async def import_xlsx(pac_id: str, file: UploadFile = File(...), request: Request = None):
    user = await get_current_user(request)
    pac = await db.pacs.find_one({'pac_id': pac_id, 'user_id': user.user_id})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC not found")
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active
    imported_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            item_id = f"item_{uuid.uuid4().hex[:12]}"
            quantidade = float(row[4]) if row[4] else 0
            valorUnitario = float(row[5]) if row[5] else 0
            item_doc = {'item_id': item_id, 'pac_id': pac_id, 'tipo': row[0] or 'Material de Consumo', 'catmat': row[1] or '', 'descricao': row[2] or '', 'unidade': row[3] or 'Unidade', 'quantidade': quantidade, 'valorUnitario': valorUnitario, 'valorTotal': quantidade * valorUnitario, 'prioridade': row[6] or 'Média', 'justificativa': row[7] or '', 'imagemUrl': None, 'created_at': datetime.now(timezone.utc)}
            await db.pac_items.insert_one(item_doc)
            imported_count += 1
        except Exception:
            continue
    await recalculate_pac_totals(pac_id)
    return {'message': f'{imported_count} itens importados com sucesso'}

# ===== ESTATÍSTICAS DO PAC GERAL (DASHBOARD) =====
@api_router.get("/pacs-geral/stats")
async def get_pacs_geral_stats(request: Request):
    """
    Retorna estatísticas agregadas de todos os PACs Gerais,
    agrupadas por Subitem de Classificação Orçamentária.
    """
    user = await get_current_user(request)
    
    # Buscar todos os items de PAC Geral
    all_items = await db.pac_geral_items.find({}, {'_id': 0}).to_list(10000)
    
    # Agrupar por subitem_classificacao
    stats_by_subitem = {}
    total_geral = 0
    
    for item in all_items:
        subitem = item.get('subitem_classificacao', 'Não Classificado') or 'Não Classificado'
        codigo = item.get('codigo_classificacao', '') or ''
        valor = item.get('valorTotal', 0) or 0
        
        # Criar chave composta: código + subitem
        key = f"{codigo} - {subitem}" if codigo else subitem
        
        if key not in stats_by_subitem:
            stats_by_subitem[key] = {
                'subitem': subitem,
                'codigo': codigo,
                'valor_total': 0,
                'quantidade_items': 0
            }
        
        stats_by_subitem[key]['valor_total'] += valor
        stats_by_subitem[key]['quantidade_items'] += 1
        total_geral += valor
    
    # Converter para lista e ordenar por valor
    stats_list = list(stats_by_subitem.values())
    stats_list.sort(key=lambda x: x['valor_total'], reverse=True)
    
    return {
        'stats_by_subitem': stats_list,
        'total_geral': total_geral,
        'total_items': len(all_items)
    }

# ===== ROTAS PAC GERAL =====

@api_router.get("/pacs-geral/anos")
async def get_pacs_geral_anos(request: Request):
    """Retorna lista de anos disponíveis nos PACs Gerais"""
    user = await get_current_user(request)
    
    # Buscar anos distintos dos PACs Gerais
    pipeline = [
        {'$match': {'ano': {'$exists': True, '$ne': None}}},
        {'$group': {'_id': '$ano'}},
        {'$sort': {'_id': -1}}
    ]
    
    result = await db.pacs_geral.aggregate(pipeline).to_list(100)
    anos = [r['_id'] for r in result if r['_id']]
    
    # Converter para inteiros se necessário
    anos_int = []
    for ano in anos:
        try:
            anos_int.append(int(ano))
        except (ValueError, TypeError):
            pass
    
    # Garantir que 2026 e o ano atual estejam na lista (PAC Geral começa em 2026)
    ano_atual = datetime.now().year
    anos_base = list(range(2026, ano_atual + 1))
    
    for ano in anos_base:
        if ano not in anos_int:
            anos_int.append(ano)
    
    anos_int.sort(reverse=True)
    return {'anos': anos_int, 'ano_atual': ano_atual}

@api_router.get("/pacs-geral", response_model=List[PACGeral])
async def get_pacs_geral(request: Request, ano: str = None):
    user = await get_current_user(request)
    # Todos os usuários podem ver todos os PACs Gerais
    
    query = {}
    if ano:
        query['ano'] = str(ano)
    
    pacs = await db.pacs_geral.find(query, {'_id': 0}).to_list(1000)
    return pacs

@api_router.get("/pacs-geral/paginado")
async def get_pacs_geral_paginado(
    request: Request,
    ano: str = None,
    page: int = 1,
    page_size: int = 20,
    search: str = None
):
    """Retorna PACs Gerais com paginação"""
    user = await get_current_user(request)
    
    query = {}
    if ano:
        query['ano'] = str(ano)
    
    if search:
        query['$or'] = [
            {'nome_secretaria': {'$regex': search, '$options': 'i'}},
            {'secretario': {'$regex': search, '$options': 'i'}},
            {'fiscal_contrato': {'$regex': search, '$options': 'i'}}
        ]
    
    # Contar total
    total = await db.pacs_geral.count_documents(query)
    
    # Buscar com paginação
    skip = (page - 1) * page_size
    pacs = await db.pacs_geral.find(query, {'_id': 0}).sort('created_at', -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        'items': pacs,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size
    }

@api_router.post("/pacs-geral", response_model=PACGeral)
async def create_pac_geral(pac_data: PACGeralCreate, request: Request):
    user = await get_current_user(request)
    pac_geral_id = f"pac_geral_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    pac_doc = {
        'pac_geral_id': pac_geral_id,
        'user_id': user.user_id,
        **pac_data.model_dump(),
        'created_at': now,
        'updated_at': now
    }
    await db.pacs_geral.insert_one(pac_doc)
    pac_doc.pop('_id', None)
    return PACGeral(**pac_doc)

@api_router.get("/pacs-geral/{pac_geral_id}", response_model=PACGeral)
async def get_pac_geral(pac_geral_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    # Admins podem ver todos, usuários padrão só seus próprios
    if not user.is_admin and pac['user_id'] != user.user_id:
        # Permitir visualização mas não edição
        pass
    return PACGeral(**pac)

@api_router.put("/pacs-geral/{pac_geral_id}", response_model=PACGeral)
async def update_pac_geral(pac_geral_id: str, pac_data: PACGeralUpdate, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    # Todos os usuários podem editar qualquer PAC Geral
    
    update_data = {k: v for k, v in pac_data.model_dump().items() if v is not None}
    if update_data:
        update_data['updated_at'] = datetime.now(timezone.utc)
        await db.pacs_geral.update_one({'pac_geral_id': pac_geral_id}, {'$set': update_data})
    
    updated_pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    return PACGeral(**updated_pac)

@api_router.delete("/pacs-geral/{pac_geral_id}")
async def delete_pac_geral(pac_geral_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    # Apenas administradores podem excluir PAC Geral
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Only administrators can delete PAC Geral")
    
    await db.pacs_geral.delete_one({'pac_geral_id': pac_geral_id})
    await db.pac_geral_items.delete_many({'pac_geral_id': pac_geral_id})
    return {'message': 'PAC Geral deleted successfully'}

# ===== ROTAS ITEMS PAC GERAL =====
@api_router.get("/pacs-geral/{pac_geral_id}/items", response_model=List[PACGeralItem])
async def get_pac_geral_items(pac_geral_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(1000)
    return items

@api_router.post("/pacs-geral/{pac_geral_id}/items", response_model=PACGeralItem)
async def create_pac_geral_item(pac_geral_id: str, item_data: PACGeralItemCreate, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    if not user.is_admin and pac['user_id'] != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    item_id = f"item_{uuid.uuid4().hex[:12]}"
    
    # Calcular quantidade total
    quantidade_total = (
        item_data.qtd_ad + item_data.qtd_fa + item_data.qtd_sa + 
        item_data.qtd_se + item_data.qtd_as + item_data.qtd_ag + 
        item_data.qtd_ob + item_data.qtd_tr + item_data.qtd_cul
    )
    
    valor_total = quantidade_total * item_data.valorUnitario
    
    item_doc = {
        'item_id': item_id,
        'pac_geral_id': pac_geral_id,
        **item_data.model_dump(),
        'quantidade_total': quantidade_total,
        'valorTotal': valor_total,
        'created_at': datetime.now(timezone.utc)
    }
    
    await db.pac_geral_items.insert_one(item_doc)
    item_doc.pop('_id', None)
    return PACGeralItem(**item_doc)

@api_router.put("/pacs-geral/{pac_geral_id}/items/{item_id}", response_model=PACGeralItem)
async def update_pac_geral_item(pac_geral_id: str, item_id: str, item_data: PACGeralItemUpdate, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    # Todos os usuários podem editar items de qualquer PAC Geral
    
    item = await db.pac_geral_items.find_one({'item_id': item_id, 'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    update_data = {k: v for k, v in item_data.model_dump().items() if v is not None}
    
    # Recalcular totais se alguma quantidade ou valor mudou
    if any(k.startswith('qtd_') for k in update_data.keys()) or 'valorUnitario' in update_data:
        updated_item = {**item, **update_data}
        quantidade_total = (
            updated_item.get('qtd_ad', 0) + updated_item.get('qtd_fa', 0) + 
            updated_item.get('qtd_sa', 0) + updated_item.get('qtd_se', 0) + 
            updated_item.get('qtd_as', 0) + updated_item.get('qtd_ag', 0) + 
            updated_item.get('qtd_ob', 0) + updated_item.get('qtd_tr', 0) + 
            updated_item.get('qtd_cul', 0)
        )
        update_data['quantidade_total'] = quantidade_total
        update_data['valorTotal'] = quantidade_total * updated_item.get('valorUnitario', 0)
    
    if update_data:
        await db.pac_geral_items.update_one(
            {'item_id': item_id, 'pac_geral_id': pac_geral_id},
            {'$set': update_data}
        )
    
    updated_item = await db.pac_geral_items.find_one({'item_id': item_id}, {'_id': 0})
    return PACGeralItem(**updated_item)

@api_router.delete("/pacs-geral/{pac_geral_id}/items/{item_id}")
async def delete_pac_geral_item(pac_geral_id: str, item_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    # Apenas administradores podem excluir items
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Only administrators can delete items")
    
    result = await db.pac_geral_items.delete_one({'item_id': item_id, 'pac_geral_id': pac_geral_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    return {'message': 'Item deleted successfully'}


# ===== EXPORTAÇÃO PAC GERAL =====
@api_router.get("/pacs-geral/{pac_geral_id}/export/xlsx")
async def export_pac_geral_xlsx(pac_geral_id: str, request: Request):
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PAC Geral"
    
    # Estilos
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    total_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    current_row = 1
    
    # Cabeçalho
    ws.merge_cells(f'A{current_row}:M{current_row}')
    ws[f'A{current_row}'] = 'PREFEITURA MUNICIPAL DE ACAIACA - MG'
    ws[f'A{current_row}'].font = Font(name='Arial', size=16, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:M{current_row}')
    ws[f'A{current_row}'] = 'PAC ACAIACA 2026 - PLANO ANUAL DE CONTRATAÇÕES COMPARTILHADO'
    ws[f'A{current_row}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:M{current_row}')
    ws[f'A{current_row}'] = 'Lei Federal nº 14.133/2021'
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, italic=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2
    
    # Dados da Secretaria
    ws[f'A{current_row}'] = 'Secretaria Responsável:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = pac['nome_secretaria']
    current_row += 1
    
    ws[f'A{current_row}'] = 'Secretário:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = pac['secretario']
    current_row += 1
    
    ws[f'A{current_row}'] = 'Telefone:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = pac['telefone']
    ws[f'E{current_row}'] = 'E-mail:'
    ws[f'E{current_row}'].font = bold_font
    ws[f'F{current_row}'] = pac['email']
    current_row += 1
    
    ws[f'A{current_row}'] = 'Secretarias Participantes:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = ', '.join(pac['secretarias_selecionadas'])
    current_row += 2
    
    # Tabela de itens
    headers = ['#', 'Código', 'Descrição', 'Und', 'AD', 'FA', 'SA', 'SE', 'AS', 'AG', 'OB', 'TR', 'CUL', 'Total', 'Valor Unit', 'Valor Total', 'Prior', 'Justificativa', 'Classificação']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Dados dos itens
    for idx, item in enumerate(items, start=1):
        classificacao_text = ''
        if item.get('codigo_classificacao'):
            classificacao_text = f"{item['codigo_classificacao']}"
            if item.get('subitem_classificacao'):
                classificacao_text += f" - {item['subitem_classificacao']}"
        
        row_data = [
            idx,
            item['catmat'],
            item['descricao'],
            item['unidade'],
            item.get('qtd_ad', 0),
            item.get('qtd_fa', 0),
            item.get('qtd_sa', 0),
            item.get('qtd_se', 0),
            item.get('qtd_as', 0),
            item.get('qtd_ag', 0),
            item.get('qtd_ob', 0),
            item.get('qtd_tr', 0),
            item.get('qtd_cul', 0),
            item['quantidade_total'],
            item['valorUnitario'],
            item['valorTotal'],
            item['prioridade'],
            item.get('justificativa', ''),
            classificacao_text
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center' if col in [1,2,4,5,6,7,8,9,10,11,12,13,14,17] else 'left',
                vertical='center',
                wrap_text=True
            )
            if col in [15, 16]:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        current_row += 1
    
    # Totais
    total = sum(item['valorTotal'] for item in items)
    
    ws.merge_cells(f'A{current_row}:O{current_row}')
    ws[f'A{current_row}'] = 'TOTAL GERAL ESTIMADO:'
    ws[f'A{current_row}'].font = bold_font
    ws[f'A{current_row}'].fill = total_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{current_row}'].border = border
    
    ws[f'P{current_row}'] = total
    ws[f'P{current_row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'P{current_row}'].fill = total_fill
    ws[f'P{current_row}'].number_format = 'R$ #,##0.00'
    ws[f'P{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'P{current_row}'].border = border
    
    # Assinaturas
    current_row += 3
    ws[f'A{current_row}'] = '_' * 40
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'H{current_row}'] = '_' * 40
    ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    ws[f'A{current_row}'] = 'Fiscal do Contrato'
    ws[f'A{current_row}'].font = bold_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'H{current_row}'] = 'Gestor do Contrato'
    ws[f'H{current_row}'].font = bold_font
    ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 8
    for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 8
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['Q'].width = 8
    ws.column_dimensions['R'].width = 35  # Justificativa
    ws.column_dimensions['S'].width = 30  # Classificação
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=PAC_Geral_{pac['nome_secretaria']}.xlsx"}
    )

# ===== IMPORTAÇÃO DE ARQUIVO PARA PAC GERAL =====
@api_router.post("/pacs-geral/{pac_geral_id}/import")
async def import_pac_geral_items(
    pac_geral_id: str, 
    file: UploadFile = File(...), 
    request: Request = None
):
    """
    Importa itens para o PAC Geral a partir de um arquivo.
    Formatos suportados: CSV, XLSX, JSON
    
    Estrutura esperada:
    - codigo (Catmat/Catser)
    - descricao
    - unidade
    - quantidade_total (ou qtd_total)
    - valor_unitario (ou valorUnitario)
    - prioridade
    - classificacao (opcional)
    - justificativa (opcional)
    """
    user = await get_current_user(request)
    
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    # Ler o arquivo
    content = await file.read()
    filename = file.filename.lower()
    
    items_to_import = []
    
    try:
        if filename.endswith('.csv'):
            # Processar CSV
            import csv
            import io
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded), delimiter=';')
            for row in reader:
                items_to_import.append(row)
                
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Processar Excel
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(content))
            ws = wb.active
            
            # Pegar cabeçalhos da primeira linha
            headers = [cell.value for cell in ws[1] if cell.value]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    item = dict(zip(headers, row))
                    items_to_import.append(item)
                    
        elif filename.endswith('.json'):
            # Processar JSON
            import json
            data = json.loads(content.decode('utf-8'))
            if isinstance(data, list):
                items_to_import = data
            elif isinstance(data, dict) and 'items' in data:
                items_to_import = data['items']
            else:
                raise HTTPException(status_code=400, detail="JSON deve conter uma lista de itens ou objeto com chave 'items'")
        else:
            raise HTTPException(status_code=400, detail="Formato de arquivo não suportado. Use CSV, XLSX ou JSON")
        
        # Processar e inserir itens
        imported_count = 0
        for item_data in items_to_import:
            # Normalizar nomes de campos
            codigo = item_data.get('codigo') or item_data.get('catmat') or item_data.get('Código') or ''
            descricao = item_data.get('descricao') or item_data.get('Descrição') or item_data.get('descricão') or ''
            unidade = item_data.get('unidade') or item_data.get('Unidade') or item_data.get('und') or 'Unidade'
            
            qtd_total = item_data.get('quantidade_total') or item_data.get('qtd_total') or item_data.get('Qtd Total') or item_data.get('quantidade') or 0
            valor_unit = item_data.get('valor_unitario') or item_data.get('valorUnitario') or item_data.get('Valor Unit') or item_data.get('valor') or 0
            
            prioridade = item_data.get('prioridade') or item_data.get('Prioridade') or item_data.get('prior') or 'Média'
            classificacao = item_data.get('classificacao') or item_data.get('Classificação') or item_data.get('codigo_classificacao') or ''
            justificativa = item_data.get('justificativa') or item_data.get('Justificativa') or ''
            
            # Converter valores numéricos
            try:
                qtd_total = float(str(qtd_total).replace(',', '.').replace('R$', '').strip()) if qtd_total else 0
                valor_unit = float(str(valor_unit).replace(',', '.').replace('R$', '').strip()) if valor_unit else 0
            except:
                qtd_total = 0
                valor_unit = 0
            
            if not descricao:
                continue
            
            item_id = f"item_{uuid.uuid4().hex[:12]}"
            item_doc = {
                'item_id': item_id,
                'pac_geral_id': pac_geral_id,
                'catmat': str(codigo)[:20],
                'descricao': str(descricao)[:500],
                'unidade': str(unidade)[:20],
                'qtd_ad': 0, 'qtd_fa': 0, 'qtd_sa': 0, 'qtd_se': 0,
                'qtd_as': 0, 'qtd_ag': 0, 'qtd_ob': 0, 'qtd_tr': 0, 'qtd_cul': 0,
                'quantidade_total': qtd_total,
                'valorUnitario': valor_unit,
                'valorTotal': qtd_total * valor_unit,
                'prioridade': str(prioridade)[:20],
                'justificativa': str(justificativa)[:500],
                'codigo_classificacao': str(classificacao)[:20] if classificacao else '',
                'subitem_classificacao': '',
                'created_at': datetime.now(timezone.utc)
            }
            
            await db.pac_geral_items.insert_one(item_doc)
            imported_count += 1
        
        return {
            'message': f'{imported_count} itens importados com sucesso',
            'total_items': imported_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

@api_router.get("/pacs-geral/{pac_geral_id}/export/pdf")
async def export_pac_geral_pdf(pac_geral_id: str, request: Request, orientation: str = "landscape", assinar: bool = True, data: str = None):
    """
    Exporta o PAC Geral para PDF com design profissional e paginação de 15 itens por página.
    Args:
        orientation: 'landscape' (paisagem) ou 'portrait' (retrato)
        assinar: Se True, adiciona assinatura digital
        data: Data da assinatura (formato DD/MM/YYYY HH:MM:SS) - retroativa se necessário
    """
    user = await get_current_user(request)
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral not found")
    
    items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='PAC GERAL - PLANO ANUAL DE CONTRATAÇÕES CONSOLIDADO',
        subtitulo=f'Exercício {pac.get("ano", "2026")} - Lei Federal nº 14.133/2021',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=35*mm,  # Espaço para cabeçalho oficial
        bottomMargin=25*mm,  # Espaço para rodapé oficial
        title=f'PAC Geral {pac.get("nome_secretaria", "")} {pac.get("ano", "2026")}'
    )
    
    elements = []
    custom_styles, cor_primaria, cor_secundaria, cor_destaque = get_professional_styles()
    base_styles = getSampleStyleSheet()
    
    # Caixa de informações
    elements.append(create_info_box(pac, custom_styles, is_pac_geral=True))
    
    # Secretarias participantes
    secretarias = pac.get('secretarias_selecionadas', [])
    if secretarias:
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph(f'<b>Secretarias Participantes:</b> {", ".join(secretarias)}', custom_styles['small']))
    
    elements.append(Spacer(1, 6*mm))
    
    # Título da seção de itens
    elements.append(Paragraph('<b>DETALHAMENTO DOS ITENS DO PLANO ANUAL DE CONTRATAÇÕES</b>', custom_styles['section_header']))
    elements.append(Spacer(1, 3*mm))
    
    # Tabela de TODOS os itens (sem paginação forçada para economizar papel)
    if items:
        # Criar tabela para todos os itens
        table_data = [['#', 'Código\nCATMAT', 'Descrição do Objeto', 'Justificativa', 'Und', 'Qtd', 'Valor\nUnitário', 'Valor\nTotal', 'Prior.', 'Classificação\nOrçamentária']]
        
        for idx, item in enumerate(items, start=1):
            classificacao_text = ''
            if item.get('codigo_classificacao'):
                classificacao_text = f"{item['codigo_classificacao']}"
                if item.get('subitem_classificacao'):
                    classificacao_text += f"\n{item['subitem_classificacao']}"
            
            justificativa = item.get('justificativa', '') or 'N/I'
            descricao = item.get('descricao', '')
            
            table_data.append([
                str(idx),
                item.get('catmat', ''),
                Paragraph(f"<font size=7>{descricao}</font>", custom_styles['table_cell']),
                Paragraph(f"<font size=6>{justificativa}</font>", custom_styles['table_cell']),
                item.get('unidade', ''),
                str(int(item.get('quantidade_total', 0))),
                f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {item.get('valorTotal', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                item.get('prioridade', ''),
                Paragraph(f"<font size=6>{classificacao_text}</font>", custom_styles['table_cell'])
            ])
        
        if orientation.lower() == 'portrait':
            col_widths = [0.5*cm, 1.2*cm, 3*cm, 2.5*cm, 0.8*cm, 0.8*cm, 1.4*cm, 1.5*cm, 0.8*cm, 2.2*cm]
        else:
            col_widths = [0.5*cm, 1.4*cm, 5*cm, 4.5*cm, 1*cm, 0.9*cm, 1.6*cm, 1.8*cm, 1*cm, 4.5*cm]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), cor_primaria),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (5, -1), 'CENTER'),
            ('ALIGN', (6, 1), (7, -1), 'RIGHT'),
            ('ALIGN', (8, 1), (8, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#D6EAF8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('BOX', (0, 0), (-1, -1), 1, cor_primaria),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
    
    # Linha de total após todos os itens
    elements.append(Spacer(1, 4*mm))
    total = sum(item.get('valorTotal', 0) for item in items)
    elements.append(create_total_row(total, custom_styles, orientation))
    
    # Seção de assinaturas
    elements.extend(create_signature_section(pac, custom_styles, is_pac_geral=True))
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    # Adicionar assinatura digital ao PDF
    try:
        signed_buffer, validation_code = await add_signature_to_pdf(
            buffer, user, f"PAC Geral - {pac.get('nome_secretaria', '')}", pac_geral_id, None, data
        )
        return StreamingResponse(
            signed_buffer, 
            media_type='application/pdf', 
            headers={
                'Content-Disposition': f'attachment; filename=PAC_Geral_{pac.get("nome_secretaria", "").replace(" ", "_")}_{pac.get("ano", "2026")}.pdf',
                'X-Validation-Code': validation_code
            }
        )
    except HTTPException:
        # Re-raise HTTPException para validação de CPF/Cargo
        raise
    except Exception as e:
        logging.error(f"Erro ao adicionar assinatura ao PDF: {e}")
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=PAC_Geral_{pac.get('nome_secretaria', '').replace(' ', '_')}_{pac.get('ano', '2026')}.pdf"}
        )


# ============ ROTAS PAC GERAL OBRAS E SERVIÇOS ============

@api_router.get("/classificacao/obras-servicos")
async def get_classificacao_obras_servicos():
    """Retorna códigos de classificação para Obras e Serviços (Lei 14.133/2021 + Portaria 448)"""
    return CLASSIFICACAO_OBRAS_SERVICOS

@api_router.get("/pacs-geral-obras", response_model=List[PACGeralObras])
async def get_pacs_geral_obras(request: Request, ano: str = None):
    """Lista todos os PAC Geral Obras do usuário"""
    user = await get_current_user(request)
    query = {'user_id': user.user_id} if not user.is_admin else {}
    if ano:
        query['ano'] = ano
    pacs = await db.pacs_geral_obras.find(query, {'_id': 0}).to_list(100)
    return [PACGeralObras(**p) for p in pacs]

@api_router.post("/pacs-geral-obras", response_model=PACGeralObras)
async def create_pac_geral_obras(pac_data: PACGeralObrasCreate, request: Request):
    """Cria um novo PAC Geral Obras"""
    user = await get_current_user(request)
    pac_obras_id = f"pacobras_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    pac_doc = {
        'pac_obras_id': pac_obras_id,
        'user_id': user.user_id,
        **pac_data.model_dump(),
        'created_at': now,
        'updated_at': now
    }
    await db.pacs_geral_obras.insert_one(pac_doc)
    return PACGeralObras(**pac_doc)

@api_router.get("/pacs-geral-obras/{pac_obras_id}", response_model=PACGeralObras)
async def get_pac_geral_obras(pac_obras_id: str, request: Request):
    """Obtém um PAC Geral Obras específico"""
    user = await get_current_user(request)
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras not found")
    return PACGeralObras(**pac)

@api_router.put("/pacs-geral-obras/{pac_obras_id}", response_model=PACGeralObras)
async def update_pac_geral_obras(pac_obras_id: str, pac_data: PACGeralObrasUpdate, request: Request):
    """Atualiza um PAC Geral Obras"""
    user = await get_current_user(request)
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras not found")
    update_data = {k: v for k, v in pac_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc)
    await db.pacs_geral_obras.update_one({'pac_obras_id': pac_obras_id}, {'$set': update_data})
    updated = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    return PACGeralObras(**updated)

@api_router.delete("/pacs-geral-obras/{pac_obras_id}")
async def delete_pac_geral_obras(pac_obras_id: str, request: Request):
    """Exclui um PAC Geral Obras e seus itens"""
    user = await get_current_user(request)
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras not found")
    await db.pac_geral_obras_items.delete_many({'pac_obras_id': pac_obras_id})
    await db.pacs_geral_obras.delete_one({'pac_obras_id': pac_obras_id})
    return {'message': 'PAC Geral Obras excluído com sucesso'}

# ===== ROTAS ITEMS PAC GERAL OBRAS =====
@api_router.get("/pacs-geral-obras/{pac_obras_id}/items", response_model=List[PACGeralObrasItem])
async def get_pac_geral_obras_items(pac_obras_id: str, request: Request):
    """Lista itens de um PAC Geral Obras"""
    user = await get_current_user(request)
    items = await db.pac_geral_obras_items.find({'pac_obras_id': pac_obras_id}, {'_id': 0}).to_list(1000)
    return [PACGeralObrasItem(**item) for item in items]

@api_router.post("/pacs-geral-obras/{pac_obras_id}/items", response_model=PACGeralObrasItem)
async def create_pac_geral_obras_item(pac_obras_id: str, item_data: PACGeralObrasItemCreate, request: Request):
    """Adiciona um item ao PAC Geral Obras"""
    user = await get_current_user(request)
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras not found")
    
    item_id = f"itemobras_{uuid.uuid4().hex[:12]}"
    item_dict = item_data.model_dump()
    
    # Calcular quantidade total
    quantidade_total = sum([
        item_dict.get('qtd_ad', 0), item_dict.get('qtd_fa', 0), item_dict.get('qtd_sa', 0),
        item_dict.get('qtd_se', 0), item_dict.get('qtd_as', 0), item_dict.get('qtd_ag', 0),
        item_dict.get('qtd_ob', 0), item_dict.get('qtd_tr', 0), item_dict.get('qtd_cul', 0)
    ])
    
    item_doc = {
        'item_id': item_id,
        'pac_obras_id': pac_obras_id,
        **item_dict,
        'quantidade_total': quantidade_total,
        'valorTotal': quantidade_total * item_dict['valorUnitario'],
        'created_at': datetime.now(timezone.utc)
    }
    await db.pac_geral_obras_items.insert_one(item_doc)
    return PACGeralObrasItem(**item_doc)

@api_router.put("/pacs-geral-obras/{pac_obras_id}/items/{item_id}", response_model=PACGeralObrasItem)
async def update_pac_geral_obras_item(pac_obras_id: str, item_id: str, item_data: PACGeralObrasItemUpdate, request: Request):
    """Atualiza um item do PAC Geral Obras"""
    user = await get_current_user(request)
    item = await db.pac_geral_obras_items.find_one({'item_id': item_id, 'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    update_data = {k: v for k, v in item_data.model_dump().items() if v is not None}
    
    # Recalcular totais se necessário
    merged = {**item, **update_data}
    quantidade_total = sum([
        merged.get('qtd_ad', 0), merged.get('qtd_fa', 0), merged.get('qtd_sa', 0),
        merged.get('qtd_se', 0), merged.get('qtd_as', 0), merged.get('qtd_ag', 0),
        merged.get('qtd_ob', 0), merged.get('qtd_tr', 0), merged.get('qtd_cul', 0)
    ])
    update_data['quantidade_total'] = quantidade_total
    update_data['valorTotal'] = quantidade_total * merged.get('valorUnitario', 0)
    
    await db.pac_geral_obras_items.update_one({'item_id': item_id}, {'$set': update_data})
    updated = await db.pac_geral_obras_items.find_one({'item_id': item_id}, {'_id': 0})
    return PACGeralObrasItem(**updated)

@api_router.delete("/pacs-geral-obras/{pac_obras_id}/items/{item_id}")
async def delete_pac_geral_obras_item(pac_obras_id: str, item_id: str, request: Request):
    """Exclui um item do PAC Geral Obras"""
    user = await get_current_user(request)
    result = await db.pac_geral_obras_items.delete_one({'item_id': item_id, 'pac_obras_id': pac_obras_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {'message': 'Item excluído com sucesso'}

# ===== EXPORTAÇÃO PDF DO PAC GERAL OBRAS =====
@api_router.get("/pacs-geral-obras/{pac_obras_id}/export/pdf")
async def export_pac_geral_obras_pdf(pac_obras_id: str, request: Request, orientation: str = "landscape", assinar: bool = True, data: str = None):
    """
    Exporta PAC Geral Obras e Serviços para PDF
    Classificação conforme Lei 14.133/2021 e Portaria 448/ME
    Args:
        assinar: Se True, adiciona assinatura digital
        data: Data da assinatura (formato DD/MM/YYYY HH:MM:SS) - retroativa se necessário
    """
    user = await get_current_user(request)
    
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras não encontrado")
    
    items = await db.pac_geral_obras_items.find({'pac_obras_id': pac_obras_id}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = landscape(A4) if orientation.lower() == 'landscape' else A4
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='PLANO ANUAL DE CONTRATAÇÕES - OBRAS E SERVIÇOS DE ENGENHARIA',
        subtitulo='Lei 14.133/2021 - Nova Lei de Licitações | Portaria 448/ME',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=35*mm,  # Espaço para cabeçalho oficial
        bottomMargin=25*mm,  # Espaço para rodapé oficial
        title=f'PAC Obras - {pac["nome_secretaria"]} {pac.get("ano", "2026")}'
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos
    titulo_style = ParagraphStyle('TituloPACObras', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#1565C0'), alignment=TA_CENTER, spaceAfter=4*mm)
    subtitulo_style = ParagraphStyle('SubtituloPACObras', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#1976D2'), spaceBefore=6*mm, spaceAfter=3*mm)
    
    # ===== DADOS DA SECRETARIA =====
    elements.append(Paragraph('IDENTIFICAÇÃO', subtitulo_style))
    
    sec_data = [
        ['Secretaria:', pac.get('nome_secretaria', '-')],
        ['Secretário(a):', pac.get('secretario', '-')],
        ['Fiscal do Contrato:', pac.get('fiscal_contrato', '-') or '-'],
        ['Telefone:', pac.get('telefone', '-')],
        ['E-mail:', pac.get('email', '-')],
        ['Endereço:', pac.get('endereco', '-')],
        ['Ano:', pac.get('ano', '2026')],
        ['Tipo:', pac.get('tipo_contratacao', 'OBRAS')]
    ]
    
    sec_table = Table(sec_data, colWidths=[40*mm, 100*mm])
    sec_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#90CAF9')),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(sec_table)
    
    # ===== ITENS =====
    if items:
        elements.append(Paragraph('DETALHAMENTO DOS ITENS', subtitulo_style))
        
        # Secretarias selecionadas
        secs = pac.get('secretarias_selecionadas', [])
        
        # Cabeçalho dinâmico
        header = ['#', 'CATSER', 'Descrição', 'Classif.', 'Un']
        sec_labels = {'AD': 'Admin', 'FA': 'Faz', 'SA': 'Saúde', 'SE': 'Educ', 'AS': 'Assist', 'AG': 'Agric', 'OB': 'Obras', 'TR': 'Transp', 'CUL': 'Cult'}
        for s in secs:
            header.append(sec_labels.get(s, s))
        header.extend(['Total', 'Unit.', 'TOTAL'])
        
        items_data = [header]
        
        for idx, item in enumerate(items, 1):
            row = [
                str(idx),
                item.get('catmat', '-'),
                Paragraph(item.get('descricao', '-')[:40], ParagraphStyle('DescItem', fontSize=6)),
                item.get('codigo_classificacao', '-'),
                item.get('unidade', '-')
            ]
            
            # Quantidades por secretaria
            for s in secs:
                qtd = item.get(f'qtd_{s.lower()}', 0)
                row.append(str(int(qtd)) if qtd else '-')
            
            row.extend([
                str(int(item.get('quantidade_total', 0))),
                f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {item.get('valorTotal', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
            items_data.append(row)
        
        # Calcular larguras das colunas
        num_secs = len(secs)
        col_widths = [8*mm, 18*mm, 50*mm, 18*mm, 12*mm]  # Fixas
        col_widths.extend([12*mm] * num_secs)  # Secretarias
        col_widths.extend([15*mm, 22*mm, 28*mm])  # Total, Unit, TOTAL
        
        items_table = Table(items_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#90CAF9')),
            ('PADDING', (0, 0), (-1, -1), 2),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
        ]))
        elements.append(items_table)
        
        # Total
        total = sum(item.get('valorTotal', 0) for item in items)
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph(f"<b>VALOR TOTAL: R$ {total:,.2f}</b>".replace(',', 'X').replace('.', ',').replace('X', '.'), ParagraphStyle('TotalPAC', fontSize=11, alignment=TA_RIGHT, textColor=colors.HexColor('#1B5E20'))))
    
    # ===== ASSINATURA =====
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph('_' * 50, ParagraphStyle('Linha', alignment=TA_CENTER)))
    elements.append(Paragraph(f'<b>{pac.get("secretario", "Secretário(a)")}</b>', ParagraphStyle('Assinatura', alignment=TA_CENTER, fontSize=9)))
    elements.append(Paragraph(f'{pac.get("nome_secretaria", "")}', ParagraphStyle('Cargo', alignment=TA_CENTER, fontSize=7, textColor=colors.gray)))
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    filename = f"PAC_Obras_{pac['nome_secretaria'].replace(' ', '_')}_{pac.get('ano', '2026')}.pdf"
    
    # Adicionar assinatura digital ao PDF
    try:
        signed_buffer, validation_code = await add_signature_to_pdf(
            buffer, user, f"PAC Obras - {pac['nome_secretaria']}", pac_obras_id, None, data
        )
        return StreamingResponse(
            signed_buffer,
            media_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'X-Validation-Code': validation_code
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Erro ao adicionar assinatura ao PDF de PAC Obras: {e}")
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )


# ============ ROTAS DE GESTÃO PROCESSUAL ============

@api_router.get("/processos/anos")
async def get_processos_anos(request: Request):
    """Retorna lista de anos disponíveis nos processos"""
    user = await get_current_user(request)
    
    # Como muitos processos não têm o campo ano, extrair do numero_processo
    processos = await db.processos.find({}, {'numero_processo': 1, 'ano': 1}).to_list(1000)
    
    anos = set()
    for p in processos:
        if p.get('ano'):
            anos.add(int(p['ano']))
        else:
            # Extrair do numero_processo (ex: PRC - 0006/2025)
            numero = p.get('numero_processo', '')
            match = re.search(r'/(\d{4})$', numero)
            if match:
                anos.add(int(match.group(1)))
    
    # Garantir que o ano atual esteja na lista
    ano_atual = datetime.now().year
    anos.add(ano_atual)
    
    anos_list = sorted(list(anos), reverse=True)
    
    # Definir o ano padrão como o ano corrente (se tiver dados) ou o mais recente com dados
    ano_padrao = ano_atual
    if ano_atual not in anos and anos_list:
        ano_padrao = anos_list[0]  # Ano mais recente com dados
    
    return {'anos': anos_list, 'ano_atual': ano_padrao}


@api_router.get("/processos", response_model=List[Processo])
async def get_processos(request: Request, ano: int = None):
    """Lista todos os processos, opcionalmente filtrados por ano"""
    user = await get_current_user(request)
    
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    # Fix null string values in datetime fields e extrair ano do numero_processo
    for p in processos:
        for field in ['data_inicio', 'data_autuacao', 'data_contrato']:
            if p.get(field) == 'null' or p.get(field) == '':
                p[field] = None
        
        # Extrair ano do numero_processo se não existir (ex: PRC - 0006/2025)
        if not p.get('ano'):
            numero = p.get('numero_processo', '')
            match = re.search(r'/(\d{4})$', numero)
            if match:
                p['ano'] = int(match.group(1))
            else:
                p['ano'] = datetime.now().year  # Ano atual como padrão
    
    # Aplicar filtro de ano após extração
    if ano:
        processos = [p for p in processos if p.get('ano') == ano]
    
    return [Processo(**p) for p in processos]

@api_router.get("/processos/paginado")
async def get_processos_paginado(
    request: Request, 
    ano: int = None,
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    status: str = None,
    modalidade: str = None
):
    """Lista processos com paginação e filtros no backend"""
    user = await get_current_user(request)
    
    # Buscar todos os processos
    processos = await db.processos.find({}, {'_id': 0}).to_list(5000)
    
    # Fix null values e extrair ano
    for p in processos:
        for field in ['data_inicio', 'data_autuacao', 'data_contrato']:
            if p.get(field) == 'null' or p.get(field) == '':
                p[field] = None
        
        if not p.get('ano'):
            numero = p.get('numero_processo', '')
            match = re.search(r'/(\d{4})$', numero)
            if match:
                p['ano'] = int(match.group(1))
            else:
                p['ano'] = datetime.now().year  # Ano atual
    
    # Aplicar filtros
    if ano:
        processos = [p for p in processos if p.get('ano') == ano]
    
    if status:
        processos = [p for p in processos if p.get('status') == status]
    
    if modalidade:
        processos = [p for p in processos if p.get('modalidade') == modalidade]
    
    if search:
        search_lower = search.lower()
        processos = [p for p in processos if 
            search_lower in (p.get('numero_processo', '') or '').lower() or
            search_lower in (p.get('objeto', '') or '').lower() or
            search_lower in (p.get('secretaria', '') or '').lower() or
            search_lower in (p.get('responsavel', '') or '').lower()
        ]
    
    # Calcular paginação
    total = len(processos)
    total_pages = (total + page_size - 1) // page_size
    
    # Aplicar paginação
    start = (page - 1) * page_size
    end = start + page_size
    paginated_items = processos[start:end]
    
    return {
        'items': [Processo(**p).model_dump() for p in paginated_items],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages
    }

@api_router.get("/processos/stats")
async def get_processos_stats(request: Request, data_inicio: str = None, data_fim: str = None):
    """Estatísticas avançadas dos processos para dashboard"""
    user = await get_current_user(request)
    
    # Query base
    query = {}
    
    # Filtro por período
    if data_inicio or data_fim:
        date_filter = {}
        if data_inicio:
            date_filter['$gte'] = data_inicio
        if data_fim:
            date_filter['$lte'] = data_fim
        if date_filter:
            query['data_inicio'] = date_filter
    
    processos = await db.processos.find(query, {'_id': 0}).to_list(1000)
    
    # Filtrar processos com dados válidos
    processos_validos = [p for p in processos if p.get('numero_processo') and p.get('numero_processo') != 'Processo']
    
    # 1. Estatísticas por Status
    stats_by_status = {}
    for p in processos_validos:
        status = p.get('status', 'Não Definido') or 'Não Definido'
        if status not in stats_by_status:
            stats_by_status[status] = {'status': status, 'quantidade': 0}
        stats_by_status[status]['quantidade'] += 1
    
    # 2. Estatísticas por Modalidade
    stats_by_modalidade = {}
    for p in processos_validos:
        modalidade = p.get('modalidade', 'Não Definido') or 'Não Definido'
        if modalidade not in stats_by_modalidade:
            stats_by_modalidade[modalidade] = {'modalidade': modalidade, 'quantidade': 0, 'concluidos': 0}
        stats_by_modalidade[modalidade]['quantidade'] += 1
        if p.get('status') == 'Concluído':
            stats_by_modalidade[modalidade]['concluidos'] += 1
    
    # 3. Tempo médio de finalização (em dias)
    tempos_finalizacao = []
    tempos_por_mes = {}
    tempos_por_modalidade = {}
    
    for p in processos_validos:
        if p.get('status') == 'Concluído' and p.get('data_inicio') and p.get('data_contrato'):
            try:
                data_inicio_str = p.get('data_inicio')
                data_contrato_str = p.get('data_contrato')
                
                # Parse das datas
                if isinstance(data_inicio_str, str) and data_inicio_str not in ['None', 'null', '']:
                    if 'T' in data_inicio_str:
                        data_inicio_dt = datetime.fromisoformat(data_inicio_str.replace('Z', '+00:00'))
                    else:
                        data_inicio_dt = datetime.strptime(data_inicio_str.split(' ')[0], '%Y-%m-%d')
                    
                    if isinstance(data_contrato_str, str) and data_contrato_str not in ['None', 'null', '']:
                        if 'T' in data_contrato_str:
                            data_contrato_dt = datetime.fromisoformat(data_contrato_str.replace('Z', '+00:00'))
                        else:
                            data_contrato_dt = datetime.strptime(data_contrato_str.split(' ')[0], '%Y-%m-%d')
                        
                        dias = (data_contrato_dt - data_inicio_dt).days
                        if dias >= 0 and dias < 365:  # Filtrar outliers
                            tempos_finalizacao.append(dias)
                            
                            # Por mês
                            mes_ano = data_inicio_dt.strftime('%Y-%m')
                            if mes_ano not in tempos_por_mes:
                                tempos_por_mes[mes_ano] = []
                            tempos_por_mes[mes_ano].append(dias)
                            
                            # Por modalidade
                            modalidade = p.get('modalidade', 'Não Definido')
                            if modalidade not in tempos_por_modalidade:
                                tempos_por_modalidade[modalidade] = []
                            tempos_por_modalidade[modalidade].append(dias)
            except Exception as e:
                continue
    
    tempo_medio_geral = round(sum(tempos_finalizacao) / len(tempos_finalizacao), 1) if tempos_finalizacao else 0
    
    # Tempo médio por mês
    tempo_medio_por_mes = []
    for mes, tempos in sorted(tempos_por_mes.items()):
        tempo_medio_por_mes.append({
            'mes': mes,
            'tempo_medio': round(sum(tempos) / len(tempos), 1) if tempos else 0,
            'quantidade': len(tempos)
        })
    
    # Tempo médio por modalidade
    tempo_medio_por_modalidade = []
    for modalidade, tempos in tempos_por_modalidade.items():
        tempo_medio_por_modalidade.append({
            'modalidade': modalidade,
            'tempo_medio': round(sum(tempos) / len(tempos), 1) if tempos else 0,
            'quantidade': len(tempos)
        })
    tempo_medio_por_modalidade.sort(key=lambda x: x['tempo_medio'], reverse=True)
    
    # 4. Processos por Responsável/Usuário
    processos_por_responsavel = {}
    for p in processos_validos:
        responsavel = p.get('responsavel', 'Não Atribuído') or 'Não Atribuído'
        if responsavel not in processos_por_responsavel:
            processos_por_responsavel[responsavel] = {
                'responsavel': responsavel,
                'quantidade': 0,
                'concluidos': 0,
                'em_andamento': 0
            }
        processos_por_responsavel[responsavel]['quantidade'] += 1
        if p.get('status') == 'Concluído':
            processos_por_responsavel[responsavel]['concluidos'] += 1
        else:
            processos_por_responsavel[responsavel]['em_andamento'] += 1
    
    lista_responsaveis = list(processos_por_responsavel.values())
    lista_responsaveis.sort(key=lambda x: x['quantidade'], reverse=True)
    
    # 5. Processos por Secretaria
    processos_por_secretaria = {}
    for p in processos_validos:
        secretaria = p.get('secretaria', 'Não Informada') or 'Não Informada'
        if secretaria not in processos_por_secretaria:
            processos_por_secretaria[secretaria] = {'secretaria': secretaria, 'quantidade': 0}
        processos_por_secretaria[secretaria]['quantidade'] += 1
    
    lista_secretarias = list(processos_por_secretaria.values())
    lista_secretarias.sort(key=lambda x: x['quantidade'], reverse=True)
    
    # 6. Processos por mês (timeline)
    processos_por_mes = {}
    for p in processos_validos:
        data_inicio_str = p.get('data_inicio')
        if data_inicio_str and data_inicio_str not in ['None', 'null', '']:
            try:
                if 'T' in str(data_inicio_str):
                    dt = datetime.fromisoformat(str(data_inicio_str).replace('Z', '+00:00'))
                else:
                    dt = datetime.strptime(str(data_inicio_str).split(' ')[0], '%Y-%m-%d')
                mes_ano = dt.strftime('%Y-%m')
                if mes_ano not in processos_por_mes:
                    processos_por_mes[mes_ano] = {'mes': mes_ano, 'quantidade': 0, 'concluidos': 0}
                processos_por_mes[mes_ano]['quantidade'] += 1
                if p.get('status') == 'Concluído':
                    processos_por_mes[mes_ano]['concluidos'] += 1
            except:
                pass
    
    timeline = list(processos_por_mes.values())
    timeline.sort(key=lambda x: x['mes'])
    
    # Calcular médias gerais
    total_processos = len(processos_validos)
    total_concluidos = sum(1 for p in processos_validos if p.get('status') == 'Concluído')
    total_em_andamento = total_processos - total_concluidos
    taxa_conclusao = round((total_concluidos / total_processos * 100), 1) if total_processos > 0 else 0
    
    return {
        'total_processos': total_processos,
        'total_concluidos': total_concluidos,
        'total_em_andamento': total_em_andamento,
        'taxa_conclusao': taxa_conclusao,
        'tempo_medio_dias': tempo_medio_geral,
        'stats_by_status': list(stats_by_status.values()),
        'stats_by_modalidade': list(stats_by_modalidade.values()),
        'tempo_medio_por_mes': tempo_medio_por_mes,
        'tempo_medio_por_modalidade': tempo_medio_por_modalidade,
        'processos_por_responsavel': lista_responsaveis,
        'processos_por_secretaria': lista_secretarias,
        'timeline': timeline
    }

@api_router.post("/processos", response_model=Processo)
async def create_processo(processo_data: ProcessoCreate, request: Request):
    """Cria um novo processo - qualquer usuário autenticado"""
    user = await get_current_user(request)
    processo_id = f"proc_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    processo_doc = {
        'processo_id': processo_id,
        'user_id': user.user_id,
        **processo_data.model_dump(),
        'created_at': now,
        'updated_at': now
    }
    
    # Extrair ano do numero_processo se não fornecido (ex: PRC - 0006/2025)
    # Usar o ano do payload se fornecido explicitamente
    ano_fornecido = processo_doc.get('ano')
    if ano_fornecido is None or ano_fornecido == 0:
        numero = processo_doc.get('numero_processo', '')
        match = re.search(r'/(\d{4})$', numero)
        if match:
            processo_doc['ano'] = int(match.group(1))
        else:
            processo_doc['ano'] = now.year
    
    # Converter datas para o formato correto
    if processo_doc.get('data_inicio'):
        processo_doc['data_inicio'] = processo_doc['data_inicio'].isoformat() if isinstance(processo_doc['data_inicio'], datetime) else processo_doc['data_inicio']
    if processo_doc.get('data_autuacao'):
        processo_doc['data_autuacao'] = processo_doc['data_autuacao'].isoformat() if isinstance(processo_doc['data_autuacao'], datetime) else processo_doc['data_autuacao']
    if processo_doc.get('data_contrato'):
        processo_doc['data_contrato'] = processo_doc['data_contrato'].isoformat() if isinstance(processo_doc['data_contrato'], datetime) else processo_doc['data_contrato']
    
    await db.processos.insert_one(processo_doc)
    return Processo(**processo_doc)

@api_router.get("/processos/{processo_id}", response_model=Processo)
async def get_processo(processo_id: str, request: Request):
    """Obtém um processo específico"""
    user = await get_current_user(request)
    processo = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0})
    if not processo:
        raise HTTPException(status_code=404, detail="Processo not found")
    return Processo(**processo)

@api_router.put("/processos/{processo_id}", response_model=Processo)
async def update_processo(processo_id: str, processo_update: ProcessoUpdate, request: Request):
    """Atualiza um processo - qualquer usuário autenticado"""
    user = await get_current_user(request)
    processo = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0})
    if not processo:
        raise HTTPException(status_code=404, detail="Processo not found")
    
    update_data = {k: v for k, v in processo_update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc)
    
    # Converter datas
    for date_field in ['data_inicio', 'data_autuacao', 'data_contrato']:
        if date_field in update_data and update_data[date_field]:
            if isinstance(update_data[date_field], datetime):
                update_data[date_field] = update_data[date_field].isoformat()
    
    await db.processos.update_one({'processo_id': processo_id}, {'$set': update_data})
    updated = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0})
    return Processo(**updated)

@api_router.delete("/processos/{processo_id}")
async def delete_processo(processo_id: str, request: Request):
    """Exclui um processo - APENAS ADMINISTRADORES"""
    user = await get_current_user(request)
    
    # Verificar se é admin
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem excluir processos")
    
    processo = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0})
    if not processo:
        raise HTTPException(status_code=404, detail="Processo not found")
    
    await db.processos.delete_one({'processo_id': processo_id})
    return {'message': 'Processo excluído com sucesso'}

@api_router.post("/processos/migrate-fields")
async def migrate_processos_fields(request: Request):
    """
    Migra campos antigos para nova nomenclatura:
    - modalidade -> modalidade_contratacao
    - situacao -> status (se status vazio)
    ADMIN ONLY
    """
    user = await get_current_user(request)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem executar migrações")
    
    processos = await db.processos.find({}).to_list(None)
    migrated_count = 0
    
    for proc in processos:
        update_data = {}
        
        # Migrar modalidade para modalidade_contratacao
        if proc.get('modalidade') and not proc.get('modalidade_contratacao'):
            update_data['modalidade_contratacao'] = proc['modalidade']
        
        # Se não tem status mas tem situação antiga, usar como status
        # Nota: O campo status já existe mas com valores diferentes
        # Precisamos mapear os valores antigos para os novos
        old_status = proc.get('status', '')
        status_mapping = {
            'Iniciado': 'Em Elaboração',
            'Publicado': 'Em Licitação',
            'Aguardando Jurídico': 'Em Licitação',
            'Homologado': 'Homologado',
            'Concluído': 'Concluído',
            'Cancelado': 'Cancelado'
        }
        
        if old_status in status_mapping:
            update_data['status'] = status_mapping[old_status]
        
        if update_data:
            await db.processos.update_one(
                {'processo_id': proc['processo_id']},
                {'$set': update_data}
            )
            migrated_count += 1
    
    return {
        'message': f'Migração concluída: {migrated_count} processos atualizados',
        'total_processos': len(processos),
        'migrated': migrated_count
    }

@api_router.post("/processos/import")
async def import_processos(file: UploadFile = File(...), request: Request = None):
    """Importa processos de um arquivo Excel/CSV"""
    user = await get_current_user(request)
    
    content = await file.read()
    filename = file.filename.lower()
    
    items_to_import = []
    
    try:
        if filename.endswith('.csv'):
            import csv
            import io
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded), delimiter=';')
            for row in reader:
                items_to_import.append(row)
                
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(content))
            ws = wb.active
            
            headers = [cell.value for cell in ws[1] if cell.value]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    item = dict(zip(headers, row))
                    items_to_import.append(item)
        else:
            raise HTTPException(status_code=400, detail="Formato não suportado. Use CSV ou XLSX")
        
        imported_count = 0
        for item_data in items_to_import:
            # Mapear campos da planilha
            numero = str(item_data.get('Processo', item_data.get('Column1', '')))
            if not numero:
                continue
            
            processo_id = f"proc_{uuid.uuid4().hex[:12]}"
            now = datetime.now(timezone.utc)
            
            # Processar datas
            data_inicio = item_data.get('Data de Início', item_data.get('Column6'))
            data_autuacao = item_data.get('Data de Autuação', item_data.get('Column7'))
            data_contrato = item_data.get('Data do Contrato', item_data.get('Column8'))
            
            processo_doc = {
                'processo_id': processo_id,
                'user_id': user.user_id,
                'numero_processo': numero,
                'status': str(item_data.get('Status', item_data.get('Column2', 'Iniciado'))),
                'modalidade': str(item_data.get('Modalidade', item_data.get('Column3', ''))),
                'objeto': str(item_data.get('Objeto', item_data.get('Column4', '')))[:1000],
                'situacao': str(item_data.get('Situação', item_data.get('Column5', '')))[:50],
                'responsavel': str(item_data.get('Responsável', item_data.get('Column6', '')))[:100] if not isinstance(item_data.get('Responsável', item_data.get('Column6')), datetime) else '',
                'data_inicio': str(data_inicio) if data_inicio else None,
                'data_autuacao': str(data_autuacao) if data_autuacao else None,
                'data_contrato': str(data_contrato) if data_contrato else None,
                'secretaria': str(item_data.get('Secretaria', item_data.get('Column9', '')))[:200],
                'secretario': str(item_data.get('Secretário', item_data.get('Column10', '')))[:100],
                'observacoes': str(item_data.get('Observações', item_data.get('Column11', '')))[:500] if item_data.get('Observações', item_data.get('Column11')) else '',
                'created_at': now,
                'updated_at': now
            }
            
            await db.processos.insert_one(processo_doc)
            imported_count += 1
        
        return {'message': f'{imported_count} processos importados com sucesso', 'total_items': imported_count}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

@api_router.get("/processos/export/pdf")
async def export_processos_pdf(request: Request, orientation: str = "landscape", assinar: bool = True, data: str = None):
    """
    Exporta todos os processos para PDF com margens padronizadas
    Args:
        assinar: Se True, adiciona assinatura digital
        data: Data da assinatura (formato DD/MM/YYYY HH:MM:SS) - retroativa se necessário
    """
    user = await get_current_user(request)
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    
    # Margens padronizadas conforme Lei 14.133/2021
    # 5cm (esquerda/direita), 3cm (superior/inferior)
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='GESTÃO PROCESSUAL - RELATÓRIO DE PROCESSOS',
        subtitulo='Lei Federal nº 14.133/2021',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm, 
        rightMargin=15*mm, 
        topMargin=35*mm,  # Espaço para cabeçalho oficial
        bottomMargin=25*mm  # Espaço para rodapé oficial
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=14,
        textColor=colors.HexColor('#1F4788'), spaceAfter=4,
        alignment=TA_CENTER, fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle', parent=styles['Heading2'], fontSize=11,
        textColor=colors.HexColor('#1F4788'), spaceAfter=6,
        alignment=TA_CENTER, fontName='Helvetica-Bold'
    )
    
    # Tabela de processos - campos COMPLETOS
    if orientation.lower() == 'portrait':
        table_data = [['#', 'Processo', 'Status', 'Modalidade', 'Objeto', 'Secretaria']]
        for idx, p in enumerate(processos, start=1):
            table_data.append([
                str(idx),
                p.get('numero_processo', ''),  # COMPLETO
                p.get('status', ''),  # COMPLETO
                Paragraph(f"<font size=6>{p.get('modalidade', '')}</font>", styles['Normal']),  # COMPLETO
                Paragraph(f"<font size=6>{p.get('objeto', '')}</font>", styles['Normal']),  # COMPLETO
                Paragraph(f"<font size=6>{p.get('secretaria', '')}</font>", styles['Normal'])  # COMPLETO
            ])
        col_widths = [0.5*cm, 2*cm, 1.8*cm, 3*cm, 5.5*cm, 3.5*cm]
    else:
        table_data = [['#', 'Processo', 'Status', 'Modalidade', 'Objeto', 'Responsável', 'Secretaria', 'Observações']]
        for idx, p in enumerate(processos, start=1):
            table_data.append([
                str(idx),
                p.get('numero_processo', ''),  # COMPLETO
                p.get('status', ''),  # COMPLETO
                Paragraph(f"<font size=6>{p.get('modalidade', '')}</font>", styles['Normal']),  # COMPLETO
                Paragraph(f"<font size=6>{p.get('objeto', '')}</font>", styles['Normal']),  # COMPLETO
                p.get('responsavel', ''),  # COMPLETO
                Paragraph(f"<font size=6>{p.get('secretaria', '')}</font>", styles['Normal']),  # COMPLETO
                Paragraph(f"<font size=5>{p.get('observacoes', '')}</font>", styles['Normal'])  # COMPLETO
            ])
        col_widths = [0.5*cm, 2.2*cm, 1.6*cm, 2.5*cm, 6.5*cm, 2*cm, 4*cm, 4*cm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('ALIGN', (0, 1), (2, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 8*mm))
    
    # Rodapé
    footer_text = f'<font size=6><i>Total de {len(processos)} processos</i></font>'
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', alignment=TA_CENTER, textColor=colors.grey)))
    
    # Build com callback oficial para cabeçalho/rodapé em todas as páginas
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    orientation_name = 'Paisagem' if orientation.lower() == 'landscape' else 'Retrato'
    
    # Adicionar assinatura digital ao PDF
    try:
        signed_buffer, validation_code = await add_signature_to_pdf(
            buffer, user, "Relatório de Gestão Processual", f"processos_{datetime.now().strftime('%Y%m%d%H%M%S')}", None, data
        )
        return StreamingResponse(
            signed_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=Gestao_Processual_{orientation_name}.pdf",
                "X-Validation-Code": validation_code
            }
        )
    except HTTPException:
        # Re-raise HTTPException para validação de CPF/Cargo
        raise
    except Exception as e:
        logging.error(f"Erro ao adicionar assinatura ao PDF de processos: {e}")
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Gestao_Processual_{orientation_name}.pdf"}
        )

@api_router.get("/processos/export/xlsx")
async def export_processos_xlsx(request: Request):
    """Exporta todos os processos para Excel"""
    user = await get_current_user(request)
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Gestão Processual"
    
    # Cabeçalho
    headers = ['#', 'Processo', 'Status', 'Modalidade', 'Objeto', 'Situação', 'Responsável', 
               'Data Início', 'Data Autuação', 'Data Contrato', 'Secretaria', 'Secretário', 'Observações']
    
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row, p in enumerate(processos, start=2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=p.get('numero_processo', ''))
        ws.cell(row=row, column=3, value=p.get('status', ''))
        ws.cell(row=row, column=4, value=p.get('modalidade', ''))
        ws.cell(row=row, column=5, value=p.get('objeto', ''))
        ws.cell(row=row, column=6, value=p.get('situacao', ''))
        ws.cell(row=row, column=7, value=p.get('responsavel', ''))
        ws.cell(row=row, column=8, value=p.get('data_inicio', ''))
        ws.cell(row=row, column=9, value=p.get('data_autuacao', ''))
        ws.cell(row=row, column=10, value=p.get('data_contrato', ''))
        ws.cell(row=row, column=11, value=p.get('secretaria', ''))
        ws.cell(row=row, column=12, value=p.get('secretario', ''))
        ws.cell(row=row, column=13, value=p.get('observacoes', ''))
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 30
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['M'].width = 40
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Gestao_Processual.xlsx"}
    )

# ============ BACKUP E RESTAURAÇÃO DO SISTEMA ============

import json

# ============ FUNÇÃO HELPER PARA SERIALIZAÇÃO JSON ============
def serialize_for_json(obj):
    """Converte objetos para tipos serializáveis em JSON"""
    if isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {k: serialize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [serialize_for_json(item) for item in obj]
    elif hasattr(obj, '__dict__'):
        return serialize_for_json(obj.__dict__)
    return obj

def serialize_document(doc):
    """Serializa um documento MongoDB removendo _id e convertendo datetime"""
    if doc is None:
        return None
    serialized = {}
    for key, value in doc.items():
        if key == '_id':
            continue
        serialized[key] = serialize_for_json(value)
    return serialized

@api_router.get("/backup/export")
async def export_backup(request: Request):
    """
    Exporta todos os dados do sistema para um arquivo JSON.
    Permite restaurar os dados após fork/redeploy.
    Apenas administradores podem fazer backup.
    """
    user = await get_current_user(request)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem fazer backup")
    
    try:
        # Coletar todos os dados das coleções
        backup_data = {
            'metadata': {
                'version': '2.1',
                'exported_at': datetime.now(timezone.utc).isoformat(),
                'exported_by': user.email,
                'system': 'Planejamento Acaiaca'
            },
            'users': [],
            'pacs': [],
            'pac_items': [],
            'pacs_geral': [],
            'pac_geral_items': [],
            'pacs_geral_obras': [],
            'pac_geral_obras_items': [],
            'processos': [],
            'mrosc_projetos': [],
            'mrosc_rh': [],
            'mrosc_despesas': [],
            'mrosc_documentos': [],
            'document_signatures': [],
            'user_sessions': []
        }
        
        # Exportar todas as coleções usando a função helper
        collections_map = [
            ('users', 'users'),
            ('pacs', 'pacs'),
            ('pac_items', 'pac_items'),
            ('pacs_geral', 'pacs_geral'),
            ('pac_geral_items', 'pac_geral_items'),
            ('pacs_geral_obras', 'pacs_geral_obras'),
            ('pac_geral_obras_items', 'pac_geral_obras_items'),
            ('processos', 'processos'),
            ('mrosc_projetos', 'mrosc_projetos'),
            ('mrosc_rh', 'mrosc_rh'),
            ('mrosc_despesas', 'mrosc_despesas'),
            ('mrosc_documentos', 'mrosc_documentos'),
            ('document_signatures', 'document_signatures'),
        ]
        
        for collection_name, backup_key in collections_map:
            try:
                collection = db[collection_name]
                docs = await collection.find({}).to_list(100000)
                backup_data[backup_key] = [serialize_document(doc) for doc in docs]
            except Exception as e:
                logging.warning(f"Erro ao exportar {collection_name}: {str(e)}")
                backup_data[backup_key] = []
        
        # Gerar JSON
        json_content = json.dumps(backup_data, ensure_ascii=False, indent=2, default=str)
        
        # Criar resposta como download
        buffer = BytesIO(json_content.encode('utf-8'))
        buffer.seek(0)
        
        filename = f"backup_planejamento_acaiaca_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        return StreamingResponse(
            buffer,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        logging.error(f"Erro ao gerar backup: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar backup: {str(e)}")


@api_router.post("/backup/restore")
async def restore_backup(request: Request, file: UploadFile = File(...)):
    """
    Restaura todos os dados do sistema a partir de um arquivo de backup JSON.
    ATENÇÃO: Esta operação substitui TODOS os dados existentes!
    Apenas administradores podem restaurar backup.
    """
    user = await get_current_user(request)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem restaurar backup")
    
    try:
        # Ler arquivo
        contents = await file.read()
        backup_data = json.loads(contents.decode('utf-8'))
        
        # Validar estrutura do backup
        required_keys = ['metadata', 'users', 'pacs', 'pac_items', 'pacs_geral', 'pac_geral_items', 'processos']
        for key in required_keys:
            if key not in backup_data:
                raise HTTPException(status_code=400, detail=f"Arquivo de backup inválido: falta a chave '{key}'")
        
        # Contadores
        stats = {
            'users': 0,
            'pacs': 0,
            'pac_items': 0,
            'pacs_geral': 0,
            'pac_geral_items': 0,
            'processos': 0
        }
        
        # Restaurar usuários (não sobrescreve o admin atual)
        current_admin_email = user.email
        for u in backup_data['users']:
            # Não sobrescrever o usuário admin atual
            if u.get('email') == current_admin_email:
                continue
            
            # Converter datas
            if 'created_at' in u and isinstance(u['created_at'], str):
                u['created_at'] = datetime.fromisoformat(u['created_at'].replace('Z', '+00:00'))
            
            # Upsert por user_id
            await db.users.update_one(
                {'user_id': u['user_id']},
                {'$set': u},
                upsert=True
            )
            stats['users'] += 1
        
        # Restaurar PACs
        for p in backup_data['pacs']:
            for field in ['created_at', 'updated_at']:
                if field in p and isinstance(p[field], str):
                    p[field] = datetime.fromisoformat(p[field].replace('Z', '+00:00'))
            
            await db.pacs.update_one(
                {'pac_id': p['pac_id']},
                {'$set': p},
                upsert=True
            )
            stats['pacs'] += 1
        
        # Restaurar PAC Items
        for item in backup_data['pac_items']:
            if 'created_at' in item and isinstance(item['created_at'], str):
                item['created_at'] = datetime.fromisoformat(item['created_at'].replace('Z', '+00:00'))
            
            await db.pac_items.update_one(
                {'item_id': item['item_id']},
                {'$set': item},
                upsert=True
            )
            stats['pac_items'] += 1
        
        # Restaurar PACs Geral
        for pg in backup_data['pacs_geral']:
            for field in ['created_at', 'updated_at']:
                if field in pg and isinstance(pg[field], str):
                    pg[field] = datetime.fromisoformat(pg[field].replace('Z', '+00:00'))
            
            await db.pacs_geral.update_one(
                {'pac_geral_id': pg['pac_geral_id']},
                {'$set': pg},
                upsert=True
            )
            stats['pacs_geral'] += 1
        
        # Restaurar PAC Geral Items
        for item in backup_data['pac_geral_items']:
            if 'created_at' in item and isinstance(item['created_at'], str):
                item['created_at'] = datetime.fromisoformat(item['created_at'].replace('Z', '+00:00'))
            
            await db.pac_geral_items.update_one(
                {'item_id': item['item_id']},
                {'$set': item},
                upsert=True
            )
            stats['pac_geral_items'] += 1
        
        # Restaurar PACs Geral Obras
        for pgo in backup_data.get('pacs_geral_obras', []):
            for field in ['created_at', 'updated_at']:
                if field in pgo and isinstance(pgo[field], str):
                    try:
                        pgo[field] = datetime.fromisoformat(pgo[field].replace('Z', '+00:00'))
                    except:
                        pass
            await db.pacs_geral_obras.update_one(
                {'pac_obras_id': pgo['pac_obras_id']},
                {'$set': pgo},
                upsert=True
            )
            stats['pacs_geral_obras'] = stats.get('pacs_geral_obras', 0) + 1
        
        # Restaurar PAC Geral Obras Items
        for item in backup_data.get('pac_geral_obras_items', []):
            if 'created_at' in item and isinstance(item['created_at'], str):
                try:
                    item['created_at'] = datetime.fromisoformat(item['created_at'].replace('Z', '+00:00'))
                except:
                    pass
            await db.pac_geral_obras_items.update_one(
                {'item_id': item['item_id']},
                {'$set': item},
                upsert=True
            )
            stats['pac_geral_obras_items'] = stats.get('pac_geral_obras_items', 0) + 1
        
        # Restaurar Processos
        for proc in backup_data['processos']:
            for field in ['created_at', 'updated_at', 'data_inicio', 'data_autuacao', 'data_contrato']:
                if field in proc and isinstance(proc[field], str):
                    try:
                        proc[field] = datetime.fromisoformat(proc[field].replace('Z', '+00:00'))
                    except:
                        proc[field] = None
            
            await db.processos.update_one(
                {'processo_id': proc['processo_id']},
                {'$set': proc},
                upsert=True
            )
            stats['processos'] += 1
        
        # Restaurar MROSC Projetos
        for proj in backup_data.get('mrosc_projetos', []):
            for field in ['created_at', 'updated_at', 'data_inicio', 'data_fim']:
                if field in proj and isinstance(proj[field], str):
                    try:
                        proj[field] = datetime.fromisoformat(proj[field].replace('Z', '+00:00'))
                    except:
                        pass
            await db.mrosc_projetos.update_one(
                {'projeto_id': proj['projeto_id']},
                {'$set': proj},
                upsert=True
            )
            stats['mrosc_projetos'] = stats.get('mrosc_projetos', 0) + 1
        
        # Restaurar MROSC RH
        for rh in backup_data.get('mrosc_rh', []):
            if 'created_at' in rh and isinstance(rh['created_at'], str):
                try:
                    rh['created_at'] = datetime.fromisoformat(rh['created_at'].replace('Z', '+00:00'))
                except:
                    pass
            await db.mrosc_rh.update_one(
                {'rh_id': rh['rh_id']},
                {'$set': rh},
                upsert=True
            )
            stats['mrosc_rh'] = stats.get('mrosc_rh', 0) + 1
        
        # Restaurar MROSC Despesas
        for desp in backup_data.get('mrosc_despesas', []):
            if 'created_at' in desp and isinstance(desp['created_at'], str):
                try:
                    desp['created_at'] = datetime.fromisoformat(desp['created_at'].replace('Z', '+00:00'))
                except:
                    pass
            await db.mrosc_despesas.update_one(
                {'despesa_id': desp['despesa_id']},
                {'$set': desp},
                upsert=True
            )
            stats['mrosc_despesas'] = stats.get('mrosc_despesas', 0) + 1
        
        # Restaurar MROSC Documentos
        for doc in backup_data.get('mrosc_documentos', []):
            for field in ['created_at', 'data_documento', 'validated_at']:
                if field in doc and isinstance(doc[field], str):
                    try:
                        doc[field] = datetime.fromisoformat(doc[field].replace('Z', '+00:00'))
                    except:
                        pass
            await db.mrosc_documentos.update_one(
                {'documento_id': doc['documento_id']},
                {'$set': doc},
                upsert=True
            )
            stats['mrosc_documentos'] = stats.get('mrosc_documentos', 0) + 1
        
        # Restaurar Assinaturas de Documentos
        for sig in backup_data.get('document_signatures', []):
            if 'created_at' in sig and isinstance(sig['created_at'], str):
                try:
                    sig['created_at'] = datetime.fromisoformat(sig['created_at'].replace('Z', '+00:00'))
                except:
                    pass
            await db.document_signatures.update_one(
                {'validation_code': sig['validation_code']},
                {'$set': sig},
                upsert=True
            )
            stats['document_signatures'] = stats.get('document_signatures', 0) + 1
        
        return {
            'success': True,
            'message': 'Backup restaurado com sucesso!',
            'stats': stats,
            'backup_metadata': backup_data.get('metadata', {})
        }
    
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Arquivo de backup inválido: não é um JSON válido")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao restaurar backup: {str(e)}")


@api_router.get("/backup/info")
async def get_backup_info(request: Request):
    """
    Retorna informações sobre os dados atuais do sistema para backup.
    """
    user = await get_current_user(request)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem ver informações de backup")
    
    # Contar documentos em cada coleção
    users_count = await db.users.count_documents({})
    pacs_count = await db.pacs.count_documents({})
    pac_items_count = await db.pac_items.count_documents({})
    pacs_geral_count = await db.pacs_geral.count_documents({})
    pac_geral_items_count = await db.pac_geral_items.count_documents({})
    pacs_geral_obras_count = await db.pacs_geral_obras.count_documents({})
    pac_geral_obras_items_count = await db.pac_geral_obras_items.count_documents({})
    processos_count = await db.processos.count_documents({})
    mrosc_projetos_count = await db.mrosc_projetos.count_documents({})
    mrosc_rh_count = await db.mrosc_rh.count_documents({})
    mrosc_despesas_count = await db.mrosc_despesas.count_documents({})
    mrosc_documentos_count = await db.mrosc_documentos.count_documents({})
    document_signatures_count = await db.document_signatures.count_documents({})
    
    total = (users_count + pacs_count + pac_items_count + pacs_geral_count + 
             pac_geral_items_count + pacs_geral_obras_count + pac_geral_obras_items_count +
             processos_count + mrosc_projetos_count + mrosc_rh_count + mrosc_despesas_count +
             mrosc_documentos_count + document_signatures_count)
    
    return {
        'system': 'Planejamento Acaiaca',
        'version': '2.0',
        'current_data': {
            'users': users_count,
            'pacs': pacs_count,
            'pac_items': pac_items_count,
            'pacs_geral': pacs_geral_count,
            'pac_geral_items': pac_geral_items_count,
            'pacs_geral_obras': pacs_geral_obras_count,
            'pac_geral_obras_items': pac_geral_obras_items_count,
            'processos': processos_count,
            'mrosc_projetos': mrosc_projetos_count,
            'mrosc_rh': mrosc_rh_count,
            'mrosc_despesas': mrosc_despesas_count,
            'mrosc_documentos': mrosc_documentos_count,
            'document_signatures': document_signatures_count
        },
        'total_records': total,
        'backup_available': True,
        'restore_available': True
    }


# ============ PORTAL PÚBLICO - TRANSPARÊNCIA ============
# Endpoints públicos para acesso sem autenticação
# Conformidade com Lei de Acesso à Informação (Lei 12.527/2011)

public_router = APIRouter(prefix="/api/public")

@public_router.get("/pacs")
async def public_get_pacs(
    secretaria: str = None,
    ano: str = "2026",
    page: int = 1,
    limit: int = 50
):
    """
    Lista pública de PACs Individuais.
    Não requer autenticação - Portal de Transparência.
    """
    query = {}
    if secretaria:
        query['secretaria'] = {'$regex': secretaria, '$options': 'i'}
    if ano:
        query['ano'] = ano
    
    skip = (page - 1) * limit
    total = await db.pacs.count_documents(query)
    
    pacs = await db.pacs.find(query, {
        '_id': 0,
        'user_id': 0  # Ocultar informações internas
    }).skip(skip).limit(limit).to_list(limit)
    
    return {
        'success': True,
        'data': pacs,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': total,
            'pages': (total + limit - 1) // limit
        }
    }

@public_router.get("/pacs/{pac_id}")
async def public_get_pac(pac_id: str):
    """Detalhes públicos de um PAC específico."""
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0, 'user_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC não encontrado")
    return {'success': True, 'data': pac}

@public_router.get("/pacs/{pac_id}/items")
async def public_get_pac_items(pac_id: str):
    """Lista pública de itens de um PAC."""
    pac = await db.pacs.find_one({'pac_id': pac_id})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC não encontrado")
    
    items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(1000)
    return {'success': True, 'data': items, 'total': len(items)}

@public_router.get("/pacs-geral")
async def public_get_pacs_geral(page: int = 1, limit: int = 50):
    """Lista pública de PACs Gerais."""
    skip = (page - 1) * limit
    total = await db.pacs_geral.count_documents({})
    
    pacs = await db.pacs_geral.find({}, {
        '_id': 0,
        'user_id': 0
    }).skip(skip).limit(limit).to_list(limit)
    
    return {
        'success': True,
        'data': pacs,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': total,
            'pages': (total + limit - 1) // limit
        }
    }

@public_router.get("/pacs-geral/{pac_geral_id}")
async def public_get_pac_geral(pac_geral_id: str):
    """Detalhes públicos de um PAC Geral."""
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0, 'user_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
    return {'success': True, 'data': pac}

@public_router.get("/pacs-geral/{pac_geral_id}/items")
async def public_get_pac_geral_items(pac_geral_id: str):
    """Lista pública de itens de um PAC Geral."""
    pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
    
    items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(1000)
    return {'success': True, 'data': items, 'total': len(items)}

@public_router.get("/processos")
async def public_get_processos(
    status: str = None,
    modalidade: str = None,
    secretaria: str = None,
    page: int = 1,
    limit: int = 50
):
    """
    Lista pública de processos licitatórios.
    Não requer autenticação - Portal de Transparência.
    """
    query = {}
    if status:
        query['status'] = {'$regex': status, '$options': 'i'}
    if modalidade:
        query['modalidade'] = {'$regex': modalidade, '$options': 'i'}
    if secretaria:
        query['secretaria'] = {'$regex': secretaria, '$options': 'i'}
    
    skip = (page - 1) * limit
    total = await db.processos.count_documents(query)
    
    processos = await db.processos.find(query, {
        '_id': 0,
        'user_id': 0
    }).skip(skip).limit(limit).to_list(limit)
    
    return {
        'success': True,
        'data': processos,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': total,
            'pages': (total + limit - 1) // limit
        }
    }

@public_router.get("/processos/{processo_id}")
async def public_get_processo(processo_id: str):
    """Detalhes públicos de um processo."""
    processo = await db.processos.find_one({'processo_id': processo_id}, {'_id': 0, 'user_id': 0})
    if not processo:
        raise HTTPException(status_code=404, detail="Processo não encontrado")
    return {'success': True, 'data': processo}

@public_router.get("/dashboard/stats")
async def public_dashboard_stats():
    """
    Estatísticas públicas consolidadas para o Dashboard de Transparência.
    """
    # Total de PACs
    total_pacs = await db.pacs.count_documents({})
    total_pacs_geral = await db.pacs_geral.count_documents({})
    total_processos = await db.processos.count_documents({})
    
    # Processos por status
    processos_por_status = await db.processos.aggregate([
        {'$group': {'_id': '$status', 'quantidade': {'$sum': 1}}}
    ]).to_list(100)
    
    # Total de itens e valores
    pac_items = await db.pac_items.find({}, {'_id': 0, 'valorTotal': 1}).to_list(10000)
    pac_geral_items = await db.pac_geral_items.find({}, {'_id': 0, 'valorTotal': 1}).to_list(10000)
    
    valor_total_pac = sum(item.get('valorTotal', 0) for item in pac_items)
    valor_total_pac_geral = sum(item.get('valorTotal', 0) for item in pac_geral_items)
    
    # Classificações orçamentárias mais utilizadas
    classificacoes = await db.pac_geral_items.aggregate([
        {'$group': {
            '_id': '$subitem_classificacao',
            'quantidade': {'$sum': 1},
            'valor_total': {'$sum': '$valorTotal'}
        }},
        {'$sort': {'valor_total': -1}},
        {'$limit': 10}
    ]).to_list(10)
    
    return {
        'success': True,
        'data': {
            'resumo': {
                'total_pacs_individuais': total_pacs,
                'total_pacs_gerais': total_pacs_geral,
                'total_processos': total_processos,
                'valor_total_pac': valor_total_pac,
                'valor_total_pac_geral': valor_total_pac_geral
            },
            'processos_por_status': [
                {'status': p['_id'] or 'Não informado', 'quantidade': p['quantidade']}
                for p in processos_por_status
            ],
            'classificacoes_principais': [
                {
                    'subitem': c['_id'] or 'Não classificado',
                    'quantidade': c['quantidade'],
                    'valor_total': c['valor_total']
                }
                for c in classificacoes
            ]
        },
        'meta': {
            'generated_at': datetime.now(timezone.utc).isoformat(),
            'system': 'Planejamento Acaiaca - Portal de Transparência'
        }
    }

@public_router.get("/classificacoes")
async def public_get_classificacoes():
    """Retorna os códigos de classificação orçamentária (público)."""
    codigos = {
        "339030": {
            "nome": "Material de Consumo",
            "subitens": [
                "Material de Consumo - Combustíveis e Lubrificantes Automotivos",
                "Gás Engarrafado",
                "Gêneros de Alimentação",
                "Material Farmacológico",
                "Material Odontológico",
                "Material Educativo e Esportivo",
                "Material de Expediente",
                "Material de Processamento de Dados",
                "Material de Limpeza e Higienização",
                "Uniformes, Tecidos e Aviamentos",
                "Material para Manutenção de Bens Imóveis",
                "Material Elétrico e Eletrônico",
                "Material Hospitalar"
            ]
        },
        "339036": {
            "nome": "Outros Serviços de Terceiros (Pessoa Física)",
            "subitens": [
                "Serviços Técnicos Profissionais",
                "Manutenção e Conservação de Equipamentos",
                "Manutenção e Conservação de Veículos",
                "Serviços de Limpeza e Conservação",
                "Serviços Médicos e Odontológicos"
            ]
        },
        "339039": {
            "nome": "Outros Serviços de Terceiros (Pessoa Jurídica)",
            "subitens": [
                "Serviços Técnicos Profissionais",
                "Manutenção de Software",
                "Manutenção e Conservação de Veículos",
                "Serviços de Energia Elétrica",
                "Serviços de Telecomunicações",
                "Vigilância Ostensiva",
                "Limpeza e Conservação"
            ]
        },
        "449052": {
            "nome": "Equipamentos e Material Permanente",
            "subitens": [
                "Aparelhos e Equipamentos de Comunicação",
                "Equipamentos Médico-Hospitalares",
                "Mobiliário em Geral",
                "Equipamentos de Processamento de Dados",
                "Máquinas e Utensílios de Escritório",
                "Veículos de Tração Mecânica"
            ]
        }
    }
    return {'success': True, 'data': codigos}


# ============ ENDPOINTS PÚBLICOS DE EXPORTAÇÃO ============
# Seguem as mesmas regras de geração do backend

@public_router.get("/pacs/{pac_id}/export/pdf")
async def public_export_pac_pdf(pac_id: str, orientation: str = "landscape"):
    """
    Exporta PAC Individual para PDF (público) - Formato DOEM.
    """
    pac = await db.pacs.find_one({'pac_id': pac_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC não encontrado")
    items = await db.pac_items.find({'pac_id': pac_id}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='PAC - PLANO ANUAL DE CONTRATAÇÕES',
        subtitulo=f'Exercício {pac.get("ano", "2026")} - Lei Federal nº 14.133/2021',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=55*mm,  # Espaço para cabeçalho oficial
        bottomMargin=35*mm,  # Espaço para rodapé oficial
        title=f'PAC {pac.get("secretaria", "")} 2026'
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Dados da secretaria
    header_data = [
        [f"<b>Secretaria:</b> {pac.get('secretaria', '')}", f"<b>Secretário(a):</b> {pac.get('secretario', '')}"],
        [f"<b>Fiscal do Contrato:</b> {pac.get('fiscal', '')}", f"<b>Telefone:</b> {pac.get('telefone', '')}"],
        [f"<b>E-mail:</b> {pac.get('email', '')}", f"<b>Endereço:</b> {pac.get('endereco', '')}"]
    ]
    
    for row in header_data:
        elements.append(Paragraph(f"{row[0]} | {row[1]}", ParagraphStyle('Normal', fontSize=8, spaceAfter=1)))
    
    elements.append(Spacer(1, 4*mm))
    
    # Tabela de itens
    elements.append(Paragraph('<b>DETALHAMENTO DOS ITENS</b>', ParagraphStyle('Header', fontSize=10, fontName='Helvetica-Bold', spaceAfter=3)))
    
    table_data = [['#', 'Código\nCATMAT', 'Descrição', 'Und', 'Qtd', 'V.Unit', 'V.Total', 'Prior.', 'Justificativa', 'Classif.']]
    
    for idx, item in enumerate(items, start=1):
        classificacao_text = ''
        if item.get('codigo_classificacao'):
            classificacao_text = f"{item['codigo_classificacao']}"
            if item.get('subitem_classificacao'):
                classificacao_text += f"\n{item['subitem_classificacao']}"
        
        descricao = item.get('descricao', '')[:80] if len(item.get('descricao', '')) > 80 else item.get('descricao', '')
        justificativa = (item.get('justificativa', '') or 'N/I')[:60]
        
        table_data.append([
            str(idx),
            item.get('catmat', ''),
            Paragraph(f"<font size=6>{descricao}</font>", styles['Normal']),
            item.get('unidade', ''),
            str(int(item.get('quantidade', 0))),
            f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {item.get('valorTotal', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            item.get('prioridade', ''),
            Paragraph(f"<font size=5>{justificativa}</font>", styles['Normal']),
            Paragraph(f"<font size=5>{classificacao_text}</font>", styles['Normal'])
        ])
    
    total = sum(item.get('valorTotal', 0) for item in items)
    table_data.append(['', '', Paragraph('<b>TOTAL:</b>', styles['Normal']), '', '', '', f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), '', '', ''])
    
    if orientation.lower() == 'portrait':
        col_widths = [0.5*cm, 1.2*cm, 3*cm, 0.8*cm, 0.7*cm, 1.5*cm, 1.6*cm, 0.8*cm, 2.5*cm, 2.2*cm]
    else:
        col_widths = [0.5*cm, 1.3*cm, 5*cm, 0.9*cm, 0.8*cm, 1.6*cm, 1.8*cm, 0.9*cm, 5*cm, 4.5*cm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -2), 6),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),
        ('ALIGN', (5, 1), (6, -1), 'RIGHT'),
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table)
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    filename = f'PAC_{pac.get("secretaria", "").replace(" ", "_")}_2026.pdf'
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@public_router.get("/pacs-geral/{pac_geral_id}/export/pdf")
async def public_export_pac_geral_pdf(pac_geral_id: str, orientation: str = "landscape"):
    """Exporta PAC Geral para PDF (público) - Formato DOEM."""
    pac_geral = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
    if not pac_geral:
        raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
    items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='PAC GERAL - PLANO ANUAL DE CONTRATAÇÕES COMPARTILHADO',
        subtitulo=f'Exercício {pac_geral.get("ano", "2026")} - Lei Federal nº 14.133/2021',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=55*mm,
        bottomMargin=35*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Dados
    elements.append(Paragraph(f"<b>Secretaria Responsável:</b> {pac_geral.get('nome_secretaria', '')}", ParagraphStyle('Normal', fontSize=9)))
    elements.append(Paragraph(f"<b>Secretário(a):</b> {pac_geral.get('secretario', '')} | <b>Fiscal:</b> {pac_geral.get('fiscal_contrato', '')}", ParagraphStyle('Normal', fontSize=9)))
    
    secretarias = pac_geral.get('secretarias_selecionadas', [])
    if secretarias:
        elements.append(Paragraph(f"<b>Secretarias Participantes:</b> {', '.join(secretarias)}", ParagraphStyle('Normal', fontSize=8)))
    
    elements.append(Spacer(1, 4*mm))
    
    # Tabela de itens
    elements.append(Paragraph('<b>DETALHAMENTO DOS ITENS</b>', ParagraphStyle('Header', fontSize=9, fontName='Helvetica-Bold', spaceAfter=2)))
    
    table_data = [['#', 'CATMAT', 'Descrição', 'Justificativa', 'Und', 'Qtd', 'V.Unit', 'V.Total', 'Prior.', 'Classif.']]
    
    for idx, item in enumerate(items, start=1):
        classificacao_text = ''
        if item.get('codigo_classificacao'):
            classificacao_text = f"{item['codigo_classificacao']}"
            if item.get('subitem_classificacao'):
                classificacao_text += f"\n{item['subitem_classificacao']}"
        
        justificativa = (item.get('justificativa', '') or 'N/I')[:50]
        descricao = item.get('descricao', '')[:70]
        
        table_data.append([
            str(idx),
            item.get('catmat', ''),
            Paragraph(f"<font size=6>{descricao}</font>", styles['Normal']),
            Paragraph(f"<font size=5>{justificativa}</font>", styles['Normal']),
            item.get('unidade', ''),
            str(int(item.get('quantidade_total', 0))),
            f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            f"R$ {item.get('valorTotal', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            item.get('prioridade', ''),
            Paragraph(f"<font size=5>{classificacao_text}</font>", styles['Normal'])
        ])
    
    total = sum(item.get('valorTotal', 0) for item in items)
    total_qtd = sum(item.get('quantidade_total', 0) for item in items)
    table_data.append([
        '', '', Paragraph('<b>TOTAL:</b>', styles['Normal']), '', '',
        str(int(total_qtd)), '', f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), '', ''
    ])
    
    if orientation.lower() == 'portrait':
        col_widths = [0.5*cm, 1*cm, 3*cm, 2.5*cm, 0.7*cm, 0.8*cm, 1.4*cm, 1.5*cm, 0.7*cm, 2.2*cm]
    else:
        col_widths = [0.5*cm, 1.3*cm, 5*cm, 4.5*cm, 0.9*cm, 1*cm, 1.8*cm, 2*cm, 0.9*cm, 4*cm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -2), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    elements.append(table)
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    filename = f'PAC_Geral_{pac_geral.get("nome_secretaria", "").replace(" ", "_")}.pdf'
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ===== ROTAS PÚBLICAS PAC GERAL OBRAS =====
@public_router.get("/pacs-geral-obras")
async def public_get_pacs_geral_obras(ano: str = None, page: int = 1, limit: int = 50):
    """Lista todos os PAC Geral Obras (público - transparência)"""
    query = {}
    if ano:
        query['ano'] = ano
    pacs = await db.pacs_geral_obras.find(query, {'_id': 0}).to_list(limit)
    return {'data': pacs, 'total': len(pacs)}

@public_router.get("/pacs-geral-obras/{pac_obras_id}")
async def public_get_pac_geral_obras(pac_obras_id: str):
    """Obtém detalhes de um PAC Geral Obras (público)"""
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras not found")
    return {'data': pac}

@public_router.get("/pacs-geral-obras/{pac_obras_id}/items")
async def public_get_pac_geral_obras_items(pac_obras_id: str):
    """Lista itens de um PAC Geral Obras (público)"""
    items = await db.pac_geral_obras_items.find({'pac_obras_id': pac_obras_id}, {'_id': 0}).to_list(1000)
    return {'data': items}

@public_router.get("/pacs-geral-obras/{pac_obras_id}/export/pdf")
async def public_export_pac_geral_obras_pdf(pac_obras_id: str, orientation: str = "landscape"):
    """Exporta PAC Geral Obras para PDF (público) - Formato DOEM"""
    pac = await db.pacs_geral_obras.find_one({'pac_obras_id': pac_obras_id}, {'_id': 0})
    if not pac:
        raise HTTPException(status_code=404, detail="PAC Geral Obras not found")
    
    items = await db.pac_geral_obras_items.find({'pac_obras_id': pac_obras_id}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='PAC GERAL - OBRAS E SERVIÇOS DE ENGENHARIA',
        subtitulo=f'Lei Federal nº 14.133/2021 | {pac.get("nome_secretaria", "")}',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=55*mm,
        bottomMargin=35*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Dados do PAC
    info_style = ParagraphStyle('Info', fontSize=8)
    elements.append(Paragraph(f'<b>Responsável:</b> {pac.get("secretario", "")} | <b>Fiscal:</b> {pac.get("fiscal_contrato", "-")}', info_style))
    elements.append(Paragraph(f'<b>Email:</b> {pac.get("email", "")} | <b>Ano:</b> {pac.get("ano", "2026")}', info_style))
    elements.append(Spacer(1, 4*mm))
    
    # Tabela de itens
    if items:
        valor_total = sum(item.get('valorTotal', 0) for item in items)
        elements.append(Paragraph(f'<b>Total de Itens: {len(items)} | Valor Total: R$ {valor_total:,.2f}</b>'.replace(',', 'X').replace('.', ',').replace('X', '.'), ParagraphStyle('Info', fontSize=9, fontName='Helvetica-Bold')))
        elements.append(Spacer(1, 2*mm))
        
        table_data = [['#', 'Descrição', 'Classificação', 'Und', 'Qtd', 'V.Unit', 'V.Total']]
        for idx, item in enumerate(items, start=1):
            table_data.append([
                str(idx),
                Paragraph(f"<font size=5>{item.get('descricao', '')[:60]}</font>", styles['Normal']),
                Paragraph(f"<font size=5>{item.get('codigo_classificacao', '')} - {item.get('subitem_classificacao', '')}</font>", styles['Normal']),
                item.get('unidade', ''),
                str(item.get('quantidade_total', 0)),
                f"R$ {item.get('valorUnitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {item.get('valorTotal', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        col_widths = [0.5*cm, 6*cm, 5*cm, 1.2*cm, 1*cm, 2*cm, 2.5*cm] if orientation.lower() == 'landscape' else [0.5*cm, 3.5*cm, 3*cm, 0.8*cm, 0.7*cm, 1.5*cm, 2*cm]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 6),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (-2, 1), (-1, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph('Nenhum item cadastrado.', ParagraphStyle('Center', fontSize=10, alignment=1)))
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    filename = f'PAC_Obras_{pac.get("nome_secretaria", "").replace(" ", "_")}.pdf'
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

@public_router.get("/processos/export/pdf")
async def public_export_processos_pdf(orientation: str = "landscape"):
    """Exporta todos os processos para PDF (público) - Formato DOEM."""
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    buffer = BytesIO()
    page_size = A4 if orientation.lower() == 'portrait' else landscape(A4)
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='GESTÃO PROCESSUAL - RELATÓRIO DE PROCESSOS',
        subtitulo=f'Lei Federal nº 14.133/2021 | Total: {len(processos)} processos',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=55*mm,
        bottomMargin=35*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Tabela de processos
    if orientation.lower() == 'portrait':
        table_data = [['#', 'Processo', 'Status', 'Modalidade', 'Objeto', 'Secretaria']]
        for idx, p in enumerate(processos, start=1):
            table_data.append([
                str(idx),
                p.get('numero_processo', ''),
                p.get('status', ''),
                Paragraph(f"<font size=5>{p.get('modalidade', '')}</font>", styles['Normal']),
                Paragraph(f"<font size=5>{p.get('objeto', '')[:50]}</font>", styles['Normal']),
                Paragraph(f"<font size=5>{p.get('secretaria', '')}</font>", styles['Normal'])
            ])
        col_widths = [0.5*cm, 1.8*cm, 1.5*cm, 2.5*cm, 5*cm, 3*cm]
    else:
        table_data = [['#', 'Processo', 'Status', 'Modalidade', 'Objeto', 'Responsável', 'Secretaria', 'Observações']]
        for idx, p in enumerate(processos, start=1):
            table_data.append([
                str(idx),
                p.get('numero_processo', ''),
                p.get('status', ''),
                Paragraph(f"<font size=6>{p.get('modalidade', '')}</font>", styles['Normal']),
                Paragraph(f"<font size=6>{p.get('objeto', '')}</font>", styles['Normal']),
                p.get('responsavel', ''),
                Paragraph(f"<font size=6>{p.get('secretaria', '')}</font>", styles['Normal']),
                Paragraph(f"<font size=5>{p.get('observacoes', '')}</font>", styles['Normal'])
            ])
        col_widths = [0.5*cm, 2.2*cm, 1.6*cm, 2.5*cm, 6.5*cm, 2*cm, 4*cm, 4*cm]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    elements.append(table)
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    filename = f'Processos_PAC_Acaiaca_{datetime.now().strftime("%Y%m%d")}.pdf'
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ===== ROTAS PÚBLICAS MROSC =====
@public_router.get("/mrosc/projetos")
async def public_get_projetos_mrosc():
    """Lista projetos MROSC (público - transparência)"""
    # Mostrar todos os projetos para transparência
    projetos = await db.mrosc_projetos.find({}, {'_id': 0}).to_list(100)
    return {'data': projetos, 'total': len(projetos)}

@public_router.get("/mrosc/projetos/{projeto_id}")
async def public_get_projeto_mrosc(projeto_id: str):
    """Obtém detalhes de um projeto MROSC (público)"""
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    return {'data': projeto}

@public_router.get("/mrosc/projetos/{projeto_id}/resumo")
async def public_get_resumo_mrosc(projeto_id: str):
    """Resumo financeiro de um projeto MROSC (público)"""
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Buscar RH e despesas
    rhs = await db.mrosc_rh.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    despesas = await db.mrosc_despesas.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(500)
    
    # Calcular totais
    total_rh = sum(rh.get('custo_total_projeto', 0) for rh in rhs)
    total_despesas = sum(d.get('valor_total', 0) for d in despesas)
    valor_total_projeto = projeto.get('valor_total', 0)
    repasse_publico = projeto.get('valor_repasse_publico', 0)
    contrapartida = projeto.get('valor_contrapartida', 0)
    
    return {
        'data': {
            'projeto': projeto,
            'total_rh': total_rh,
            'total_despesas': total_despesas,
            'total_executado': total_rh + total_despesas,
            'valor_total_projeto': valor_total_projeto,
            'repasse_publico': repasse_publico,
            'contrapartida': contrapartida,
            'saldo': valor_total_projeto - (total_rh + total_despesas),
            'quantidade_rh': len(rhs),
            'quantidade_despesas': len(despesas)
        }
    }

@public_router.get("/mrosc/estatisticas")
async def public_get_estatisticas_mrosc():
    """Estatísticas gerais do MROSC (público)"""
    projetos = await db.mrosc_projetos.find({}, {'_id': 0}).to_list(100)
    
    total_projetos = len(projetos)
    total_aprovados = len([p for p in projetos if p.get('status') == 'APROVADO'])
    total_em_analise = len([p for p in projetos if p.get('status') == 'ANALISE'])
    valor_total = sum(p.get('valor_total', 0) for p in projetos)
    valor_repasse = sum(p.get('valor_repasse_publico', 0) for p in projetos)
    
    return {
        'data': {
            'total_projetos': total_projetos,
            'total_aprovados': total_aprovados,
            'total_em_analise': total_em_analise,
            'valor_total': valor_total,
            'valor_repasse_publico': valor_repasse
        }
    }

# ============ MÓDULO MROSC - PRESTAÇÃO DE CONTAS (Lei 13.019/2014) ============
mrosc_router = APIRouter(prefix="/api/mrosc", tags=["Prestação de Contas MROSC"])

def calcular_encargos_clt(salario_bruto: float, numero_meses: int) -> dict:
    """Calcula encargos trabalhistas CLT conforme MROSC"""
    provisao_ferias = salario_bruto / 12
    provisao_13_salario = salario_bruto / 12
    base_fgts = salario_bruto + provisao_ferias + provisao_13_salario
    fgts = base_fgts * 0.08
    inss_patronal = salario_bruto * 0.20
    custo_mensal = salario_bruto + provisao_ferias + provisao_13_salario + fgts + inss_patronal
    return {
        'provisao_ferias': round(provisao_ferias, 2),
        'provisao_13_salario': round(provisao_13_salario, 2),
        'fgts': round(fgts, 2),
        'inss_patronal': round(inss_patronal, 2),
        'custo_mensal_total': round(custo_mensal, 2),
        'custo_total_projeto': round(custo_mensal * numero_meses, 2)
    }

def calcular_media_orcamentos(orc1: float, orc2: float, orc3: float) -> float:
    """Calcula média dos três orçamentos"""
    valores = [v for v in [orc1, orc2, orc3] if v and v > 0]
    return round(sum(valores) / len(valores), 2) if valores else 0.0

# Endpoint para naturezas de despesa
@mrosc_router.get("/naturezas-despesa")
async def get_naturezas_despesa():
    """Retorna naturezas de despesa conforme MROSC"""
    return NATUREZAS_DESPESA_MROSC

# ===== PROJETOS MROSC =====
@mrosc_router.get("/projetos", response_model=List[ProjetoMROSC])
async def get_projetos_mrosc(request: Request):
    """Lista todos os projetos MROSC"""
    user = await get_current_user(request)
    query = {} if user.is_admin else {'user_id': user.user_id}
    projetos = await db.mrosc_projetos.find(query, {'_id': 0}).to_list(100)
    return [ProjetoMROSC(**p) for p in projetos]

@mrosc_router.post("/projetos", response_model=ProjetoMROSC)
async def create_projeto_mrosc(projeto_data: ProjetoMROSCCreate, request: Request):
    """Cria um novo projeto MROSC"""
    user = await get_current_user(request)
    projeto_id = f"mrosc_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    projeto_doc = {
        'projeto_id': projeto_id,
        'user_id': user.user_id,
        **projeto_data.model_dump(),
        'created_at': now,
        'updated_at': now
    }
    await db.mrosc_projetos.insert_one(projeto_doc)
    return ProjetoMROSC(**projeto_doc)

@mrosc_router.get("/projetos/{projeto_id}", response_model=ProjetoMROSC)
async def get_projeto_mrosc(projeto_id: str, request: Request):
    """Obtém um projeto MROSC específico"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    return ProjetoMROSC(**projeto)

@mrosc_router.put("/projetos/{projeto_id}", response_model=ProjetoMROSC)
async def update_projeto_mrosc(projeto_id: str, projeto_data: ProjetoMROSCUpdate, request: Request):
    """Atualiza um projeto MROSC"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    update_data = {k: v for k, v in projeto_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc)
    await db.mrosc_projetos.update_one({'projeto_id': projeto_id}, {'$set': update_data})
    updated = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    return ProjetoMROSC(**updated)

@mrosc_router.delete("/projetos/{projeto_id}")
async def delete_projeto_mrosc(projeto_id: str, request: Request):
    """Exclui um projeto MROSC e todos os dados relacionados"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    # Excluir dados relacionados
    await db.mrosc_rh.delete_many({'projeto_id': projeto_id})
    await db.mrosc_despesas.delete_many({'projeto_id': projeto_id})
    await db.mrosc_documentos.delete_many({'projeto_id': projeto_id})
    await db.mrosc_projetos.delete_one({'projeto_id': projeto_id})
    return {'message': 'Projeto excluído com sucesso'}

# ===== RECURSOS HUMANOS MROSC =====
@mrosc_router.get("/projetos/{projeto_id}/rh", response_model=List[RecursoHumanoMROSC])
async def get_rh_mrosc(projeto_id: str, request: Request):
    """Lista recursos humanos de um projeto"""
    user = await get_current_user(request)
    rhs = await db.mrosc_rh.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    return [RecursoHumanoMROSC(**rh) for rh in rhs]

@mrosc_router.post("/projetos/{projeto_id}/rh", response_model=RecursoHumanoMROSC)
async def create_rh_mrosc(projeto_id: str, rh_data: RecursoHumanoMROSCCreate, request: Request):
    """Adiciona um recurso humano ao projeto"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    rh_id = f"rh_{uuid.uuid4().hex[:12]}"
    rh_dict = rh_data.model_dump()
    
    # Calcular encargos CLT automaticamente
    encargos = calcular_encargos_clt(rh_dict['salario_bruto'], rh_dict['numero_meses'])
    
    # Calcular média de orçamentos se fornecidos
    media = calcular_media_orcamentos(
        rh_dict.get('orcamento_1', 0),
        rh_dict.get('orcamento_2', 0),
        rh_dict.get('orcamento_3', 0)
    )
    
    # Incluir benefícios no custo mensal total
    custo_mensal = encargos['custo_mensal_total'] + rh_dict.get('vale_transporte', 0) + rh_dict.get('vale_alimentacao', 0)
    custo_total = custo_mensal * rh_dict['numero_meses']
    
    rh_doc = {
        'rh_id': rh_id,
        'projeto_id': projeto_id,
        **rh_dict,
        'provisao_ferias': encargos['provisao_ferias'],
        'provisao_13_salario': encargos['provisao_13_salario'],
        'fgts': encargos['fgts'],
        'inss_patronal': encargos['inss_patronal'],
        'custo_mensal_total': round(custo_mensal, 2),
        'custo_total_projeto': round(custo_total, 2),
        'media_orcamentos': media,
        'created_at': datetime.now(timezone.utc)
    }
    await db.mrosc_rh.insert_one(rh_doc)
    return RecursoHumanoMROSC(**rh_doc)

@mrosc_router.delete("/projetos/{projeto_id}/rh/{rh_id}")
async def delete_rh_mrosc(projeto_id: str, rh_id: str, request: Request):
    """Remove um recurso humano do projeto"""
    user = await get_current_user(request)
    result = await db.mrosc_rh.delete_one({'rh_id': rh_id, 'projeto_id': projeto_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Recurso humano não encontrado")
    return {'message': 'Recurso humano excluído com sucesso'}

# ===== DESPESAS MROSC =====
@mrosc_router.get("/projetos/{projeto_id}/despesas", response_model=List[DespesaMROSC])
async def get_despesas_mrosc(projeto_id: str, request: Request):
    """Lista despesas de um projeto"""
    user = await get_current_user(request)
    despesas = await db.mrosc_despesas.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(1000)
    return [DespesaMROSC(**d) for d in despesas]

@mrosc_router.post("/projetos/{projeto_id}/despesas", response_model=DespesaMROSC)
async def create_despesa_mrosc(projeto_id: str, despesa_data: DespesaMROSCCreate, request: Request):
    """Adiciona uma despesa ao projeto"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    despesa_id = f"desp_{uuid.uuid4().hex[:12]}"
    despesa_dict = despesa_data.model_dump()
    
    # Calcular média de orçamentos
    media = calcular_media_orcamentos(
        despesa_dict['orcamento_1'],
        despesa_dict['orcamento_2'],
        despesa_dict['orcamento_3']
    )
    
    valor_total = despesa_dict['quantidade'] * despesa_dict['valor_unitario']
    
    despesa_doc = {
        'despesa_id': despesa_id,
        'projeto_id': projeto_id,
        **despesa_dict,
        'media_orcamentos': media,
        'valor_total': round(valor_total, 2),
        'created_at': datetime.now(timezone.utc)
    }
    await db.mrosc_despesas.insert_one(despesa_doc)
    return DespesaMROSC(**despesa_doc)

@mrosc_router.delete("/projetos/{projeto_id}/despesas/{despesa_id}")
async def delete_despesa_mrosc(projeto_id: str, despesa_id: str, request: Request):
    """Remove uma despesa do projeto"""
    user = await get_current_user(request)
    result = await db.mrosc_despesas.delete_one({'despesa_id': despesa_id, 'projeto_id': projeto_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")
    return {'message': 'Despesa excluída com sucesso'}

# ===== IMPORTAÇÃO DE PLANILHA EXCEL =====
from utils.mrosc_importer import MROSCExcelImporter

@mrosc_router.get("/importar/template")
async def download_template_mrosc(request: Request):
    """
    Baixa o template de planilha Excel para importação MROSC.
    Retorna arquivo .xlsx com as colunas esperadas e exemplos.
    """
    user = await get_current_user(request)
    importer = MROSCExcelImporter()
    template_bytes = importer.generate_template()
    
    return Response(
        content=template_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=template_mrosc_importacao.xlsx"
        }
    )

@mrosc_router.post("/projetos/{projeto_id}/importar/preview")
async def preview_importacao_mrosc(
    projeto_id: str,
    request: Request,
    file: UploadFile = File(...)
):
    """
    Faz preview da importação sem salvar os dados.
    Retorna os dados que seriam importados e eventuais erros.
    """
    user = await get_current_user(request)
    
    # Verificar se projeto existe
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Verificar tipo de arquivo
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser uma planilha Excel (.xlsx ou .xls)")
    
    # Ler arquivo
    file_bytes = await file.read()
    
    # Importar
    importer = MROSCExcelImporter()
    result = importer.import_from_bytes(file_bytes)
    
    return {
        'preview': True,
        'projeto_id': projeto_id,
        'projeto_nome': projeto.get('nome_projeto'),
        **result
    }

@mrosc_router.post("/projetos/{projeto_id}/importar/confirmar")
async def confirmar_importacao_mrosc(
    projeto_id: str,
    request: Request,
    file: UploadFile = File(...),
    substituir_existentes: bool = False
):
    """
    Confirma a importação e salva os dados no banco.
    
    Args:
        projeto_id: ID do projeto
        file: Arquivo Excel
        substituir_existentes: Se True, remove dados existentes antes de importar
    """
    user = await get_current_user(request)
    
    # Verificar se projeto existe
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Verificar tipo de arquivo
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser uma planilha Excel (.xlsx ou .xls)")
    
    # Ler arquivo
    file_bytes = await file.read()
    
    # Importar
    importer = MROSCExcelImporter()
    result = importer.import_from_bytes(file_bytes)
    
    if not result['success']:
        raise HTTPException(status_code=400, detail=f"Erros na importação: {result['errors']}")
    
    # Se substituir, remover dados existentes
    if substituir_existentes:
        await db.mrosc_rh.delete_many({'projeto_id': projeto_id})
        await db.mrosc_despesas.delete_many({'projeto_id': projeto_id})
    
    # Salvar RH
    rh_inserted = 0
    for rh_item in result['rh']['items']:
        rh_item['projeto_id'] = projeto_id
        rh_item['created_at'] = datetime.now(timezone.utc)
        # Remover campo auxiliar
        rh_item.pop('row_number', None)
        await db.mrosc_rh.insert_one(rh_item)
        rh_inserted += 1
    
    # Salvar Despesas
    despesas_inserted = 0
    for despesa_item in result['despesas']['items']:
        despesa_item['projeto_id'] = projeto_id
        despesa_item['created_at'] = datetime.now(timezone.utc)
        # Remover campo auxiliar
        despesa_item.pop('row_number', None)
        await db.mrosc_despesas.insert_one(despesa_item)
        despesas_inserted += 1
    
    # Registrar no histórico
    historico = {
        'historico_id': f"hist_{uuid.uuid4().hex[:12]}",
        'projeto_id': projeto_id,
        'acao': 'IMPORTACAO_EXCEL',
        'usuario_id': user.user_id,
        'usuario_nome': user.name,
        'data': datetime.now(timezone.utc),
        'observacao': f"Importados {rh_inserted} itens de RH e {despesas_inserted} despesas via planilha Excel"
    }
    await db.mrosc_historico.insert_one(historico)
    
    return {
        'success': True,
        'projeto_id': projeto_id,
        'rh_inserted': rh_inserted,
        'despesas_inserted': despesas_inserted,
        'totals': result['totals'],
        'message': f'Importação concluída: {rh_inserted} itens de RH e {despesas_inserted} despesas'
    }

# ===== RESUMO ORÇAMENTÁRIO =====
@mrosc_router.get("/projetos/{projeto_id}/resumo")
async def get_resumo_orcamentario_mrosc(projeto_id: str, request: Request):
    """Retorna resumo orçamentário do projeto"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Somar recursos humanos
    rhs = await db.mrosc_rh.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    total_rh = sum(rh.get('custo_total_projeto', 0) for rh in rhs)
    
    # Somar despesas
    despesas = await db.mrosc_despesas.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(1000)
    total_despesas = sum(d.get('valor_total', 0) for d in despesas)
    
    # Agrupar despesas por natureza
    despesas_por_natureza = {}
    for d in despesas:
        nat = d.get('natureza_despesa', 'OUTROS')
        if nat not in despesas_por_natureza:
            despesas_por_natureza[nat] = 0
        despesas_por_natureza[nat] += d.get('valor_total', 0)
    
    total_geral = total_rh + total_despesas
    
    return {
        'projeto': projeto,
        'resumo': {
            'total_recursos_humanos': round(total_rh, 2),
            'total_despesas': round(total_despesas, 2),
            'total_geral': round(total_geral, 2),
            'quantidade_rh': len(rhs),
            'quantidade_despesas': len(despesas)
        },
        'despesas_por_natureza': despesas_por_natureza,
        'diferenca_orcamento': round(projeto.get('valor_total', 0) - total_geral, 2)
    }

# ===== DOCUMENTOS/COMPROVANTES MROSC =====
UPLOAD_DIR = ROOT_DIR / "uploads" / "mrosc"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

class DocumentoMROSCCreate(BaseModel):
    tipo_documento: str  # NOTA_FISCAL, RECIBO, CONTRATO, COMPROVANTE, OUTRO
    numero_documento: str
    data_documento: datetime
    valor: float
    despesa_id: Optional[str] = None
    observacoes: Optional[str] = None

@mrosc_router.get("/projetos/{projeto_id}/documentos")
async def get_documentos_mrosc(projeto_id: str, request: Request):
    """Lista todos os documentos de um projeto"""
    user = await get_current_user(request)
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    documentos = await db.mrosc_documentos.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(500)
    return documentos

@mrosc_router.post("/projetos/{projeto_id}/documentos/upload")
async def upload_documento_mrosc(
    projeto_id: str,
    request: Request,
    file: UploadFile = File(...),
    tipo_documento: str = Form("COMPROVANTE"),
    numero_documento: str = Form(""),
    data_documento: str = Form(""),
    valor: float = Form(0.0),
    despesa_id: str = Form(None),
    rh_id: str = Form(None),
    descricao: str = Form(""),
    observacoes: str = Form("")
):
    """Faz upload de um documento/comprovante (PDF, JPG, PNG) para o projeto MROSC"""
    user = await get_current_user(request)
    
    # Verifica se o projeto existe
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Tipos de arquivo permitidos
    extensoes_permitidas = {'.pdf', '.jpg', '.jpeg', '.png'}
    extensao = Path(file.filename).suffix.lower()
    
    if extensao not in extensoes_permitidas:
        raise HTTPException(
            status_code=400, 
            detail=f"Tipo de arquivo não permitido. Permitidos: PDF, JPG, JPEG, PNG"
        )
    
    # Define o content-type baseado na extensão
    content_types = {
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png'
    }
    content_type = content_types.get(extensao, 'application/octet-stream')
    
    # Limite de 10MB
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Limite: 10MB")
    
    # Gera nome único para o arquivo
    documento_id = f"doc_{uuid.uuid4().hex[:12]}"
    safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '_', file.filename)
    arquivo_nome = f"{documento_id}_{safe_filename}"
    arquivo_path = UPLOAD_DIR / arquivo_nome
    
    # Salva o arquivo
    with open(arquivo_path, "wb") as f:
        f.write(contents)
    
    # Parseia a data do documento
    try:
        if data_documento:
            dt_documento = datetime.fromisoformat(data_documento.replace('Z', '+00:00'))
        else:
            dt_documento = datetime.now(timezone.utc)
    except:
        dt_documento = datetime.now(timezone.utc)
    
    # Cria o registro no banco
    documento_doc = {
        'documento_id': documento_id,
        'projeto_id': projeto_id,
        'despesa_id': despesa_id if despesa_id and despesa_id != "null" and despesa_id != "" else None,
        'rh_id': rh_id if rh_id and rh_id != "null" and rh_id != "" else None,
        'tipo_documento': tipo_documento,
        'numero_documento': numero_documento,
        'data_documento': dt_documento,
        'valor': valor if valor else 0.0,
        'arquivo_url': f"/api/mrosc/documentos/{documento_id}/download",
        'arquivo_nome': file.filename,
        'arquivo_tipo': content_type,
        'arquivo_tamanho': len(contents),
        'descricao': descricao,
        'validado': False,
        'validado_por': None,
        'data_validacao': None,
        'observacoes': observacoes,
        'created_at': datetime.now(timezone.utc)
    }
    
    await db.mrosc_documentos.insert_one(documento_doc)
    
    # Remove _id para retorno
    documento_doc.pop('_id', None)
    
    return documento_doc

@mrosc_router.get("/documentos/{documento_id}/download")
async def download_documento_mrosc(documento_id: str):
    """Faz download de um documento (PDF ou imagem)"""
    documento = await db.mrosc_documentos.find_one({'documento_id': documento_id}, {'_id': 0})
    if not documento:
        raise HTTPException(status_code=404, detail="Documento não encontrado")
    
    # Define o content-type
    content_type = documento.get('arquivo_tipo', 'application/pdf')
    
    # Procura o arquivo
    for file in UPLOAD_DIR.iterdir():
        if file.name.startswith(documento_id):
            return StreamingResponse(
                open(file, "rb"),
                media_type=content_type,
                headers={
                    "Content-Disposition": f"inline; filename=\"{documento['arquivo_nome']}\""
                }
            )
    
    raise HTTPException(status_code=404, detail="Arquivo não encontrado")

@mrosc_router.delete("/projetos/{projeto_id}/documentos/{documento_id}")
async def delete_documento_mrosc(projeto_id: str, documento_id: str, request: Request):
    """Remove um documento do projeto"""
    user = await get_current_user(request)
    
    documento = await db.mrosc_documentos.find_one({'documento_id': documento_id, 'projeto_id': projeto_id}, {'_id': 0})
    if not documento:
        raise HTTPException(status_code=404, detail="Documento não encontrado")
    
    # Remove o arquivo físico
    for file in UPLOAD_DIR.iterdir():
        if file.name.startswith(documento_id):
            file.unlink()
            break
    
    # Remove do banco
    await db.mrosc_documentos.delete_one({'documento_id': documento_id})
    
    return {'message': 'Documento excluído com sucesso'}

@mrosc_router.put("/projetos/{projeto_id}/documentos/{documento_id}/validar")
async def validar_documento_mrosc(projeto_id: str, documento_id: str, request: Request):
    """Marca um documento como validado"""
    user = await get_current_user(request)
    
    documento = await db.mrosc_documentos.find_one({'documento_id': documento_id, 'projeto_id': projeto_id}, {'_id': 0})
    if not documento:
        raise HTTPException(status_code=404, detail="Documento não encontrado")
    
    await db.mrosc_documentos.update_one(
        {'documento_id': documento_id},
        {'$set': {
            'validado': True,
            'validado_por': user.name if hasattr(user, 'name') else user.get('name', 'Admin'),
            'data_validacao': datetime.now(timezone.utc)
        }}
    )
    
    return {'message': 'Documento validado com sucesso'}

# ===== WORKFLOW DE PRESTAÇÃO DE CONTAS MROSC =====

class SubmeterPrestacaoRequest(BaseModel):
    observacoes: Optional[str] = None

class PedidoCorrecaoRequest(BaseModel):
    motivo: str

class AprovarPrestacaoRequest(BaseModel):
    observacoes: Optional[str] = None

# ===== FUNÇÃO DE NOTIFICAÇÃO MROSC =====
def enviar_notificacao_mrosc(destinatario: str, assunto: str, projeto_nome: str, mensagem: str, acao: str):
    """Envia notificação por email sobre mudança de status no MROSC"""
    try:
        corpo_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: #2E7D32; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0;">Prestação de Contas - MROSC</h1>
                <p style="margin: 5px 0 0 0; font-size: 14px;">Prefeitura Municipal de Acaiaca</p>
            </div>
            <div style="padding: 20px; background: #f5f5f5;">
                <h2 style="color: #1F4E78;">{assunto}</h2>
                <p><strong>Projeto:</strong> {projeto_nome}</p>
                <p>{mensagem}</p>
                <div style="background: white; padding: 15px; border-radius: 8px; margin: 15px 0;">
                    <p style="margin: 0; color: #666;"><strong>Ação:</strong> {acao}</p>
                </div>
                <p style="font-size: 12px; color: #888;">
                    Acesse o sistema para mais detalhes: <a href="https://structurefix-1.preview.emergentagent.com/prestacao-contas">Planejamento Acaiaca</a>
                </p>
            </div>
            <div style="background: #1F4E78; color: white; padding: 10px; text-align: center; font-size: 12px;">
                <p style="margin: 0;">CNPJ: 18.295.287/0001-90 | Lei 13.019/2014 - MROSC</p>
            </div>
        </body>
        </html>
        """
        enviar_email_smtp(destinatario, f"[MROSC] {assunto}", corpo_html)
        return True
    except Exception as e:
        logging.error(f"Erro ao enviar notificação MROSC: {e}")
        return False

@mrosc_router.post("/projetos/{projeto_id}/submeter")
async def submeter_prestacao_contas(projeto_id: str, request: Request, data: SubmeterPrestacaoRequest = None):
    """
    Usuário externo submete a prestação de contas para análise.
    Após submissão, não pode mais editar até que admin solicite correção.
    """
    user = await get_current_user(request)
    
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    if projeto.get('submetido') and not projeto.get('correcao_solicitada'):
        raise HTTPException(status_code=400, detail="Projeto já foi submetido e aguarda análise")
    
    user_name = user.name if hasattr(user, 'name') else user.get('name', 'Usuário')
    
    await db.mrosc_projetos.update_one(
        {'projeto_id': projeto_id},
        {'$set': {
            'submetido': True,
            'data_submissao': datetime.now(timezone.utc),
            'submetido_por': user_name,
            'pode_editar': False,  # Bloqueia edição
            'correcao_solicitada': False,  # Reseta se estava em correção
            'status': 'SUBMETIDO',
            'updated_at': datetime.now(timezone.utc)
        }}
    )
    
    # Notificar admins sobre nova submissão
    admins = await db.users.find({'is_admin': True}, {'_id': 0, 'email': 1}).to_list(50)
    for admin in admins:
        enviar_notificacao_mrosc(
            admin['email'],
            f"Nova Prestação de Contas Submetida",
            projeto.get('nome_projeto', ''),
            f"O usuário {user_name} submeteu a prestação de contas do projeto para análise.",
            "Acesse o sistema para receber e analisar a documentação."
        )
    
    return {'message': 'Prestação de contas submetida com sucesso! Aguarde análise do gestor.'}


@mrosc_router.post("/projetos/{projeto_id}/receber")
async def receber_prestacao_contas(projeto_id: str, request: Request):
    """
    Administrador recebe/confirma recebimento da prestação de contas.
    Muda status para EM_ANALISE.
    """
    user = await get_current_user(request)
    
    # Verificar se é admin
    user_is_admin = user.is_admin if hasattr(user, 'is_admin') else user.get('is_admin', False)
    if not user_is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem receber prestação de contas")
    
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    if not projeto.get('submetido'):
        raise HTTPException(status_code=400, detail="Projeto ainda não foi submetido")
    
    user_name = user.name if hasattr(user, 'name') else user.get('name', 'Admin')
    
    await db.mrosc_projetos.update_one(
        {'projeto_id': projeto_id},
        {'$set': {
            'status': 'EM_ANALISE',
            'recebido_por': user_name,
            'data_recebimento': datetime.now(timezone.utc),
            'updated_at': datetime.now(timezone.utc)
        }}
    )
    
    return {'message': 'Prestação de contas recebida. Iniciando análise.'}


@mrosc_router.post("/projetos/{projeto_id}/solicitar-correcao")
async def solicitar_correcao_prestacao(projeto_id: str, request: Request, data: PedidoCorrecaoRequest):
    """
    Administrador solicita correção. Habilita edição para o usuário externo.
    """
    user = await get_current_user(request)
    
    user_is_admin = user.is_admin if hasattr(user, 'is_admin') else user.get('is_admin', False)
    if not user_is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem solicitar correção")
    
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    user_name = user.name if hasattr(user, 'name') else user.get('name', 'Admin')
    
    await db.mrosc_projetos.update_one(
        {'projeto_id': projeto_id},
        {'$set': {
            'correcao_solicitada': True,
            'data_correcao_solicitada': datetime.now(timezone.utc),
            'motivo_correcao': data.motivo,
            'solicitado_por': user_name,
            'pode_editar': True,  # Habilita edição novamente
            'status': 'CORRECAO_SOLICITADA',
            'updated_at': datetime.now(timezone.utc)
        }}
    )
    
    # Notificar criador do projeto sobre correção solicitada
    criador = await db.users.find_one({'user_id': projeto.get('user_id')}, {'_id': 0, 'email': 1, 'name': 1})
    if criador:
        enviar_notificacao_mrosc(
            criador['email'],
            f"Correção Solicitada em Prestação de Contas",
            projeto.get('nome_projeto', ''),
            f"O gestor {user_name} solicitou correções no seu projeto.<br><br><strong>Motivo:</strong> {data.motivo}",
            "Acesse o sistema para fazer as correções necessárias e resubmeter."
        )
    
    return {'message': 'Correção solicitada. O usuário externo pode editar novamente.'}


@mrosc_router.post("/projetos/{projeto_id}/aprovar")
async def aprovar_prestacao_contas(projeto_id: str, request: Request, data: AprovarPrestacaoRequest = None):
    """
    Administrador aprova a prestação de contas.
    """
    user = await get_current_user(request)
    
    user_is_admin = user.is_admin if hasattr(user, 'is_admin') else user.get('is_admin', False)
    if not user_is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem aprovar prestação de contas")
    
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    if not projeto.get('submetido'):
        raise HTTPException(status_code=400, detail="Projeto precisa ser submetido antes de aprovar")
    
    user_name = user.name if hasattr(user, 'name') else user.get('name', 'Admin')
    
    await db.mrosc_projetos.update_one(
        {'projeto_id': projeto_id},
        {'$set': {
            'aprovado': True,
            'data_aprovacao': datetime.now(timezone.utc),
            'aprovado_por': user_name,
            'observacoes_aprovacao': data.observacoes if data else None,
            'pode_editar': False,
            'correcao_solicitada': False,
            'status': 'APROVADO',
            'updated_at': datetime.now(timezone.utc)
        }}
    )
    
    # Notificar criador do projeto sobre aprovação
    criador = await db.users.find_one({'user_id': projeto.get('user_id')}, {'_id': 0, 'email': 1, 'name': 1})
    if criador:
        obs_msg = f"<br><br><strong>Observações:</strong> {data.observacoes}" if data and data.observacoes else ""
        enviar_notificacao_mrosc(
            criador['email'],
            f"🎉 Prestação de Contas APROVADA!",
            projeto.get('nome_projeto', ''),
            f"Parabéns! Sua prestação de contas foi aprovada pelo gestor {user_name}.{obs_msg}",
            "A prestação de contas foi concluída com sucesso."
        )
    
    return {'message': 'Prestação de contas aprovada com sucesso!'}


@mrosc_router.get("/projetos/{projeto_id}/historico")
async def get_historico_prestacao(projeto_id: str, request: Request):
    """
    Retorna o histórico de ações do workflow do projeto
    """
    user = await get_current_user(request)
    
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    historico = []
    
    if projeto.get('created_at'):
        historico.append({
            'acao': 'Projeto Criado',
            'data': projeto.get('created_at'),
            'usuario': None
        })
    
    if projeto.get('data_submissao'):
        historico.append({
            'acao': 'Submetido para Análise',
            'data': projeto.get('data_submissao'),
            'usuario': projeto.get('submetido_por')
        })
    
    if projeto.get('data_recebimento'):
        historico.append({
            'acao': 'Recebido pelo Gestor',
            'data': projeto.get('data_recebimento'),
            'usuario': projeto.get('recebido_por')
        })
    
    if projeto.get('data_correcao_solicitada'):
        historico.append({
            'acao': 'Correção Solicitada',
            'data': projeto.get('data_correcao_solicitada'),
            'usuario': projeto.get('solicitado_por'),
            'motivo': projeto.get('motivo_correcao')
        })
    
    if projeto.get('data_aprovacao'):
        historico.append({
            'acao': 'Aprovado',
            'data': projeto.get('data_aprovacao'),
            'usuario': projeto.get('aprovado_por'),
            'observacoes': projeto.get('observacoes_aprovacao')
        })
    
    # Ordenar por data
    historico.sort(key=lambda x: x.get('data') or datetime.min, reverse=True)
    
    return {
        'projeto_id': projeto_id,
        'status_atual': projeto.get('status'),
        'pode_editar': projeto.get('pode_editar', True),
        'historico': historico
    }


# ===== ASSINATURA DE DOCUMENTOS MROSC =====
@mrosc_router.post("/projetos/{projeto_id}/assinar")
async def assinar_documento_mrosc(projeto_id: str, signature_request: SignatureRequest, request: Request):
    """
    Endpoint para assinar documento MROSC com confirmação.
    
    Requer confirmação explícita e permite data retroativa/futura.
    
    Args:
        projeto_id: ID do projeto
        signature_request: Dados da assinatura (confirmar_assinatura, data_assinatura, observacoes)
    
    Returns:
        PDF assinado com a data especificada
    """
    user = await get_current_user(request)
    
    # Verificar se a confirmação foi dada
    if not signature_request.confirmar_assinatura:
        # Retornar informações para confirmação
        projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
        if not projeto:
            raise HTTPException(status_code=404, detail="Projeto não encontrado")
        
        # Buscar dados de assinatura do usuário
        user_doc = await db.users.find_one({'user_id': user.user_id}, {'_id': 0})
        user_signature = user_doc.get('signature_data') or {} if user_doc else {}
        
        cpf = user_signature.get('cpf', '').strip()
        cargo = user_signature.get('cargo', '').strip()
        
        if not cpf or not cargo:
            return {
                "status": "DADOS_INCOMPLETOS",
                "message": "Para assinar documentos, você precisa preencher seu CPF e Cargo no perfil.",
                "cpf_preenchido": bool(cpf),
                "cargo_preenchido": bool(cargo),
                "redirect_to": "/perfil"
            }
        
        return {
            "status": "AGUARDANDO_CONFIRMACAO",
            "message": "Você está prestes a assinar digitalmente este documento. Esta ação é irreversível.",
            "projeto": {
                "id": projeto_id,
                "nome": projeto.get('nome_projeto'),
                "organizacao": projeto.get('organizacao_parceira'),
                "valor_total": projeto.get('valor_total')
            },
            "assinante": {
                "nome": user.name,
                "cpf_mascarado": mask_cpf(cpf),
                "cargo": cargo,
                "email": user.email
            },
            "data_sugerida": datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M:%S'),
            "instrucoes": "Envie novamente com 'confirmar_assinatura': true e opcionalmente 'data_assinatura': 'DD/MM/YYYY HH:MM:SS' para assinar."
        }
    
    # Confirmar assinatura
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Buscar dados de assinatura do usuário
    user_doc = await db.users.find_one({'user_id': user.user_id}, {'_id': 0})
    user_signature = user_doc.get('signature_data') or {} if user_doc else {}
    
    cpf = user_signature.get('cpf', '').strip()
    cargo = user_signature.get('cargo', '').strip()
    
    if not cpf:
        raise HTTPException(
            status_code=400, 
            detail="CPF é obrigatório para assinar documentos. Por favor, atualize seu perfil."
        )
    
    if not cargo:
        raise HTTPException(
            status_code=400, 
            detail="Cargo é obrigatório para assinar documentos. Por favor, atualize seu perfil."
        )
    
    # Validar formato da data se fornecida
    signature_date = signature_request.data_assinatura
    if signature_date:
        try:
            # Validar formato DD/MM/YYYY HH:MM:SS
            datetime.strptime(signature_date, '%d/%m/%Y %H:%M:%S')
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="Formato de data inválido. Use: DD/MM/YYYY HH:MM:SS"
            )
    
    return {
        "status": "ASSINATURA_CONFIRMADA",
        "message": "Use os endpoints de PDF para gerar o documento assinado.",
        "projeto_id": projeto_id,
        "data_assinatura": signature_date or datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M:%S'),
        "assinante": user.name,
        "endpoints": {
            "pdf_simples": f"/api/mrosc/projetos/{projeto_id}/relatorio/pdf?assinar=true&data={signature_date or ''}",
            "pdf_consolidado": f"/api/mrosc/projetos/{projeto_id}/relatorio/consolidado/pdf?assinar=true&data={signature_date or ''}"
        }
    }

# ===== RELATÓRIO PDF DE PRESTAÇÃO DE CONTAS MROSC =====
@mrosc_router.get("/projetos/{projeto_id}/relatorio/pdf")
async def export_relatorio_mrosc_pdf(projeto_id: str, request: Request, assinar: bool = False, data: str = None):
    """
    Gera relatório PDF consolidado de prestação de contas conforme MROSC (Lei 13.019/2014)
    Inclui: dados do projeto, recursos humanos, despesas e documentos anexados
    """
    user = await get_current_user(request)
    
    # Buscar dados do projeto
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    # Buscar RH, despesas e documentos
    rhs = await db.mrosc_rh.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    despesas = await db.mrosc_despesas.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(1000)
    documentos = await db.mrosc_documentos.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(500)
    
    # Calcular totais
    total_rh = sum(rh.get('custo_total_projeto', 0) for rh in rhs)
    total_despesas = sum(d.get('valor_total', 0) for d in despesas)
    total_geral = total_rh + total_despesas
    total_docs_valor = sum(d.get('valor', 0) for d in documentos)
    
    buffer = BytesIO()
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='RELATÓRIO DE PRESTAÇÃO DE CONTAS',
        subtitulo='Lei 13.019/2014 - Marco Regulatório das Organizações da Sociedade Civil (MROSC)',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15*mm,
        rightMargin=15*mm,
        topMargin=35*mm,  # Espaço para cabeçalho oficial
        bottomMargin=25*mm,  # Espaço para rodapé oficial
        title=f'Prestação de Contas - {projeto["nome_projeto"]}'
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos customizados
    titulo_style = ParagraphStyle(
        'TituloMROSC',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        alignment=TA_CENTER,
        spaceAfter=6*mm
    )
    
    subtitulo_style = ParagraphStyle(
        'SubtituloMROSC',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#2E7D32'),
        spaceBefore=8*mm,
        spaceAfter=4*mm
    )
    
    normal_style = ParagraphStyle(
        'NormalMROSC',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )
    
    # ===== DADOS DO PROJETO =====
    elements.append(Paragraph('1. IDENTIFICAÇÃO DO PROJETO', subtitulo_style))
    
    projeto_data = [
        ['Nome do Projeto:', projeto.get('nome_projeto', '-')],
        ['Objeto:', Paragraph(projeto.get('objeto', '-'), normal_style)],
        ['Organização Parceira (OSC):', projeto.get('organizacao_parceira', '-')],
        ['CNPJ da OSC:', projeto.get('cnpj_parceira', '-')],
        ['Responsável OSC:', projeto.get('responsavel_osc', '-')],
        ['Período:', f"{projeto.get('data_inicio').strftime('%d/%m/%Y') if projeto.get('data_inicio') else '-'} a {projeto.get('data_conclusao').strftime('%d/%m/%Y') if projeto.get('data_conclusao') else '-'}"],
        ['Prazo:', f"{projeto.get('prazo_meses', 0)} meses"],
        ['Status:', projeto.get('status', '-')]
    ]
    
    projeto_table = Table(projeto_data, colWidths=[50*mm, 120*mm])
    projeto_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C8E6C9')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(projeto_table)
    
    # ===== RESUMO FINANCEIRO =====
    elements.append(Paragraph('2. RESUMO FINANCEIRO', subtitulo_style))
    
    financeiro_data = [
        ['Descrição', 'Valor (R$)'],
        ['Valor Total do Projeto', f"R$ {projeto.get('valor_total', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Repasse Público', f"R$ {projeto.get('valor_repasse_publico', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Contrapartida', f"R$ {projeto.get('valor_contrapartida', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Total Recursos Humanos', f"R$ {total_rh:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Total Despesas', f"R$ {total_despesas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Total Executado', f"R$ {total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Saldo', f"R$ {(projeto.get('valor_total', 0) - total_geral):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
    ]
    
    fin_table = Table(financeiro_data, colWidths=[100*mm, 70*mm])
    fin_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(fin_table)
    
    # ===== RECURSOS HUMANOS =====
    if rhs:
        elements.append(Paragraph('3. RECURSOS HUMANOS', subtitulo_style))
        
        rh_header = ['Função', 'Regime', 'Salário', 'Encargos', 'Meses', 'Custo Total']
        rh_data = [rh_header]
        
        for rh in rhs:
            encargos = rh.get('provisao_ferias', 0) + rh.get('provisao_13_salario', 0) + rh.get('fgts', 0) + rh.get('inss_patronal', 0)
            rh_data.append([
                rh.get('nome_funcao', '-'),
                rh.get('regime_contratacao', '-'),
                f"R$ {rh.get('salario_bruto', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {encargos:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                str(rh.get('numero_meses', 0)),
                f"R$ {rh.get('custo_total_projeto', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        rh_table = Table(rh_data, colWidths=[45*mm, 20*mm, 30*mm, 30*mm, 15*mm, 30*mm])
        rh_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#90CAF9')),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E3F2FD')]),
        ]))
        elements.append(rh_table)
    
    # ===== DESPESAS =====
    if despesas:
        elements.append(Paragraph('4. DESPESAS', subtitulo_style))
        
        desp_header = ['Natureza', 'Descrição', 'Qtd', 'Unit.', 'Total']
        desp_data = [desp_header]
        
        for d in despesas:
            desp_data.append([
                d.get('natureza_despesa', '-'),
                Paragraph(d.get('descricao', '-')[:50], ParagraphStyle('DescDesp', fontSize=7)),
                f"{d.get('quantidade', 0)} {d.get('unidade', '')}",
                f"R$ {d.get('valor_unitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {d.get('valor_total', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        desp_table = Table(desp_data, colWidths=[25*mm, 65*mm, 25*mm, 25*mm, 30*mm])
        desp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E65100')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#FFCC80')),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF3E0')]),
        ]))
        elements.append(desp_table)
    
    # ===== DOCUMENTOS =====
    if documentos:
        elements.append(Paragraph('5. DOCUMENTOS COMPROBATÓRIOS', subtitulo_style))
        
        docs_header = ['Tipo', 'Número', 'Data', 'Valor', 'Status']
        docs_data = [docs_header]
        
        for documento in documentos:
            data_doc = documento.get('data_documento', None)
            if data_doc:
                if isinstance(data_doc, datetime):
                    data_doc = data_doc.strftime('%d/%m/%Y')
                else:
                    data_doc = str(data_doc)[:10]
            else:
                data_doc = '-'
            docs_data.append([
                documento.get('tipo_documento', '-'),
                documento.get('numero_documento', '-') or '-',
                data_doc,
                f"R$ {documento.get('valor', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                'Validado' if documento.get('validado') else 'Pendente'
            ])
        
        docs_table = Table(docs_data, colWidths=[35*mm, 40*mm, 30*mm, 35*mm, 30*mm])
        docs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6A1B9A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CE93D8')),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3E5F5')]),
        ]))
        elements.append(docs_table)
        
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph(f"<b>Total de documentos:</b> {len(documentos)} | <b>Validados:</b> {len([d for d in documentos if d.get('validado')])} | <b>Valor comprovado:</b> R$ {total_docs_valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), normal_style))
    
    # ===== ASSINATURA =====
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph('_' * 60, ParagraphStyle('Linha', alignment=TA_CENTER)))
    user_name = user.name if hasattr(user, 'name') else user.get('name', 'Responsável') if isinstance(user, dict) else 'Responsável'
    elements.append(Paragraph(f'<b>{user_name}</b>', ParagraphStyle('Assinatura', alignment=TA_CENTER, fontSize=10)))
    elements.append(Paragraph(f'Responsável pela Prestação de Contas', ParagraphStyle('Cargo', alignment=TA_CENTER, fontSize=8, textColor=colors.gray)))
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    # Adicionar assinatura digital ao PDF
    try:
        doc_info = {
            'tipo': 'PRESTAÇÃO DE CONTAS MROSC',
            'id': projeto_id,
            'titulo': f"Prestação de Contas - {projeto.get('nome_projeto', 'N/A')}"
        }
        signed_buffer, validation_code = await add_signature_to_pdf(
            buffer, 
            user, 
            'MROSC_PRESTACAO_CONTAS', 
            projeto_id,
            doc_info,
            data  # Data da assinatura (pode ser retroativa)
        )
        buffer = signed_buffer
        logging.info(f"PDF de prestação de contas MROSC assinado com código: {validation_code}")
    except HTTPException as e:
        # Se falhar por falta de dados de assinatura, retorna PDF sem assinatura
        logging.warning(f"PDF sem assinatura digital: {e.detail}")
    except Exception as e:
        logging.error(f"Erro ao adicionar assinatura ao PDF MROSC: {e}")
    
    filename = f"Prestacao_Contas_{projeto['nome_projeto'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@mrosc_router.get("/projetos/{projeto_id}/relatorio/consolidado/pdf")
@mrosc_router.get("/projetos/{projeto_id}/relatorio/consolidado")
async def gerar_relatorio_consolidado_mrosc(projeto_id: str, request: Request, assinar: bool = False, data: str = None):
    """
    Gera relatório PDF consolidado com todos os anexos incorporados
    Inclui: dados do projeto, concedente, RH, despesas, histórico e documentos
    """
    from PyPDF2 import PdfMerger
    from PIL import Image as PILImage
    from reportlab.platypus import PageBreak
    
    user = await get_current_user(request)
    
    projeto = await db.mrosc_projetos.find_one({'projeto_id': projeto_id}, {'_id': 0})
    if not projeto:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    
    rhs = await db.mrosc_rh.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    despesas = await db.mrosc_despesas.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    documentos = await db.mrosc_documentos.find({'projeto_id': projeto_id}, {'_id': 0}).to_list(100)
    historico = await db.mrosc_historico.find({'projeto_id': projeto_id}, {'_id': 0}).sort('data', 1).to_list(100)
    
    total_rh = sum(rh.get('custo_total_projeto', 0) for rh in rhs)
    total_despesas = sum(d.get('valor_total', 0) for d in despesas)
    total_geral = total_rh + total_despesas
    
    # Create main PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1F4E78'), alignment=TA_CENTER, spaceAfter=5)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#2E7D32'), spaceBefore=12, spaceAfter=6)
    
    # ===== CAPA =====
    elements.append(Spacer(1, 30*mm))
    elements.append(Paragraph("PREFEITURA MUNICIPAL DE ACAIACA", title_style))
    elements.append(Paragraph("Estado de Minas Gerais", ParagraphStyle('Estado', fontSize=10, alignment=TA_CENTER)))
    elements.append(Spacer(1, 20*mm))
    elements.append(Paragraph("PRESTAÇÃO DE CONTAS CONSOLIDADA", ParagraphStyle('TitDoc', fontSize=18, alignment=TA_CENTER, textColor=colors.HexColor('#2E7D32'))))
    elements.append(Paragraph("MROSC - Lei 13.019/2014", ParagraphStyle('Lei', fontSize=12, alignment=TA_CENTER, textColor=colors.gray)))
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"<b>{projeto.get('nome_projeto', 'N/A')}</b>", ParagraphStyle('NomeProjeto', fontSize=14, alignment=TA_CENTER)))
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph(f"Organização: {projeto.get('organizacao_parceira', 'N/A')}", ParagraphStyle('Org', fontSize=11, alignment=TA_CENTER)))
    elements.append(Paragraph(f"CNPJ: {projeto.get('cnpj_parceira', 'N/A')}", ParagraphStyle('CNPJ', fontSize=10, alignment=TA_CENTER)))
    elements.append(Spacer(1, 30*mm))
    elements.append(Paragraph(f"VALOR TOTAL: R$ {total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), ParagraphStyle('Valor', fontSize=14, alignment=TA_CENTER, textColor=colors.HexColor('#1F4E78'))))
    elements.append(Spacer(1, 20*mm))
    elements.append(Paragraph(f"Acaiaca - MG, {datetime.now(timezone.utc).strftime('%d/%m/%Y')}", ParagraphStyle('Data', fontSize=10, alignment=TA_CENTER)))
    elements.append(PageBreak())
    
    # ===== 1. DADOS DO PROJETO =====
    elements.append(Paragraph("1. DADOS DO PROJETO", section_style))
    
    projeto_data = [
        ['Campo', 'Valor'],
        ['Nome do Projeto', projeto.get('nome_projeto', 'N/A')],
        ['Objeto', str(projeto.get('objeto', projeto.get('objeto_detalhado', 'N/A')))[:100]],
        ['Organização Parceira', projeto.get('organizacao_parceira', 'N/A')],
        ['CNPJ da Parceira', projeto.get('cnpj_parceira', 'N/A')],
        ['Responsável OSC', projeto.get('responsavel_osc', 'N/A')],
        ['Valor Repasse Público', f"R$ {projeto.get('valor_repasse_publico', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Valor Contrapartida', f"R$ {projeto.get('valor_contrapartida', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Status', projeto.get('status', 'ELABORACAO')]
    ]
    
    projeto_table = Table(projeto_data, colWidths=[120, 340])
    projeto_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(projeto_table)
    
    # ===== 2. RECURSOS HUMANOS =====
    elements.append(Paragraph("2. RECURSOS HUMANOS", section_style))
    
    if rhs:
        rh_table_data = [['#', 'Função', 'Regime', 'Salário Bruto', 'Custo Mensal', 'Meses', 'Custo Total']]
        
        for i, rh in enumerate(rhs, 1):
            rh_table_data.append([
                str(i),
                rh.get('nome_funcao', '')[:25],
                rh.get('regime_contratacao', '')[:10],
                f"R$ {rh.get('salario_bruto', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {rh.get('custo_mensal_total', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                str(rh.get('numero_meses', '')),
                f"R$ {rh.get('custo_total_projeto', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        rh_table_data.append(['', '', '', '', '', 'TOTAL:', f"R$ {total_rh:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        rh_table = Table(rh_table_data, colWidths=[20, 90, 50, 70, 70, 35, 80])
        rh_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E3F2FD')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(rh_table)
    else:
        elements.append(Paragraph("Nenhum recurso humano cadastrado.", ParagraphStyle('Empty', fontSize=10, textColor=colors.gray)))
    
    # ===== 3. DESPESAS =====
    elements.append(Paragraph("3. DESPESAS", section_style))
    
    if despesas:
        desp_table_data = [['#', 'Item', 'Natureza', 'Qtd', 'V.Unit', 'V.Total']]
        
        for i, d in enumerate(despesas, 1):
            desp_table_data.append([
                str(i),
                d.get('item_despesa', '')[:30],
                d.get('natureza_despesa', '')[:8],
                str(d.get('quantidade', 0)),
                f"R$ {d.get('valor_unitario', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"R$ {d.get('valor_total', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        desp_table_data.append(['', '', '', '', 'TOTAL:', f"R$ {total_despesas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        desp_table = Table(desp_table_data, colWidths=[20, 150, 50, 40, 70, 80])
        desp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F5E9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(desp_table)
    else:
        elements.append(Paragraph("Nenhuma despesa cadastrada.", ParagraphStyle('Empty', fontSize=10, textColor=colors.gray)))
    
    # ===== 4. RESUMO FINANCEIRO =====
    elements.append(Paragraph("4. RESUMO FINANCEIRO", section_style))
    
    resumo_data = [
        ['Categoria', 'Valor'],
        ['Total Recursos Humanos', f"R$ {total_rh:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Total Despesas', f"R$ {total_despesas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['TOTAL GERAL', f"R$ {total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
    ]
    
    resumo_table = Table(resumo_data, colWidths=[200, 150])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6F00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF3E0')),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(resumo_table)
    
    # ===== 5. DOCUMENTOS ANEXADOS =====
    elements.append(Paragraph("5. DOCUMENTOS ANEXADOS", section_style))
    
    if documentos:
        doc_table_data = [['#', 'Tipo', 'Número', 'Valor', 'Validado']]
        
        for i, documento in enumerate(documentos, 1):
            doc_table_data.append([
                str(i),
                documento.get('tipo_documento', '')[:20],
                documento.get('numero_documento', '')[:20],
                f"R$ {documento.get('valor', 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if documento.get('valor') else '-',
                '✓' if documento.get('validado') else '✗'
            ])
        
        doc_list_table = Table(doc_table_data, colWidths=[25, 100, 100, 70, 45])
        doc_list_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7B1FA2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(doc_list_table)
        elements.append(Spacer(1, 5))
        elements.append(Paragraph("<i>Os documentos originais estão anexados nas páginas seguintes.</i>", ParagraphStyle('Nota', fontSize=8, textColor=colors.gray)))
    else:
        elements.append(Paragraph("Nenhum documento anexado.", ParagraphStyle('Empty', fontSize=10, textColor=colors.gray)))
    
    # ===== FOOTER =====
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("_" * 60, ParagraphStyle('Linha', alignment=TA_CENTER)))
    elements.append(Paragraph("Documento gerado automaticamente pelo Sistema Planejamento Acaiaca", ParagraphStyle('Footer', fontSize=8, alignment=TA_CENTER, textColor=colors.gray)))
    elements.append(Paragraph(f"Data: {datetime.now(timezone.utc).strftime('%d/%m/%Y às %H:%M:%S')}", ParagraphStyle('Footer2', fontSize=8, alignment=TA_CENTER, textColor=colors.gray)))
    
    # Build main PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Merge with attachments
    merger = PdfMerger()
    merger.append(buffer)
    
    upload_dir = Path('/app/backend/uploads/mrosc')
    
    for documento in documentos:
        doc_id = documento.get('documento_id')
        arquivo_nome = documento.get('arquivo_nome', '')
        ext = arquivo_nome.split('.')[-1].lower() if '.' in arquivo_nome else ''
        
        filepath = upload_dir / projeto_id / f"{doc_id}.{ext}"
        
        if filepath.exists():
            try:
                if ext == 'pdf':
                    merger.append(str(filepath))
                elif ext in ['jpg', 'jpeg', 'png']:
                    # Convert image to PDF
                    img = PILImage.open(filepath)
                    if img.mode in ('RGBA', 'LA', 'P'):
                        img = img.convert('RGB')
                    
                    max_width = 1800
                    if img.width > max_width:
                        ratio = max_width / img.width
                        new_size = (int(img.width * ratio), int(img.height * ratio))
                        img = img.resize(new_size, PILImage.Resampling.LANCZOS)
                    
                    img_pdf = BytesIO()
                    img.save(img_pdf, format='PDF', resolution=100.0)
                    img_pdf.seek(0)
                    merger.append(img_pdf)
            except Exception as e:
                logging.error(f"Erro ao processar anexo {doc_id}: {e}")
                continue
    
    final_buffer = BytesIO()
    merger.write(final_buffer)
    merger.close()
    final_buffer.seek(0)
    
    # Adicionar assinatura digital ao PDF consolidado
    try:
        doc_info = {
            'tipo': 'RELATÓRIO CONSOLIDADO MROSC',
            'id': projeto_id,
            'titulo': f"Relatório Consolidado - {projeto.get('nome_projeto', 'N/A')}"
        }
        signed_buffer, validation_code = await add_signature_to_pdf(
            final_buffer, 
            user, 
            'MROSC_CONSOLIDADO', 
            projeto_id,
            doc_info,
            data  # Data da assinatura (pode ser retroativa)
        )
        final_buffer = signed_buffer
        logging.info(f"PDF consolidado MROSC assinado com código: {validation_code}")
    except HTTPException as e:
        # Se falhar por falta de dados de assinatura, retorna PDF sem assinatura
        logging.warning(f"PDF consolidado sem assinatura digital: {e.detail}")
    except Exception as e:
        logging.error(f"Erro ao adicionar assinatura ao PDF consolidado MROSC: {e}")
    
    filename = f"MROSC_Consolidado_{projeto.get('nome_projeto', 'projeto')[:15].replace(' ', '_')}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        final_buffer,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# Registrar router MROSC
app.include_router(mrosc_router)

# ===== DASHBOARD ANALÍTICO E SISTEMA DE ALERTAS =====
analytics_router = APIRouter(prefix="/api/analytics", tags=["Dashboard Analítico"])

@analytics_router.get("/dashboard")
async def get_dashboard_analytics(request: Request):
    """
    Dashboard analítico consolidado com métricas de todos os módulos
    """
    user = await get_current_user(request)
    
    # Estatísticas de PACs
    pacs = await db.pacs.find({}, {'_id': 0}).to_list(1000)
    pac_items = await db.pac_items.find({}, {'_id': 0}).to_list(10000)
    
    # Estatísticas de PACs Geral
    pacs_geral = await db.pacs_geral.find({}, {'_id': 0}).to_list(100)
    pac_geral_items = await db.pac_geral_items.find({}, {'_id': 0}).to_list(10000)
    
    # Estatísticas de PACs Obras
    pacs_obras = await db.pacs_geral_obras.find({}, {'_id': 0}).to_list(100)
    pac_obras_items = await db.pac_geral_obras_items.find({}, {'_id': 0}).to_list(10000)
    
    # Estatísticas de Processos
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    # Estatísticas de MROSC
    projetos_mrosc = await db.mrosc_projetos.find({}, {'_id': 0}).to_list(100)
    rh_mrosc = await db.mrosc_rh.find({}, {'_id': 0}).to_list(500)
    despesas_mrosc = await db.mrosc_despesas.find({}, {'_id': 0}).to_list(5000)
    docs_mrosc = await db.mrosc_documentos.find({}, {'_id': 0}).to_list(1000)
    
    # Cálculos
    total_pac = sum(item.get('valorTotal', 0) for item in pac_items)
    total_pac_geral = sum(item.get('valorTotal', 0) for item in pac_geral_items)
    total_pac_obras = sum(item.get('valorTotal', 0) for item in pac_obras_items)
    total_mrosc = sum(p.get('valor_total', 0) for p in projetos_mrosc)
    total_rh_mrosc = sum(rh.get('custo_total_projeto', 0) for rh in rh_mrosc)
    total_despesas_mrosc = sum(d.get('valor_total', 0) for d in despesas_mrosc)
    
    # Processos por status
    processos_por_status = {}
    for p in processos:
        status = p.get('status', 'Não Definido')
        processos_por_status[status] = processos_por_status.get(status, 0) + 1
    
    # Execução orçamentária MROSC
    execucao_mrosc = []
    for proj in projetos_mrosc:
        proj_id = proj.get('projeto_id')
        rh_proj = sum(r.get('custo_total_projeto', 0) for r in rh_mrosc if r.get('projeto_id') == proj_id)
        desp_proj = sum(d.get('valor_total', 0) for d in despesas_mrosc if d.get('projeto_id') == proj_id)
        docs_proj = len([d for d in docs_mrosc if d.get('projeto_id') == proj_id])
        total_exec = rh_proj + desp_proj
        percentual = (total_exec / proj.get('valor_total', 1)) * 100 if proj.get('valor_total') else 0
        execucao_mrosc.append({
            'projeto_id': proj_id,
            'nome': proj.get('nome_projeto', ''),
            'valor_total': proj.get('valor_total', 0),
            'executado': total_exec,
            'rh': rh_proj,
            'despesas': desp_proj,
            'documentos': docs_proj,
            'percentual': round(percentual, 1),
            'saldo': proj.get('valor_total', 0) - total_exec
        })
    
    # Top 5 itens por valor (PAC Geral)
    top_items = sorted(pac_geral_items, key=lambda x: x.get('valorTotal', 0), reverse=True)[:5]
    
    # Distribuição por secretaria
    distribuicao_secretarias = {}
    for item in pac_geral_items:
        for sec in ['AD', 'FA', 'SA', 'SE', 'AS', 'AG', 'OB', 'TR', 'CUL']:
            qtd = item.get(f'qtd_{sec.lower()}', 0)
            if qtd > 0:
                valor = qtd * item.get('valorUnitario', 0)
                if sec not in distribuicao_secretarias:
                    distribuicao_secretarias[sec] = {'quantidade': 0, 'valor': 0}
                distribuicao_secretarias[sec]['quantidade'] += qtd
                distribuicao_secretarias[sec]['valor'] += valor
    
    sec_labels = {
        'AD': 'Administração', 'FA': 'Fazenda', 'SA': 'Saúde', 'SE': 'Educação',
        'AS': 'Assist. Social', 'AG': 'Agricultura', 'OB': 'Obras', 'TR': 'Transporte', 'CUL': 'Cultura'
    }
    
    distribuicao_chart = [
        {'secretaria': sec_labels.get(k, k), 'codigo': k, 'valor': v['valor'], 'quantidade': v['quantidade']}
        for k, v in distribuicao_secretarias.items()
    ]
    distribuicao_chart.sort(key=lambda x: x['valor'], reverse=True)
    
    return {
        'resumo': {
            'total_geral': total_pac + total_pac_geral + total_pac_obras,
            'total_pac_individual': total_pac,
            'total_pac_geral': total_pac_geral,
            'total_pac_obras': total_pac_obras,
            'total_mrosc': total_mrosc,
            'total_executado_mrosc': total_rh_mrosc + total_despesas_mrosc
        },
        'contadores': {
            'pacs': len(pacs),
            'pacs_geral': len(pacs_geral),
            'pacs_obras': len(pacs_obras),
            'itens_pac': len(pac_items),
            'itens_pac_geral': len(pac_geral_items),
            'itens_pac_obras': len(pac_obras_items),
            'processos': len(processos),
            'projetos_mrosc': len(projetos_mrosc),
            'funcionarios_mrosc': len(rh_mrosc),
            'despesas_mrosc': len(despesas_mrosc),
            'documentos_mrosc': len(docs_mrosc)
        },
        'processos_por_status': [
            {'status': k, 'quantidade': v} for k, v in processos_por_status.items()
        ],
        'execucao_mrosc': execucao_mrosc,
        'top_itens': [
            {'descricao': i.get('descricao', '')[:50], 'valor': i.get('valorTotal', 0), 'catmat': i.get('catmat', '')}
            for i in top_items
        ],
        'distribuicao_secretarias': distribuicao_chart
    }


@analytics_router.get("/realtime")
async def get_realtime_metrics(request: Request):
    """
    Métricas em tempo real para monitoramento do sistema
    Inclui: uso por secretaria, horários de pico, tendências de gastos
    """
    user = await get_current_user(request)
    
    from datetime import timedelta
    from collections import defaultdict
    
    hoje = datetime.now(timezone.utc)
    inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    inicio_semana = hoje - timedelta(days=hoje.weekday())
    
    # ===== USO POR SECRETARIA =====
    pacs = await db.pacs.find({}, {'_id': 0}).to_list(1000)
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    uso_secretaria = defaultdict(lambda: {'pacs': 0, 'processos': 0, 'valor_total': 0, 'ultimo_acesso': None})
    
    for pac in pacs:
        sec = pac.get('secretaria', 'Não Definida')
        uso_secretaria[sec]['pacs'] += 1
        created = pac.get('created_at') or pac.get('updated_at')
        if created and (not uso_secretaria[sec]['ultimo_acesso'] or created > uso_secretaria[sec]['ultimo_acesso']):
            uso_secretaria[sec]['ultimo_acesso'] = created
    
    for proc in processos:
        sec = proc.get('secretaria', 'Não Definida')
        uso_secretaria[sec]['processos'] += 1
    
    # Calcular valores por secretaria
    pac_items = await db.pac_items.find({}, {'_id': 0}).to_list(10000)
    pac_lookup = {p['pac_id']: p.get('secretaria', 'Não Definida') for p in pacs}
    
    for item in pac_items:
        sec = pac_lookup.get(item.get('pac_id'), 'Não Definida')
        uso_secretaria[sec]['valor_total'] += item.get('valorTotal', 0)
    
    uso_secretaria_list = []
    for sec, data in uso_secretaria.items():
        ultimo = data['ultimo_acesso']
        if isinstance(ultimo, datetime):
            ultimo = ultimo.isoformat()
        uso_secretaria_list.append({
            'secretaria': sec,
            'pacs': data['pacs'],
            'processos': data['processos'],
            'valor_total': data['valor_total'],
            'ultimo_acesso': ultimo
        })
    uso_secretaria_list.sort(key=lambda x: x['valor_total'], reverse=True)
    
    # ===== ATIVIDADE POR HORÁRIO (últimos 7 dias) =====
    atividade_horario = [0] * 24
    
    # Buscar documentos criados nos últimos 7 dias
    sete_dias_atras = hoje - timedelta(days=7)
    
    async def count_by_hour(collection, date_field='created_at'):
        docs = await collection.find({
            date_field: {'$gte': sete_dias_atras}
        }, {'_id': 0, date_field: 1}).to_list(10000)
        
        for doc in docs:
            dt = doc.get(date_field)
            if isinstance(dt, datetime):
                atividade_horario[dt.hour] += 1
    
    await count_by_hour(db.pacs)
    await count_by_hour(db.processos)
    await count_by_hour(db.mrosc_projetos)
    await count_by_hour(db.pac_items)
    
    atividade_horario_data = [
        {'hora': f'{h:02d}:00', 'atividade': atividade_horario[h]}
        for h in range(24)
    ]
    
    # ===== TENDÊNCIA DE GASTOS (últimos 6 meses) =====
    tendencia_gastos = []
    
    for i in range(6, 0, -1):
        mes = hoje - timedelta(days=i*30)
        mes_inicio = mes.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        mes_fim = (mes_inicio + timedelta(days=32)).replace(day=1)
        
        # PAC items criados nesse mês
        items_mes = await db.pac_items.find({
            'created_at': {'$gte': mes_inicio, '$lt': mes_fim}
        }, {'valorTotal': 1, '_id': 0}).to_list(10000)
        
        valor_mes = sum(i.get('valorTotal', 0) for i in items_mes)
        
        tendencia_gastos.append({
            'mes': mes_inicio.strftime('%b/%Y'),
            'valor': valor_mes,
            'quantidade': len(items_mes)
        })
    
    # ===== MÉTRICAS DE DESEMPENHO =====
    total_usuarios = await db.users.count_documents({})
    usuarios_ativos = await db.users.count_documents({'is_active': True})
    
    # Documentos criados hoje
    inicio_hoje = hoje.replace(hour=0, minute=0, second=0, microsecond=0)
    docs_hoje = {
        'pacs': await db.pacs.count_documents({'created_at': {'$gte': inicio_hoje}}),
        'processos': await db.processos.count_documents({'created_at': {'$gte': inicio_hoje}}),
        'pac_items': await db.pac_items.count_documents({'created_at': {'$gte': inicio_hoje}}),
        'mrosc': await db.mrosc_projetos.count_documents({'created_at': {'$gte': inicio_hoje}})
    }
    
    # Documentos criados esta semana
    docs_semana = {
        'pacs': await db.pacs.count_documents({'created_at': {'$gte': inicio_semana}}),
        'processos': await db.processos.count_documents({'created_at': {'$gte': inicio_semana}}),
        'pac_items': await db.pac_items.count_documents({'created_at': {'$gte': inicio_semana}}),
        'mrosc': await db.mrosc_projetos.count_documents({'created_at': {'$gte': inicio_semana}})
    }
    
    # ===== STATUS DOS MÓDULOS =====
    status_modulos = {
        'PAC Individual': {
            'total': len(pacs),
            'itens': len(pac_items),
            'status': 'online'
        },
        'PAC Geral': {
            'total': await db.pacs_geral.count_documents({}),
            'itens': await db.pac_geral_items.count_documents({}),
            'status': 'online'
        },
        'PAC Obras': {
            'total': await db.pacs_geral_obras.count_documents({}),
            'itens': await db.pac_obras_items.count_documents({}),
            'status': 'online'
        },
        'Processos': {
            'total': len(processos),
            'status': 'online'
        },
        'MROSC': {
            'total': await db.mrosc_projetos.count_documents({}),
            'documentos': await db.mrosc_documentos.count_documents({}),
            'status': 'online'
        }
    }
    
    # ===== TOP USUÁRIOS ATIVOS (por documentos criados) =====
    top_usuarios = []
    users = await db.users.find({'is_active': True}, {'_id': 0, 'user_id': 1, 'name': 1, 'email': 1}).to_list(100)
    
    for usr in users:
        uid = usr.get('user_id')
        count = await db.pacs.count_documents({'user_id': uid})
        count += await db.processos.count_documents({'user_id': uid})
        count += await db.mrosc_projetos.count_documents({'user_id': uid})
        if count > 0:
            top_usuarios.append({
                'user_id': uid,
                'name': usr.get('name', 'N/A'),
                'email': usr.get('email', ''),
                'documentos': count
            })
    
    top_usuarios.sort(key=lambda x: x['documentos'], reverse=True)
    top_usuarios = top_usuarios[:10]
    
    return {
        'timestamp': hoje.isoformat(),
        'uso_por_secretaria': uso_secretaria_list,
        'atividade_por_horario': atividade_horario_data,
        'tendencia_gastos': tendencia_gastos,
        'metricas_desempenho': {
            'usuarios_totais': total_usuarios,
            'usuarios_ativos': usuarios_ativos,
            'documentos_hoje': docs_hoje,
            'documentos_semana': docs_semana
        },
        'status_modulos': status_modulos,
        'top_usuarios': top_usuarios,
        'horario_pico': max(atividade_horario_data, key=lambda x: x['atividade']) if any(a['atividade'] for a in atividade_horario_data) else {'hora': '09:00', 'atividade': 0}
    }


# ===== SISTEMA DE ALERTAS =====
alertas_router = APIRouter(prefix="/api/alertas", tags=["Sistema de Alertas"])

class AlertaCreate(BaseModel):
    tipo: str  # PRAZO, DOCUMENTO, CONTRATO, PRESTACAO_CONTAS, SISTEMA
    titulo: str
    mensagem: str
    prioridade: str = "MEDIA"  # BAIXA, MEDIA, ALTA, CRITICA
    modulo: str  # PAC, MROSC, PROCESSO, DOEM
    referencia_id: Optional[str] = None
    data_vencimento: Optional[datetime] = None

@alertas_router.get("/")
async def get_alertas(request: Request):
    """
    Lista todos os alertas ativos do sistema, incluindo alertas automáticos
    """
    user = await get_current_user(request)
    
    alertas = []
    hoje = datetime.now(timezone.utc)
    
    # ===== ALERTAS DE MROSC =====
    projetos_mrosc = await db.mrosc_projetos.find({}, {'_id': 0}).to_list(100)
    docs_mrosc = await db.mrosc_documentos.find({}, {'_id': 0}).to_list(1000)
    
    for proj in projetos_mrosc:
        # Alerta de prazo do projeto
        data_conclusao = proj.get('data_conclusao')
        if data_conclusao:
            if isinstance(data_conclusao, str):
                try:
                    data_conclusao = datetime.fromisoformat(data_conclusao.replace('Z', '+00:00'))
                except:
                    continue
            
            # Garantir que a data tem timezone
            if data_conclusao.tzinfo is None:
                data_conclusao = data_conclusao.replace(tzinfo=timezone.utc)
            
            dias_restantes = (data_conclusao - hoje).days
            
            if dias_restantes < 0:
                alertas.append({
                    'tipo': 'PRAZO',
                    'titulo': 'Projeto MROSC Vencido',
                    'mensagem': f'O projeto "{proj.get("nome_projeto")}" está vencido há {abs(dias_restantes)} dias.',
                    'prioridade': 'CRITICA',
                    'modulo': 'MROSC',
                    'referencia_id': proj.get('projeto_id'),
                    'data_vencimento': data_conclusao.isoformat() if isinstance(data_conclusao, datetime) else data_conclusao,
                    'dias_restantes': dias_restantes
                })
            elif dias_restantes <= 30:
                alertas.append({
                    'tipo': 'PRAZO',
                    'titulo': 'Prazo de Projeto MROSC',
                    'mensagem': f'O projeto "{proj.get("nome_projeto")}" vence em {dias_restantes} dias.',
                    'prioridade': 'ALTA' if dias_restantes <= 7 else 'MEDIA',
                    'modulo': 'MROSC',
                    'referencia_id': proj.get('projeto_id'),
                    'data_vencimento': data_conclusao.isoformat() if isinstance(data_conclusao, datetime) else data_conclusao,
                    'dias_restantes': dias_restantes
                })
        
        # Alerta de documentos pendentes de validação
        docs_pendentes = [d for d in docs_mrosc if d.get('projeto_id') == proj.get('projeto_id') and not d.get('validado')]
        if len(docs_pendentes) > 0:
            alertas.append({
                'tipo': 'DOCUMENTO',
                'titulo': 'Documentos Pendentes de Validação',
                'mensagem': f'{len(docs_pendentes)} documento(s) do projeto "{proj.get("nome_projeto")}" aguardam validação.',
                'prioridade': 'MEDIA',
                'modulo': 'MROSC',
                'referencia_id': proj.get('projeto_id'),
                'quantidade': len(docs_pendentes)
            })
    
    # ===== ALERTAS DE PROCESSOS =====
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    
    for proc in processos:
        # Alerta de processos sem número
        if not proc.get('numero_processo'):
            alertas.append({
                'tipo': 'SISTEMA',
                'titulo': 'Processo sem Número',
                'mensagem': f'O processo "{proc.get("objeto", "")[:50]}" não possui número de processo.',
                'prioridade': 'BAIXA',
                'modulo': 'PROCESSO',
                'referencia_id': proc.get('processo_id')
            })
        
        # Alerta de prazo de processo
        data_abertura = proc.get('data_abertura')
        if data_abertura and proc.get('status') not in ['Concluído', 'Cancelado', 'Homologado']:
            if isinstance(data_abertura, str):
                try:
                    data_abertura = datetime.fromisoformat(data_abertura.replace('Z', '+00:00'))
                except:
                    continue
            
            # Garantir que a data tem timezone
            if data_abertura.tzinfo is None:
                data_abertura = data_abertura.replace(tzinfo=timezone.utc)
            
            dias_aberto = (hoje - data_abertura).days
            if dias_aberto > 90:
                alertas.append({
                    'tipo': 'PRAZO',
                    'titulo': 'Processo Aberto há Muito Tempo',
                    'mensagem': f'O processo "{proc.get("objeto", "")[:40]}" está aberto há {dias_aberto} dias.',
                    'prioridade': 'MEDIA' if dias_aberto < 180 else 'ALTA',
                    'modulo': 'PROCESSO',
                    'referencia_id': proc.get('processo_id'),
                    'dias_aberto': dias_aberto
                })
    
    # Ordenar por prioridade
    prioridade_ordem = {'CRITICA': 0, 'ALTA': 1, 'MEDIA': 2, 'BAIXA': 3}
    alertas.sort(key=lambda x: prioridade_ordem.get(x.get('prioridade', 'MEDIA'), 2))
    
    return {
        'total': len(alertas),
        'criticos': len([a for a in alertas if a.get('prioridade') == 'CRITICA']),
        'altos': len([a for a in alertas if a.get('prioridade') == 'ALTA']),
        'medios': len([a for a in alertas if a.get('prioridade') == 'MEDIA']),
        'baixos': len([a for a in alertas if a.get('prioridade') == 'BAIXA']),
        'alertas': alertas
    }


@alertas_router.get("/resumo")
async def get_alertas_resumo(request: Request):
    """Resumo rápido de alertas para exibição no header"""
    user = await get_current_user(request)
    alertas = await get_alertas(request)
    
    return {
        'total': alertas['total'],
        'criticos': alertas['criticos'],
        'altos': alertas['altos'],
        'tem_alertas_urgentes': alertas['criticos'] > 0 or alertas['altos'] > 0
    }


# ===== RELATÓRIOS GERENCIAIS CONSOLIDADOS =====
relatorios_router = APIRouter(prefix="/api/relatorios", tags=["Relatórios Gerenciais"])

@relatorios_router.get("/consolidado/pdf")
async def export_relatorio_consolidado_pdf(request: Request):
    """
    Gera relatório PDF consolidado de todos os módulos do sistema
    """
    user = await get_current_user(request)
    
    # Verificar se é admin
    user_is_admin = user.is_admin if hasattr(user, 'is_admin') else user.get('is_admin', False)
    if not user_is_admin:
        raise HTTPException(status_code=403, detail="Apenas administradores podem gerar relatórios consolidados")
    
    # Buscar dados de todos os módulos
    pacs = await db.pacs.find({}, {'_id': 0}).to_list(100)
    pac_items = await db.pac_items.find({}, {'_id': 0}).to_list(10000)
    pacs_geral = await db.pacs_geral.find({}, {'_id': 0}).to_list(50)
    pac_geral_items = await db.pac_geral_items.find({}, {'_id': 0}).to_list(10000)
    pacs_obras = await db.pacs_geral_obras.find({}, {'_id': 0}).to_list(50)
    pac_obras_items = await db.pac_geral_obras_items.find({}, {'_id': 0}).to_list(10000)
    processos = await db.processos.find({}, {'_id': 0}).to_list(1000)
    projetos_mrosc = await db.mrosc_projetos.find({}, {'_id': 0}).to_list(100)
    
    # Cálculos
    total_pac = sum(i.get('valorTotal', 0) for i in pac_items)
    total_pac_geral = sum(i.get('valorTotal', 0) for i in pac_geral_items)
    total_obras = sum(i.get('valorTotal', 0) for i in pac_obras_items)
    total_mrosc = sum(p.get('valor_total', 0) for p in projetos_mrosc)
    
    buffer = BytesIO()
    
    # Criar callback para cabeçalho/rodapé oficial
    from utils.pdf_utils import create_page_callback
    oficial_callback = create_page_callback(
        titulo_documento='RELATÓRIO GERENCIAL CONSOLIDADO',
        subtitulo=f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}',
        total_pages=1
    )
    
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=35*mm, bottomMargin=25*mm)
    
    elements = []
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('Titulo', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1F4E78'), alignment=TA_CENTER, spaceAfter=6*mm)
    subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#2E7D32'), spaceBefore=6*mm, spaceAfter=3*mm)
    
    # Resumo Executivo
    elements.append(Paragraph('1. RESUMO EXECUTIVO', subtitulo_style))
    resumo_data = [
        ['Indicador', 'Quantidade', 'Valor Total'],
        ['PACs Individuais', str(len(pacs)), f"R$ {total_pac:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['PACs Gerais', str(len(pacs_geral)), f"R$ {total_pac_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['PACs Obras', str(len(pacs_obras)), f"R$ {total_obras:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['Processos', str(len(processos)), '-'],
        ['Projetos MROSC', str(len(projetos_mrosc)), f"R$ {total_mrosc:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')],
        ['TOTAL GERAL', '-', f"R$ {(total_pac + total_pac_geral + total_obras + total_mrosc):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')]
    ]
    
    resumo_table = Table(resumo_data, colWidths=[80*mm, 40*mm, 50*mm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E3F2FD')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(resumo_table)
    
    # Processos por Status
    elements.append(Paragraph('2. PROCESSOS POR STATUS', subtitulo_style))
    status_count = {}
    for p in processos:
        s = p.get('status', 'Não Definido')
        status_count[s] = status_count.get(s, 0) + 1
    
    status_data = [['Status', 'Quantidade']]
    for s, c in sorted(status_count.items(), key=lambda x: x[1], reverse=True):
        status_data.append([s, str(c)])
    
    status_table = Table(status_data, colWidths=[100*mm, 40*mm])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C8E6C9')),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(status_table)
    
    # Assinatura
    elements.append(Spacer(1, 15*mm))
    user_name = user.name if hasattr(user, 'name') else user.get('name', 'Admin')
    elements.append(Paragraph('_' * 50, ParagraphStyle('Linha', alignment=TA_CENTER)))
    elements.append(Paragraph(f'<b>{user_name}</b>', ParagraphStyle('Assinatura', alignment=TA_CENTER, fontSize=10)))
    elements.append(Paragraph('Responsável pelo Relatório', ParagraphStyle('Cargo', alignment=TA_CENTER, fontSize=8, textColor=colors.gray)))
    
    # Build com callback oficial
    doc.build(elements, onFirstPage=oficial_callback, onLaterPages=oficial_callback)
    buffer.seek(0)
    
    filename = f"Relatorio_Consolidado_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buffer, media_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# Registrar routers
app.include_router(analytics_router)
app.include_router(alertas_router)
app.include_router(relatorios_router)

# ===== Router de Validação de Documentos (Público) =====
validation_router = APIRouter(prefix="/api/validar", tags=["Validação de Documentos"])

class DocumentValidationRequest(BaseModel):
    validation_code: str

class DocumentValidationResponse(BaseModel):
    is_valid: bool
    message: str
    document_info: Optional[dict] = None

@validation_router.get("/")
async def validation_info():
    """Retorna informações sobre a validação de documentos"""
    return {
        "titulo": "Validação de Documentos Assinados Digitalmente",
        "descricao": "Use este serviço para verificar a autenticidade de documentos emitidos pela Prefeitura Municipal de Acaiaca.",
        "instrucoes": [
            "Digite o código de validação presente no selo de assinatura do documento",
            "O código possui o formato: DOC-XXXXXXXX-YYYYMMDD",
            "Você também pode escanear o QR Code presente no documento"
        ]
    }

@validation_router.get("/{validation_code}")
async def validate_document_get(validation_code: str):
    """Valida um documento pelo código (via QR Code ou link direto)"""
    return await validate_document_internal(validation_code)

@validation_router.post("/verificar")
async def validate_document_post(request: DocumentValidationRequest):
    """Valida um documento pelo código (via formulário)"""
    return await validate_document_internal(request.validation_code)

async def validate_document_internal(validation_code: str) -> DocumentValidationResponse:
    """Função interna para validação de documentos"""
    try:
        # Buscar assinatura no banco
        signature = await db.document_signatures.find_one(
            {'validation_code': validation_code.strip().upper()},
            {'_id': 0}
        )
        
        if not signature:
            return DocumentValidationResponse(
                is_valid=False,
                message="Código de validação não encontrado. Verifique se o código foi digitado corretamente.",
                document_info=None
            )
        
        if not signature.get('is_valid', True):
            return DocumentValidationResponse(
                is_valid=False,
                message="Este documento foi revogado ou invalidado.",
                document_info=None
            )
        
        # Preparar informações do documento
        signers_info = []
        for signer in signature.get('signers', []):
            signers_info.append({
                'nome': signer.get('nome', 'N/A'),
                'cargo': signer.get('cargo', ''),
                'cpf_masked': mask_cpf(signer.get('cpf', '')),
                'email': signer.get('email', '')[:3] + '***@***' if signer.get('email') else ''
            })
        
        return DocumentValidationResponse(
            is_valid=True,
            message="Documento válido! A assinatura digital foi verificada com sucesso.",
            document_info={
                'tipo_documento': signature.get('document_type', 'Documento'),
                'data_assinatura': signature.get('created_at').isoformat() if signature.get('created_at') else None,
                'assinantes': signers_info,
                'hash_parcial': signature.get('hash_document', '')[:16] + '...'
            }
        )
    except Exception as e:
        logging.error(f"Erro ao validar documento: {e}")
        return DocumentValidationResponse(
            is_valid=False,
            message="Erro ao processar a validação. Tente novamente.",
            document_info=None
        )

app.include_router(validation_router)

# Registrar router WebSocket
app.include_router(ws_router, prefix="/api")

app.include_router(api_router)
app.include_router(public_router)  # Rotas públicas para transparência

# CORS Middleware
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','), allow_methods=["*"], allow_headers=["*"])

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Middleware de logging de requisições
from starlette.middleware.base import BaseHTTPMiddleware
import time

class LoggingMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        start_time = time.time()
        
        # Extrair user_id do token se presente
        user_id = None
        auth_header = request.headers.get('authorization', '')
        if auth_header.startswith('Bearer '):
            try:
                token = auth_header.split(' ')[1]
                payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                user_id = payload.get('user_id')
            except:
                pass
        
        response = await call_next(request)
        
        duration_ms = (time.time() - start_time) * 1000
        
        # Logar apenas se não for health check ou assets estáticos
        path = request.url.path
        if not path.startswith('/api/docs') and not path.startswith('/api/openapi') and not path.startswith('/api/redoc'):
            request_logger.log_request(
                method=request.method,
                path=path,
                status_code=response.status_code,
                duration_ms=duration_ms,
                user_id=user_id
            )
        
        return response

app.add_middleware(LoggingMiddleware)

log_info("🚀 Planejamento Acaiaca API iniciada com sucesso!")

@app.on_event("startup")
async def startup_event():
    """Evento executado na inicialização do servidor"""
    log_info("Conectando ao MongoDB...")
    log_info(f"Banco de dados: {os.environ.get('DB_NAME', 'pac_acaiaca')}")
    log_info("Sistema pronto para receber requisições")

@app.on_event("shutdown")
async def shutdown_db_client():
    """Evento executado no encerramento do servidor"""
    log_info("Encerrando conexões...")
    client.close()
    log_info("Sistema encerrado com sucesso")
