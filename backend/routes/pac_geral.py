"""
PAC Geral Routes - Planejamento Acaiaca
Handles CRUD operations for PAC Geral (General Annual Contracting Plans)
"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from io import BytesIO
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER

router = APIRouter(prefix="/api", tags=["PAC Geral"])


def setup_pac_geral_routes(db, get_current_user, PACGeral, PACGeralCreate, PACGeralUpdate, PACGeralItem, PACGeralItemCreate, PACGeralItemUpdate):
    """Configure PAC Geral routes with injected dependencies"""

    SECRETARIAS_SIGLAS = {
        'AD': 'Administração',
        'FA': 'Fazenda',
        'SA': 'Saúde',
        'SE': 'Educação',
        'AS': 'Assistência Social',
        'AG': 'Agricultura',
        'OB': 'Obras',
        'TR': 'Transporte',
        'CUL': 'Cultura'
    }

    # ===== STATS =====
    @router.get("/pacs-geral/stats")
    async def get_pacs_geral_stats(request: Request, ano: str = None):
        """Returns PAC Geral statistics"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        
        pacs = await db.pacs_geral.find(query, {'_id': 0}).to_list(1000)
        
        stats = {
            'total_pacs': len(pacs),
            'valor_total': 0,
            'total_items': 0,
            'por_secretaria': {}
        }
        
        for pac in pacs:
            items = await db.pac_geral_items.find({'pac_geral_id': pac['pac_geral_id']}, {'_id': 0}).to_list(10000)
            valor_pac = sum(item.get('valorTotal', 0) for item in items)
            stats['valor_total'] += valor_pac
            stats['total_items'] += len(items)
            
            nome_sec = pac.get('nome_secretaria', 'Não definida')
            if nome_sec not in stats['por_secretaria']:
                stats['por_secretaria'][nome_sec] = {'quantidade': 0, 'valor': 0}
            stats['por_secretaria'][nome_sec]['quantidade'] += 1
            stats['por_secretaria'][nome_sec]['valor'] += valor_pac
        
        return stats

    # ===== YEARS =====
    @router.get("/pacs-geral/anos")
    async def get_anos_pacs_geral(request: Request):
        """Returns available years in PACs Geral"""
        user = await get_current_user(request)
        pipeline = [
            {"$group": {"_id": "$ano"}},
            {"$sort": {"_id": -1}}
        ]
        anos = await db.pacs_geral.aggregate(pipeline).to_list(100)
        anos_list = [a['_id'] for a in anos if a['_id']]
        if not anos_list:
            anos_list = [str(datetime.now().year)]
        return anos_list

    # ===== CRUD =====
    @router.get("/pacs-geral", response_model=List[PACGeral])
    async def get_pacs_geral(request: Request, ano: str = None):
        """List all PACs Geral"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        pacs = await db.pacs_geral.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
        return [PACGeral(**p) for p in pacs]

    @router.get("/pacs-geral/paginado")
    async def get_pacs_geral_paginado(
        request: Request,
        ano: str = None,
        secretaria: str = None,
        page: int = 1,
        limit: int = 20
    ):
        """List PACs Geral with pagination"""
        user = await get_current_user(request)
        query = {}
        if ano:
            query['ano'] = ano
        if secretaria:
            query['nome_secretaria'] = {'$regex': secretaria, '$options': 'i'}
        
        skip = (page - 1) * limit
        total = await db.pacs_geral.count_documents(query)
        pacs = await db.pacs_geral.find(query, {'_id': 0}).sort('created_at', -1).skip(skip).limit(limit).to_list(limit)
        
        return {
            'data': pacs,
            'total': total,
            'page': page,
            'limit': limit,
            'pages': (total + limit - 1) // limit
        }

    @router.post("/pacs-geral", response_model=PACGeral)
    async def create_pac_geral(pac: PACGeralCreate, request: Request):
        """Creates a new PAC Geral"""
        user = await get_current_user(request)
        pac_dict = pac.model_dump()
        pac_dict['pac_geral_id'] = f"pacg_{uuid.uuid4().hex[:12]}"
        pac_dict['user_id'] = user.user_id
        pac_dict['created_at'] = datetime.now(timezone.utc)
        pac_dict['updated_at'] = datetime.now(timezone.utc)
        await db.pacs_geral.insert_one(pac_dict)
        pac_dict.pop('_id', None)
        return PACGeral(**pac_dict)

    @router.get("/pacs-geral/{pac_geral_id}", response_model=PACGeral)
    async def get_pac_geral(pac_geral_id: str, request: Request):
        """Gets a specific PAC Geral"""
        user = await get_current_user(request)
        pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        return PACGeral(**pac)

    @router.put("/pacs-geral/{pac_geral_id}", response_model=PACGeral)
    async def update_pac_geral(pac_geral_id: str, pac: PACGeralUpdate, request: Request):
        """Updates a PAC Geral"""
        user = await get_current_user(request)
        existing = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id})
        if not existing:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        
        update_data = {k: v for k, v in pac.model_dump().items() if v is not None}
        update_data['updated_at'] = datetime.now(timezone.utc)
        
        await db.pacs_geral.update_one({'pac_geral_id': pac_geral_id}, {'$set': update_data})
        updated = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
        return PACGeral(**updated)

    @router.delete("/pacs-geral/{pac_geral_id}")
    async def delete_pac_geral(pac_geral_id: str, request: Request):
        """Deletes a PAC Geral and its items"""
        user = await get_current_user(request)
        result = await db.pacs_geral.delete_one({'pac_geral_id': pac_geral_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        await db.pac_geral_items.delete_many({'pac_geral_id': pac_geral_id})
        return {'message': 'PAC Geral excluído com sucesso'}

    # ===== ITEMS =====
    @router.get("/pacs-geral/{pac_geral_id}/items", response_model=List[PACGeralItem])
    async def get_pac_geral_items(pac_geral_id: str, request: Request):
        """Lists items of a PAC Geral"""
        user = await get_current_user(request)
        items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(10000)
        return [PACGeralItem(**item) for item in items]

    @router.post("/pacs-geral/{pac_geral_id}/items", response_model=PACGeralItem)
    async def create_pac_geral_item(pac_geral_id: str, item: PACGeralItemCreate, request: Request):
        """Adds an item to a PAC Geral"""
        user = await get_current_user(request)
        pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        
        item_dict = item.model_dump()
        item_dict['item_id'] = f"itemg_{uuid.uuid4().hex[:12]}"
        item_dict['pac_geral_id'] = pac_geral_id
        
        # Calculate total quantity
        quantidade_total = (
            item_dict.get('qtd_ad', 0) + item_dict.get('qtd_fa', 0) +
            item_dict.get('qtd_sa', 0) + item_dict.get('qtd_se', 0) +
            item_dict.get('qtd_as', 0) + item_dict.get('qtd_ag', 0) +
            item_dict.get('qtd_ob', 0) + item_dict.get('qtd_tr', 0) +
            item_dict.get('qtd_cul', 0)
        )
        item_dict['quantidade_total'] = quantidade_total
        item_dict['valorTotal'] = round(quantidade_total * item_dict.get('valorUnitario', 0), 2)
        item_dict['created_at'] = datetime.now(timezone.utc)
        
        await db.pac_geral_items.insert_one(item_dict)
        item_dict.pop('_id', None)
        return PACGeralItem(**item_dict)

    @router.put("/pacs-geral/{pac_geral_id}/items/{item_id}", response_model=PACGeralItem)
    async def update_pac_geral_item(pac_geral_id: str, item_id: str, item: PACGeralItemUpdate, request: Request):
        """Updates a PAC Geral item"""
        user = await get_current_user(request)
        existing = await db.pac_geral_items.find_one({'item_id': item_id, 'pac_geral_id': pac_geral_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        
        update_data = {k: v for k, v in item.model_dump().items() if v is not None}
        
        # Recalculate totals
        qtd_fields = ['qtd_ad', 'qtd_fa', 'qtd_sa', 'qtd_se', 'qtd_as', 'qtd_ag', 'qtd_ob', 'qtd_tr', 'qtd_cul']
        quantidade_total = sum(update_data.get(f, existing.get(f, 0)) for f in qtd_fields)
        update_data['quantidade_total'] = quantidade_total
        
        valor_unitario = update_data.get('valorUnitario', existing.get('valorUnitario', 0))
        update_data['valorTotal'] = round(quantidade_total * valor_unitario, 2)
        
        await db.pac_geral_items.update_one({'item_id': item_id}, {'$set': update_data})
        updated = await db.pac_geral_items.find_one({'item_id': item_id}, {'_id': 0})
        return PACGeralItem(**updated)

    @router.delete("/pacs-geral/{pac_geral_id}/items/{item_id}")
    async def delete_pac_geral_item(pac_geral_id: str, item_id: str, request: Request):
        """Deletes a PAC Geral item"""
        user = await get_current_user(request)
        result = await db.pac_geral_items.delete_one({'item_id': item_id, 'pac_geral_id': pac_geral_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        return {'message': 'Item excluído com sucesso'}

    # ===== EXPORT XLSX =====
    @router.get("/pacs-geral/{pac_geral_id}/export/xlsx")
    async def export_pac_geral_xlsx(pac_geral_id: str, request: Request):
        """Exports PAC Geral to XLSX"""
        user = await get_current_user(request)
        
        pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        
        items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(10000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "PAC Geral"
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:R1')
        ws['A1'] = f"PAC GERAL - {pac.get('nome_secretaria', 'N/A')} - {pac.get('ano', '2026')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['#', 'CATMAT', 'Descrição', 'Unid.', 'AD', 'FA', 'SA', 'SE', 'AS', 'AG', 'OB', 'TR', 'CUL', 'Total', 'V.Unit', 'V.Total', 'Prior.', 'Class.']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        total_geral = 0
        for row, item in enumerate(items, 5):
            ws.cell(row=row, column=1, value=row-4).border = thin_border
            ws.cell(row=row, column=2, value=item.get('catmat', '')).border = thin_border
            ws.cell(row=row, column=3, value=item.get('descricao', '')).border = thin_border
            ws.cell(row=row, column=4, value=item.get('unidade', '')).border = thin_border
            ws.cell(row=row, column=5, value=item.get('qtd_ad', 0)).border = thin_border
            ws.cell(row=row, column=6, value=item.get('qtd_fa', 0)).border = thin_border
            ws.cell(row=row, column=7, value=item.get('qtd_sa', 0)).border = thin_border
            ws.cell(row=row, column=8, value=item.get('qtd_se', 0)).border = thin_border
            ws.cell(row=row, column=9, value=item.get('qtd_as', 0)).border = thin_border
            ws.cell(row=row, column=10, value=item.get('qtd_ag', 0)).border = thin_border
            ws.cell(row=row, column=11, value=item.get('qtd_ob', 0)).border = thin_border
            ws.cell(row=row, column=12, value=item.get('qtd_tr', 0)).border = thin_border
            ws.cell(row=row, column=13, value=item.get('qtd_cul', 0)).border = thin_border
            ws.cell(row=row, column=14, value=item.get('quantidade_total', 0)).border = thin_border
            ws.cell(row=row, column=15, value=item.get('valorUnitario', 0)).border = thin_border
            ws.cell(row=row, column=16, value=item.get('valorTotal', 0)).border = thin_border
            ws.cell(row=row, column=17, value=item.get('prioridade', '')).border = thin_border
            ws.cell(row=row, column=18, value=item.get('codigo_classificacao', '')).border = thin_border
            total_geral += item.get('valorTotal', 0)
        
        # Total row
        total_row = 5 + len(items)
        ws.cell(row=total_row, column=15, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=total_row, column=16, value=total_geral).font = Font(bold=True)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"PAC_Geral_{pac.get('nome_secretaria', 'geral')}_{pac.get('ano', '2026')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # ===== IMPORT =====
    @router.post("/pacs-geral/{pac_geral_id}/import")
    async def import_pac_geral_items(pac_geral_id: str, request: Request, file: UploadFile = File(...)):
        """Imports items from XLSX file"""
        user = await get_current_user(request)
        
        pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        
        try:
            contents = await file.read()
            wb = load_workbook(filename=BytesIO(contents))
            ws = wb.active
            
            imported = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                try:
                    quantidade_total = sum([
                        float(row[4] or 0), float(row[5] or 0), float(row[6] or 0),
                        float(row[7] or 0), float(row[8] or 0), float(row[9] or 0),
                        float(row[10] or 0), float(row[11] or 0), float(row[12] or 0)
                    ])
                    valor_unitario = float(row[13] or 0)
                    
                    item_dict = {
                        'item_id': f"itemg_{uuid.uuid4().hex[:12]}",
                        'pac_geral_id': pac_geral_id,
                        'catmat': str(row[0]) if row[0] else '',
                        'descricao': str(row[1]) if row[1] else '',
                        'unidade': str(row[2]) if row[2] else 'UN',
                        'qtd_ad': float(row[4] or 0),
                        'qtd_fa': float(row[5] or 0),
                        'qtd_sa': float(row[6] or 0),
                        'qtd_se': float(row[7] or 0),
                        'qtd_as': float(row[8] or 0),
                        'qtd_ag': float(row[9] or 0),
                        'qtd_ob': float(row[10] or 0),
                        'qtd_tr': float(row[11] or 0),
                        'qtd_cul': float(row[12] or 0),
                        'quantidade_total': quantidade_total,
                        'valorUnitario': valor_unitario,
                        'valorTotal': round(quantidade_total * valor_unitario, 2),
                        'prioridade': str(row[14]) if row[14] else 'Média',
                        'justificativa': str(row[15]) if len(row) > 15 and row[15] else '',
                        'created_at': datetime.now(timezone.utc)
                    }
                    await db.pac_geral_items.insert_one(item_dict)
                    imported += 1
                except Exception as e:
                    errors.append(f"Linha: {str(e)}")
            
            return {'success': True, 'imported': imported, 'errors': errors}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")

    # ===== EXPORT PDF =====
    @router.get("/pacs-geral/{pac_geral_id}/export/pdf")
    async def export_pac_geral_pdf(pac_geral_id: str, request: Request):
        """Exports PAC Geral to PDF"""
        user = await get_current_user(request)
        
        pac = await db.pacs_geral.find_one({'pac_geral_id': pac_geral_id}, {'_id': 0})
        if not pac:
            raise HTTPException(status_code=404, detail="PAC Geral não encontrado")
        
        items = await db.pac_geral_items.find({'pac_geral_id': pac_geral_id}, {'_id': 0}).to_list(10000)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=15*mm, bottomMargin=10*mm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=12, textColor=colors.HexColor('#1F4E78'), alignment=TA_CENTER, spaceAfter=8)
        
        elements.append(Paragraph(f"PAC GERAL - {pac.get('nome_secretaria', 'N/A')} - {pac.get('ano', '2026')}", title_style))
        elements.append(Spacer(1, 5))
        
        # Table with secretaria columns
        table_data = [['#', 'CATMAT', 'Descrição', 'AD', 'FA', 'SA', 'SE', 'AS', 'AG', 'OB', 'TR', 'CUL', 'Total', 'V.Total']]
        
        total = 0
        for i, item in enumerate(items, 1):
            valor_total = item.get('valorTotal', 0)
            total += valor_total
            table_data.append([
                str(i),
                item.get('catmat', '')[:8],
                item.get('descricao', '')[:30],
                str(int(item.get('qtd_ad', 0))),
                str(int(item.get('qtd_fa', 0))),
                str(int(item.get('qtd_sa', 0))),
                str(int(item.get('qtd_se', 0))),
                str(int(item.get('qtd_as', 0))),
                str(int(item.get('qtd_ag', 0))),
                str(int(item.get('qtd_ob', 0))),
                str(int(item.get('qtd_tr', 0))),
                str(int(item.get('qtd_cul', 0))),
                str(int(item.get('quantidade_total', 0))),
                f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ])
        
        table_data.append(['', '', '', '', '', '', '', '', '', '', '', '', 'TOTAL:', f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')])
        
        table = Table(table_data, colWidths=[18, 35, 100, 22, 22, 22, 22, 22, 22, 22, 22, 22, 30, 55])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"PAC_Geral_{pac.get('nome_secretaria', 'geral')}_{pac.get('ano', '2026')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    return router
